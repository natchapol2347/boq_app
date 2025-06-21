#!/usr/bin/env python3
# Hard-coded BOQ Processor
# This version uses fixed column mappings for each sheet type
# and separate database tables for each sheet type

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import numpy as np
from fuzzywuzzy import fuzz
import os
import uuid
from datetime import datetime
import re
from werkzeug.utils import secure_filename
from pathlib import Path
import sqlite3
import logging
import shutil
import openpyxl

logging.basicConfig(level=logging.INFO)

class HardCodedBOQProcessor:
    def __init__(self):
        self.app = Flask(__name__)
        CORS(self.app)
        
        # --- Database and Folder Setup ---
        self.data_dir = Path.home() / 'AppData' / 'Roaming' / 'BOQProcessor'
        os.makedirs(self.data_dir, exist_ok=True)
        self.db_path = self.data_dir / 'master_data.db'
        
        # Session management for processing sessions
        self.processing_sessions = {}
        
        self.master_data_folder = 'master_data'
        self.upload_folder = 'uploads'
        self.output_folder = 'output'
        os.makedirs(self.master_data_folder, exist_ok=True)
        os.makedirs(self.upload_folder, exist_ok=True)
        os.makedirs(self.output_folder, exist_ok=True)

        # Hardcoded markup rates
        self.markup_rates = {100: 1.00, 130: 1.30, 150: 1.50, 50: 0.50, 30: 0.30}
        
        # Sheet type definitions
        self.sheet_types = {
            'INT': {
                'pattern': 'int',
                'header_row': 9,  # 0-based
                'columns': {
                    'code': 2,          # Column B
                    'name': 3,          # Column C
                    'quantity': 4,      # Column D
                    'unit': 5,          # Column E
                    'material_cost': 6, # Column F
                    'labor_cost': 7,    # Column G
                    'total_cost': 8     # Column H
                },
                'table_name': 'interior_items'
            },
            'EE': {
                'pattern': 'ee',
                'header_row': 7,  # 0-based
                'columns': {
                    'code': 2,          # Column B
                    'name': 3,          # Column C
                    'unit': 6,          # Column F
                    'quantity': 7,      # Column G
                    'material_cost': 8, # Column H
                    'labor_cost': 10,   # Column J
                    'total_cost': 12    # Column L
                },
                'table_name': 'ee_items'
            },
            'AC': {
                'pattern': 'ac',
                'header_row': 5,  # 0-based
                'columns': {
                    'code': 2,          # Column B
                    'name': 3,          # Column C
                    'unit': 6,          # Column F
                    'quantity': 7,      # Column G
                    'material_cost': 8, # Column H
                    'labor_cost': 10,   # Column J
                    'total_cost': 12    # Column L
                },
                'table_name': 'ac_items'
            },
            'FP': {
                'pattern': 'fp',
                'header_row': 7,  # 0-based
                'columns': {
                    'code': 2,          # Column B
                    'name': 3,          # Column C
                    'unit': 6,          # Column F
                    'quantity': 7,      # Column G
                    'material_cost': 8, # Column H
                    'labor_cost': 10,   # Column J
                    'total_cost': 12    # Column L
                },
                'table_name': 'fp_items'
            },
            'DEFAULT': {
                'pattern': '',
                'header_row': 8,  # 0-based
                'columns': {
                    'code': 2,          # Column B
                    'name': 3,          # Column C
                    'quantity': 4,      # Column D
                    'unit': 5,          # Column E
                    'material_cost': 6, # Column F
                    'labor_cost': 7,    # Column G
                    'total_cost': 8     # Column H
                },
                'table_name': 'default_items'
            }
        }
        
        self._init_db()
        self._sync_default_master_excel()
        self.setup_routes()

    def _init_db(self):
        """Initialize SQLite database with separate tables for each sheet type"""
        logging.info(f"Initializing database at {self.db_path}")
        
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Create a table for each sheet type
                for sheet_type, config in self.sheet_types.items():
                    table_name = config['table_name']
                    cursor.execute(f'''
                        CREATE TABLE IF NOT EXISTS {table_name} (
                            internal_id TEXT PRIMARY KEY, 
                            code TEXT, 
                            name TEXT NOT NULL,
                            material_cost REAL DEFAULT 0, 
                            labor_cost REAL DEFAULT 0, 
                            total_cost REAL DEFAULT 0,
                            unit TEXT
                        )
                    ''')
                conn.commit()
                
                # Debug: Check tables
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                tables = cursor.fetchall()
                logging.info(f"Database tables: {tables}")
                
                for sheet_type, config in self.sheet_types.items():
                    table_name = config['table_name']
                    # Check if table has any data
                    count = cursor.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
                    logging.info(f"Table {table_name} contains {count} items")
                    
                    # Check items with costs
                    if count > 0:
                        items_with_costs = cursor.execute(
                            f"SELECT COUNT(*) FROM {table_name} WHERE material_cost > 0 OR labor_cost > 0"
                        ).fetchone()[0]
                        logging.info(f"Items with costs in {table_name}: {items_with_costs} ({items_with_costs/count*100:.1f}%)")
                
                logging.info(f"Database initialized successfully at {self.db_path}")
                
        except Exception as e:
            logging.error(f"Error initializing database: {e}", exc_info=True)
            raise

    def _sync_default_master_excel(self):
        """Sync default master Excel file to database with sheet-specific tables"""
        default_excel_path = os.path.join(self.master_data_folder, 'master.xlsx')
        
        if os.path.exists(default_excel_path):
            logging.info(f"Found {default_excel_path}. Synchronizing database...")
            
            try:
                excel_file = pd.ExcelFile(default_excel_path)
                sheet_names = excel_file.sheet_names
                logging.info(f"Excel file contains {len(sheet_names)} sheets: {sheet_names}")
                
                # Process each sheet
                for sheet_name in sheet_names:
                    if "sum" in sheet_name.lower():
                        continue
                        
                    # Determine sheet type
                    sheet_type = 'DEFAULT'
                    for key, config in self.sheet_types.items():
                        if config['pattern'] and config['pattern'].lower() in sheet_name.lower():
                            sheet_type = key
                            break
                    
                    logging.info(f"Processing sheet {sheet_name} as type {sheet_type}")
                    
                    # Get the sheet configuration
                    config = self.sheet_types[sheet_type]
                    header_row = config['header_row']
                    table_name = config['table_name']
                    
                    # Read sheet with fixed header row
                    df = pd.read_excel(default_excel_path, sheet_name=sheet_name, header=header_row)
                    
                    # Process the dataframe using fixed column positions
                    processed_df = self._process_master_sheet(df, sheet_type)
                    
                    if not processed_df.empty:
                        # Insert into the appropriate table
                        with sqlite3.connect(self.db_path) as conn:
                            # Clear existing data for this sheet type
                            conn.execute(f"DELETE FROM {table_name}")
                            
                            # Insert new data
                            for _, row in processed_df.iterrows():
                                conn.execute(
                                    f"INSERT INTO {table_name} (internal_id, code, name, material_cost, labor_cost, total_cost, unit) VALUES (?, ?, ?, ?, ?, ?, ?)",
                                    (
                                        row['internal_id'],
                                        row['code'],
                                        row['name'],
                                        row['material_cost'],
                                        row['labor_cost'],
                                        row['total_cost'],
                                        row.get('unit', '')
                                    )
                                )
                            conn.commit()
                            
                        logging.info(f"Synchronized {len(processed_df)} items into {table_name}")
                
                # Add sample costs if no costs are found
                self._ensure_costs_exist()
                        
            except Exception as e:
                logging.error(f"Error synchronizing master data: {e}", exc_info=True)
        else:
            logging.warning(f"Default master Excel file not found at {default_excel_path}")
            
            # Add sample data if no master data is available
            self._add_sample_data()

    def _process_master_sheet(self, df, sheet_type):
        """Process a master data sheet using fixed column positions"""
        if df.empty:
            return pd.DataFrame()
        
        # Get column configuration for this sheet type
        config = self.sheet_types[sheet_type]
        column_mapping = config['columns']
        
        # Create a new dataframe with selected columns
        result_data = []
        for idx, row in df.iterrows():
            try:
                # Get values from fixed positions
                code_idx = column_mapping['code'] - 1  # Convert to 0-based
                name_idx = column_mapping['name'] - 1
                material_idx = column_mapping['material_cost'] - 1
                labor_idx = column_mapping['labor_cost'] - 1
                unit_idx = column_mapping['unit'] - 1 if 'unit' in column_mapping else None
                
                # Get values (safely)
                if idx >= len(df):
                    continue
                    
                row_values = row.values
                if len(row_values) <= max(code_idx, name_idx, material_idx, labor_idx):
                    continue
                
                code = str(row_values[code_idx]) if code_idx < len(row_values) else ''
                name = str(row_values[name_idx]) if name_idx < len(row_values) else ''
                
                # Clean and convert cost values
                try:
                    material_cost = float(row_values[material_idx]) if material_idx < len(row_values) and pd.notna(row_values[material_idx]) else 0
                except (ValueError, TypeError):
                    material_cost = 0
                    
                try:
                    labor_cost = float(row_values[labor_idx]) if labor_idx < len(row_values) and pd.notna(row_values[labor_idx]) else 0
                except (ValueError, TypeError):
                    labor_cost = 0
                
                # Get unit if available
                unit = str(row_values[unit_idx]) if unit_idx is not None and unit_idx < len(row_values) else ''
                
                # Skip empty or total rows
                if not name or name.lower() in ['nan', 'none', '']:
                    continue
                    
                if any(keyword in name.lower() for keyword in ['total', 'รวม', 'sum', 'subtotal']):
                    continue
                
                # Create item
                total_cost = material_cost + labor_cost
                item = {
                    'internal_id': f"item_{uuid.uuid4().hex[:8]}",
                    'code': code,
                    'name': name,
                    'material_cost': material_cost,
                    'labor_cost': labor_cost,
                    'total_cost': total_cost,
                    'unit': unit
                }
                
                result_data.append(item)
                
                # Debug costs
                if material_cost > 0 or labor_cost > 0:
                    print(f"Found item with costs: {name[:30]}... -> Material: {material_cost}, Labor: {labor_cost}")
                    
            except Exception as e:
                print(f"Error processing row {idx}: {e}")
                continue
        
        if not result_data:
            return pd.DataFrame()
            
        result_df = pd.DataFrame(result_data)
        print(f"Processed {len(result_df)} items from sheet type {sheet_type}")
        print(f"Items with costs: {len(result_df[(result_df['material_cost'] > 0) | (result_df['labor_cost'] > 0)])}")
        
        return result_df

    def _ensure_costs_exist(self):
        """Make sure all tables have at least some items with costs"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            for sheet_type, config in self.sheet_types.items():
                table_name = config['table_name']
                
                # Check if table has costs
                cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE material_cost > 0 OR labor_cost > 0")
                count = cursor.fetchone()[0]
                
                if count == 0:
                    logging.info(f"No costs found in {table_name}, adding sample costs")
                    
                    # Add sample costs to this table
                    cursor.execute(f"UPDATE {table_name} SET material_cost = 500, labor_cost = 300, total_cost = 800")
                    conn.commit()
                    
                    cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE material_cost > 0")
                    updated = cursor.fetchone()[0]
                    logging.info(f"Added sample costs to {updated} items in {table_name}")

    def _add_sample_data(self):
        """Add sample data to the database if no master data is available"""
        logging.info("Adding sample data to database")
        
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Add sample items to each table
            for sheet_type, config in self.sheet_types.items():
                table_name = config['table_name']
                
                # Check if table already has data
                cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                count = cursor.fetchone()[0]
                
                if count > 0:
                    logging.info(f"Table {table_name} already has {count} items, skipping")
                    continue
                
                # Add 10 sample items
                for i in range(1, 11):
                    item_id = f"sample_{sheet_type.lower()}_{i}"
                    name = f"Sample {sheet_type} item {i}"
                    code = f"CODE{sheet_type}{i}"
                    material_cost = 500
                    labor_cost = 300
                    total_cost = material_cost + labor_cost
                    
                    cursor.execute(
                        f"INSERT INTO {table_name} (internal_id, code, name, material_cost, labor_cost, total_cost, unit) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (item_id, code, name, material_cost, labor_cost, total_cost, 'EA')
                    )
                
                conn.commit()
                logging.info(f"Added 10 sample items to {table_name}")

    def store_processing_session(self, session_id, data):
        """Store processing session data"""
        self.processing_sessions[session_id] = { 
            'data': data, 
            'created_at': datetime.now() 
        }

    def find_best_match(self, item_name, sheet_type):
        """Find best matching item from database using fuzzy matching - specific to sheet type"""
        if not item_name or pd.isna(item_name): 
            return None
            
        # Get the appropriate table for this sheet type
        table_name = self.sheet_types[sheet_type]['table_name']
            
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            all_items = conn.execute(f"SELECT * FROM {table_name}").fetchall()
            
        if not all_items: 
            return None
        
        sanitized_search = str(item_name).lower().strip()
        best_match = None
        best_similarity = 0
        
        for item_row in all_items:
            item_dict = dict(item_row)
            sanitized_candidate = str(item_dict['name']).lower().strip()
            similarity = fuzz.ratio(sanitized_search, sanitized_candidate)
            
            if similarity > best_similarity:
                best_similarity = similarity
                best_match = {'item': item_dict, 'similarity': similarity}
                
        return best_match

    def determine_sheet_type(self, sheet_name):
        """Determine the sheet type based on its name"""
        sheet_name_lower = sheet_name.lower()
        
        for sheet_type, config in self.sheet_types.items():
            if config['pattern'] and config['pattern'].lower() in sheet_name_lower:
                return sheet_type
                
        # Default to the DEFAULT type if no match
        return 'DEFAULT'

    def _safe_write_to_cell(self, worksheet, row_num, col_num, value):
        """Write value to Excel cell with error handling"""
        if row_num is None or col_num is None or row_num < 1 or col_num < 1:
            print(f"❌ Invalid cell coordinates: row={row_num}, col={col_num}")
            return False
            
        try:
            # Debug cell info
            print(f"  🔍 Writing to cell ({row_num}, {col_num}): value={value}, type={type(value)}")
            
            cell = worksheet.cell(row=int(row_num), column=int(col_num))
            
            # Handle merged cells
            if hasattr(cell, 'coordinate'):
                for merged_range in worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Write to the top-left cell of merged range
                        print(f"  🔄 Cell is part of merged range: {merged_range}")
                        main_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        main_cell.value = value
                        return True
            
            # Write to regular cell
            cell.value = value
            
            # Ensure proper number formatting for numeric values
            if isinstance(value, (int, float)):
                cell.number_format = '0.00'
            
            return True
            
        except Exception as e:
            print(f"❌ Error writing to cell ({row_num}, {col_num}): {e}")
            return False

    def setup_routes(self):
        @self.app.route('/api/process-boq', methods=['POST'])
        def process_boq_route():
            if 'file' not in request.files: 
                return jsonify({'success': False, 'error': 'No file uploaded'})
                
            file = request.files['file']
            filepath = os.path.join(self.upload_folder, secure_filename(file.filename))
            file.save(filepath)
            
            try:
                excel_file = pd.ExcelFile(filepath)
                session_data = {'sheets': {}, 'original_filepath': filepath}
                
                sheets_to_process = [s for s in excel_file.sheet_names if "sum" not in s.lower()]
                
                for sheet_name in sheets_to_process:
                    # Determine sheet type
                    sheet_type = self.determine_sheet_type(sheet_name)
                    sheet_config = self.sheet_types[sheet_type]
                    
                    print(f"\nProcessing sheet: {sheet_name} as type {sheet_type}")
                    
                    # Use fixed header row from configuration
                    header_row = sheet_config['header_row']
                    fixed_columns = sheet_config['columns']
                    
                    # Read with fixed header row
                    df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row)
                    
                    # Simple processing
                    processed_items = []
                    total_rows = len(df)
                    matched_count = 0
                    
                    # Process each row in the sheet
                    for idx, row in df.iterrows():
                        try:
                            # Get name from fixed column
                            name_col = fixed_columns['name'] - 1  # Convert to 0-based
                            if name_col >= len(row):
                                continue
                                
                            name = str(row.iloc[name_col]).strip()
                            
                            # Skip empty or header rows
                            if not name or name.lower() in ['nan', 'none', ''] or any(keyword in name.lower() for keyword in ['total', 'รวม', 'system', 'ระบบ']):
                                continue
                            
                            # Find match in database
                            match = self.find_best_match(name, sheet_type)
                            
                            if match and match['similarity'] >= 50:
                                processed_items.append({
                                    'original_row_index': idx,
                                    'match': match
                                })
                                matched_count += 1
                                print(f"  Match: '{name[:40]}...' -> {match['similarity']:.0f}% similarity")
                        except Exception as e:
                            print(f"Error processing row {idx}: {e}")
                            continue
                    
                    print(f"Sheet {sheet_name}: {matched_count}/{total_rows} items matched")
                    
                    # Store sheet info
                    session_data['sheets'][sheet_name] = {
                        'sheet_type': sheet_type,
                        'header_row_num': header_row,
                        'processed_matches': {item['original_row_index']: item['match'] for item in processed_items},
                        'total_rows': total_rows,
                        'matched_count': matched_count
                    }
                
                # Store session
                session_id = str(uuid.uuid4())
                self.store_processing_session(session_id, session_data)
                
                # Calculate summary
                total_items = sum(sheet['total_rows'] for sheet in session_data['sheets'].values())
                total_matches = sum(sheet['matched_count'] for sheet in session_data['sheets'].values())
                
                return jsonify({
                    'success': True, 
                    'session_id': session_id,
                    'summary': {
                        'total_items': total_items,
                        'matched_items': total_matches,
                        'match_rate': (total_matches / total_items * 100) if total_items > 0 else 0
                    }
                })
                
            except Exception as e:
                logging.error(f"Error processing BOQ file: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})

        @self.app.route('/api/generate-final-boq', methods=['POST'])
        def generate_final_boq_route():
            data = request.get_json()
            session_id = data.get('session_id')
            if not session_id or session_id not in self.processing_sessions: 
                return jsonify({'success': False, 'error': 'Invalid session'})
            
            session_data = self.processing_sessions[session_id]['data']
            original_filepath = session_data['original_filepath']
            markup_options = data.get('markup_options', [100, 130, 150, 50, 30])
            
            try:
                # Generate unique output filename
                filename = f"final_boq_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                output_filepath = os.path.join(self.output_folder, filename)
                shutil.copy(original_filepath, output_filepath)

                # Open workbooks
                workbook = openpyxl.load_workbook(output_filepath)
                data_workbook = openpyxl.load_workbook(original_filepath, data_only=True)

                # Initialize counters
                items_processed = 0
                items_failed = 0
                items_with_zero_cost = 0
                items_with_zero_qty = 0

                # Process each sheet
                for sheet_name, sheet_info in session_data['sheets'].items():
                    if sheet_name not in workbook.sheetnames: 
                        continue
                        
                    print(f"\n📋 Processing sheet: {sheet_name}")
                    worksheet = workbook[sheet_name]
                    data_worksheet = data_workbook[sheet_name]
                    
                    # Get sheet info
                    sheet_type = sheet_info['sheet_type']
                    header_row_num = sheet_info['header_row_num']
                    
                    # Get column mappings for this sheet type
                    fixed_columns = self.sheet_types[sheet_type]['columns']
                    print(f"Using fixed columns for {sheet_type}: {fixed_columns}")
                    
                    # Add markup headers
                    header_row_excel = header_row_num + 1
                    start_markup_col = worksheet.max_column + 1
                    for i, p in enumerate(markup_options):
                        self._safe_write_to_cell(worksheet, header_row_excel, start_markup_col + i, f'Markup {p}%')

                    # Process matches
                    match_count = len(sheet_info['processed_matches'])
                    print(f"Processing {match_count} matches for sheet {sheet_name}")
                    
                    for original_row_index, match_info in sheet_info['processed_matches'].items():
                        if match_info['similarity'] < 50:
                            continue
                            
                        # Calculate target Excel row (1-based)
                        # The original_row_index is already a 0-based index from the dataframe after header
                        # So we just need to add it to the header_row_excel and add 1 to convert to 1-based Excel row
                        target_row_excel = header_row_excel + 1 + int(original_row_index)
                        
                        master_item = match_info['item']
                        
                        # Debug match details
                        print(f"\n🔄 Processing match: '{master_item.get('name', '')[:40]}...'")
                        print(f"  - Similarity: {match_info['similarity']}%")
                        
                        # Get quantity from data worksheet
                        quantity = 0
                        qty_col = fixed_columns.get('quantity')
                        if qty_col:
                            qty_cell = data_worksheet.cell(row=target_row_excel, column=qty_col)
                            try:
                                quantity = float(qty_cell.value or 0)
                                print(f"  - Quantity: {quantity}")
                            except (ValueError, TypeError) as e:
                                print(f"  ⚠️ Quantity conversion error: {e}")
                                quantity = 0
                        
                        if quantity == 0:
                            items_with_zero_qty += 1
                            print(f"  ⚠️ Zero quantity detected!")
                        
                        # Get costs from master data
                        mat_cost = float(master_item.get('material_cost') or 0)
                        lab_cost = float(master_item.get('labor_cost') or 0)
                        total_cost = mat_cost + lab_cost
                        
                        print(f"  - Material cost: {mat_cost}")
                        print(f"  - Labor cost: {lab_cost}")
                        print(f"  - Total unit cost: {total_cost}")
                        
                        if total_cost == 0:
                            items_with_zero_cost += 1
                            print(f"  ⚠️ Zero cost detected!")
                        
                        # Calculate final values
                        total_with_qty = total_cost * quantity
                        
                        # Write costs to Excel
                        success_count = 0
                        
                        # Material cost
                        mat_col = fixed_columns.get('material_cost')
                        if mat_col:
                            print(f"  - Writing material cost {mat_cost} to cell ({target_row_excel}, {mat_col})")
                            if self._safe_write_to_cell(worksheet, target_row_excel, mat_col, mat_cost):
                                success_count += 1
                        
                        # Labor cost
                        lab_col = fixed_columns.get('labor_cost')
                        if lab_col:
                            print(f"  - Writing labor cost {lab_cost} to cell ({target_row_excel}, {lab_col})")
                            if self._safe_write_to_cell(worksheet, target_row_excel, lab_col, lab_cost):
                                success_count += 1
                        
                        # Total cost
                        total_col = fixed_columns.get('total_cost')
                        if total_col:
                            print(f"  - Writing total cost {total_with_qty} to cell ({target_row_excel}, {total_col})")
                            if self._safe_write_to_cell(worksheet, target_row_excel, total_col, total_with_qty):
                                success_count += 1
                        
                        # Write markup values
                        for i, p in enumerate(markup_options):
                            rate = self.markup_rates.get(p, 1.0)
                            markup_val = round(total_cost * quantity * (1 + rate), 2)
                            col_num = start_markup_col + i
                            print(f"  - Writing markup {p}% value {markup_val} to cell ({target_row_excel}, {col_num})")
                            self._safe_write_to_cell(worksheet, target_row_excel, col_num, markup_val)
                        
                        if success_count > 0:
                            items_processed += 1
                        else:
                            items_failed += 1
                            print(f"  ❌ Failed to write any values for row {target_row_excel}")

                # Save the workbook
                print(f"\n💾 Saving workbook to {output_filepath}")
                workbook.save(output_filepath)
                workbook.close()
                data_workbook.close()
                
                # Cleanup
                if os.path.exists(original_filepath): 
                    os.remove(original_filepath)
                if session_id in self.processing_sessions: 
                    del self.processing_sessions[session_id]

                print(f"\n✅ Processing complete:")
                print(f"  - Items processed successfully: {items_processed}")
                print(f"  - Items failed: {items_failed}")
                print(f"  - Items with zero cost: {items_with_zero_cost}")
                print(f"  - Items with zero quantity: {items_with_zero_qty}")
                
                return jsonify({
                    'success': True, 
                    'filename': filename,
                    'items_processed': items_processed,
                    'items_failed': items_failed,
                    'debug_info': {
                        'items_with_zero_cost': items_with_zero_cost,
                        'items_with_zero_qty': items_with_zero_qty
                    }
                })
                
            except Exception as e:
                logging.error(f"Error generating final BOQ: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})

        @self.app.route('/api/download/<filename>')
        def download_file(filename):
            filepath = os.path.join(self.output_folder, filename)
            if os.path.exists(filepath):
                return send_file(filepath, as_attachment=True)
            return jsonify({'error': 'File not found'}), 404

    def run(self, host='localhost', port=5000, debug=True):
        logging.info(f"Hard-coded BOQ Processing Server starting on http://{host}:{port}")
        self.app.run(host=host, port=port, debug=debug)

if __name__ == '__main__':
    processor = HardCodedBOQProcessor()
    processor.run()