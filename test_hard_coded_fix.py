#!/usr/bin/env python3
# Test script for the fixed hard-coded BOQ processor

import openpyxl
import os
import pandas as pd
import sqlite3
from pathlib import Path
import shutil
from datetime import datetime

def test_row_calculation_fix():
    """Test the fix for row calculation issue"""
    print("Testing row calculation fix...")
    
    # Sample file
    sample_file = "uploads/Blank BOQ AIS ASP Zeer รังสิต-1.xlsx"
    if not os.path.exists(sample_file):
        print(f"Sample file not found at {sample_file}")
        return False
    
    # Make a copy for testing
    test_file = f"uploads/test_fix_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy(sample_file, test_file)
    print(f"Created test file at {test_file}")
    
    # Define sheet types
    sheet_types = {
        'INT': {
            'pattern': 'int',
            'header_row': 9,  # 0-based
            'columns': {
                'code': 2,          # Column B
                'name': 3,          # Column C
                'quantity': 4,      # Column D
                'unit': 5,          # Column E
                'material_cost': 6, # Column F
                'labor_cost': 7,    # Column G
                'total_cost': 8     # Column H
            },
            'table_name': 'interior_items'
        },
        'EE': {
            'pattern': 'ee',
            'header_row': 7,  # 0-based
            'columns': {
                'code': 2,          # Column B
                'name': 3,          # Column C
                'unit': 6,          # Column F
                'quantity': 7,      # Column G
                'material_cost': 8, # Column H
                'labor_cost': 10,   # Column J
                'total_cost': 12    # Column L
            },
            'table_name': 'ee_items'
        }
    }
    
    # Check database
    db_path = Path.home() / 'AppData' / 'Roaming' / 'BOQProcessor' / 'master_data.db'
    if not os.path.exists(db_path):
        print(f"Database not found at {db_path}")
        return False
    
    # Process sheets
    wb = openpyxl.load_workbook(test_file)
    
    for sheet_name in ['Int. 13-05-68', 'EE 13-05-68']:  # Test specific sheets
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found in workbook")
            continue
            
        print(f"\n==== Testing sheet: {sheet_name} ====")
        
        # Determine sheet type
        sheet_type = 'DEFAULT'
        for key, config in sheet_types.items():
            if config['pattern'] and config['pattern'].lower() in sheet_name.lower():
                sheet_type = key
                break
        
        print(f"Sheet type: {sheet_type}")
        
        # Get configuration
        config = sheet_types[sheet_type]
        header_row = config['header_row']
        columns = config['columns']
        
        # Read sheet with pandas
        df = pd.read_excel(test_file, sheet_name=sheet_name, header=header_row)
        
        # Get the worksheet
        sheet = wb[sheet_name]
        
        # Header row in Excel
        header_row_excel = header_row + 1
        print(f"Header row (Excel): {header_row_excel}")
        
        # Find items with costs in database
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # Get items with costs
            table_name = config['table_name']
            cursor.execute(f"SELECT name, material_cost, labor_cost FROM {table_name} WHERE material_cost > 0 OR labor_cost > 0 LIMIT 5")
            cost_items = cursor.fetchall()
            
            if not cost_items:
                print(f"No items with costs found in {table_name}")
                continue
            
            print(f"Found {len(cost_items)} items with costs in {table_name}")
            
            # Find these items in the sheet
            for item in cost_items:
                item_name = item['name']
                material_cost = item['material_cost']
                labor_cost = item['labor_cost']
                
                print(f"\nLooking for item '{item_name}' in sheet...")
                
                # Find item in dataframe
                found = False
                for idx, row in df.iterrows():
                    name_col = columns['name'] - 1  # 0-based for pandas
                    if name_col < len(row):
                        name = str(row.iloc[name_col]).strip()
                        
                        # Simple exact match for testing
                        if name == item_name:
                            found = True
                            print(f"Found at pandas index {idx}")
                            
                            # Calculate Excel row - FIXED METHOD
                            target_row_excel = header_row_excel + 1 + idx
                            
                            print(f"Target Excel row: {target_row_excel}")
                            
                            # Get cell values
                            excel_name = sheet.cell(row=target_row_excel, column=columns['name']).value
                            print(f"Excel cell value: '{excel_name}'")
                            
                            # Try to write costs
                            mat_col = columns['material_cost']
                            lab_col = columns['labor_cost']
                            
                            # Write material cost
                            print(f"Writing material cost {material_cost} to cell ({target_row_excel}, {mat_col})")
                            sheet.cell(row=target_row_excel, column=mat_col).value = material_cost
                            
                            # Write labor cost
                            print(f"Writing labor cost {labor_cost} to cell ({target_row_excel}, {lab_col})")
                            sheet.cell(row=target_row_excel, column=lab_col).value = labor_cost
                            
                            # Check if written correctly
                            mat_written = sheet.cell(row=target_row_excel, column=mat_col).value
                            lab_written = sheet.cell(row=target_row_excel, column=lab_col).value
                            
                            print(f"Values after writing: Material={mat_written}, Labor={lab_written}")
                            
                            break
                
                if not found:
                    print(f"Item '{item_name}' not found in sheet")
    
    # Save modified workbook
    output_file = f"output/test_fix_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_file)
    wb.close()
    
    print(f"\nSaved test result to {output_file}")
    print("Open this file to check if costs were written to the correct cells")
    
    # Clean up
    if os.path.exists(test_file):
        os.remove(test_file)
    
    return True

if __name__ == "__main__":
    test_row_calculation_fix()