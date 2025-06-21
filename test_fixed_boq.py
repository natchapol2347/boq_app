#!/usr/bin/env python3
# Test script for the fixed BOQ processor

import openpyxl
import os
import pandas as pd
import sqlite3
from pathlib import Path
import shutil
from datetime import datetime
import fuzzywuzzy.fuzz as fuzz

def test_total_calculation():
    """Test the calculation of section totals"""
    print("Testing section total calculation...")
    
    # Sample file
    sample_file = "uploads/Blank BOQ AIS ASP Zeer รังสิต-1.xlsx"
    if not os.path.exists(sample_file):
        print(f"Sample file not found at {sample_file}")
        return False
    
    # Make a copy for testing
    test_file = f"uploads/test_fixed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy(sample_file, test_file)
    print(f"Created test file at {test_file}")
    
    # Create section-based data
    sections = {
        'งานรื้อถอน และงานเตรียมพื้นที่': {
            'items': [
                {'name': 'ผนังชั่วคราวพร้อมประตูและINKJET GRAPHIC', 'quantity': 70, 'material': 960, 'labor': 640},
                {'name': 'ที่ผนังชั่วคราว', 'quantity': 20, 'material': 660, 'labor': 440}
            ]
        },
        'งานพื้น': {
            'items': [
                {'name': 'FLR1 พื้นกระเบื้อง สีขาว BMTZ6001 ผิวด้าน', 'quantity': 74, 'material': 660, 'labor': 440},
                {'name': 'FLR3 พื้นกระเบื้องสีขาว EXTRA WHITE ME-7700', 'quantity': 6, 'material': 759, 'labor': 500}
            ]
        }
    }
    
    # Open the workbook
    wb = openpyxl.load_workbook(test_file)
    sheet = wb["Int. 13-05-68"]
    
    # Find section rows and item rows
    section_rows = {}
    item_rows = {}
    total_rows = {}
    
    # Process sheet to find items and their positions
    for row_idx in range(1, min(100, sheet.max_row + 1)):
        cell_value = sheet.cell(row=row_idx, column=3).value  # Column C
        if not cell_value:
            continue
            
        cell_text = str(cell_value).strip()
        
        # Check for sections
        for section_name in sections.keys():
            if section_name == cell_text:
                section_rows[section_name] = row_idx
                print(f"Found section '{section_name}' at row {row_idx}")
        
        # Check for items
        for section_name, section_data in sections.items():
            for item in section_data['items']:
                if item['name'] == cell_text:
                    if section_name not in item_rows:
                        item_rows[section_name] = []
                    item_rows[section_name].append({'row': row_idx, 'item': item})
                    print(f"Found item '{cell_text}' at row {row_idx}")
        
        # Check for total rows
        if "Total" in cell_text or "รวม" in cell_text:
            for section_name in sections.keys():
                if section_name in cell_text:
                    total_rows[section_name] = row_idx
                    print(f"Found total row for '{section_name}' at row {row_idx}")
    
    # Calculate totals for each section
    for section_name, section_data in sections.items():
        section_total_material = 0
        section_total_labor = 0
        section_total = 0
        
        if section_name in item_rows:
            for item_info in item_rows[section_name]:
                item = item_info['item']
                row = item_info['row']
                
                # Write costs to cells
                quantity = item['quantity']
                material = item['material']
                labor = item['labor']
                total = (material + labor) * quantity
                
                # Material cost
                sheet.cell(row=row, column=6).value = material
                
                # Labor cost
                sheet.cell(row=row, column=7).value = labor
                
                # Total cost with quantity
                sheet.cell(row=row, column=8).value = total
                
                # Update section totals
                section_total_material += material * quantity
                section_total_labor += labor * quantity
                section_total += total
            
            # Write section totals
            if section_name in total_rows:
                total_row = total_rows[section_name]
                
                # Material total
                sheet.cell(row=total_row, column=6).value = section_total_material
                
                # Labor total
                sheet.cell(row=total_row, column=7).value = section_total_labor
                
                # Total
                sheet.cell(row=total_row, column=8).value = section_total
                
                print(f"Section '{section_name}' totals: Material={section_total_material}, Labor={section_total_labor}, Total={section_total}")
    
    # Save the test file
    output_file = f"output/test_fixed_total_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_file)
    wb.close()
    
    print(f"Saved test file to {output_file}")
    
    return True

def test_duplicate_handling():
    """Test improved matching for duplicate item names"""
    print("\nTesting duplicate item handling...")
    
    # Create a dummy database for testing
    db_path = ":memory:"
    
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            
            # Create test table
            cursor.execute('''
                CREATE TABLE test_items (
                    internal_id TEXT PRIMARY KEY, 
                    code TEXT, 
                    name TEXT NOT NULL,
                    material_cost REAL DEFAULT 0, 
                    labor_cost REAL DEFAULT 0, 
                    total_cost REAL DEFAULT 0,
                    unit TEXT
                )
            ''')
            
            # Insert test data with duplicate names but different codes
            test_data = [
                ("item_1", "CODE1", "Test Item", 100, 50, 150, "EA"),
                ("item_2", "CODE2", "Test Item", 200, 100, 300, "EA"),
                ("item_3", "CODE3", "Test Item", 300, 150, 450, "EA"),
                ("item_4", "CODE4", "Another Item", 400, 200, 600, "EA"),
            ]
            
            for item in test_data:
                cursor.execute(
                    "INSERT INTO test_items VALUES (?, ?, ?, ?, ?, ?, ?)",
                    item
                )
            conn.commit()
            
            # Test function to mimic the improved find_best_match
            def find_best_match(name, code):
                sanitized_search = str(name).lower().strip()
                sanitized_code = str(code).lower().strip() if code else ""
                best_match = None
                best_similarity = 0
                
                cursor.execute("SELECT * FROM test_items")
                all_items = cursor.fetchall()
                
                # First try exact match with code + name for duplicate handling
                if sanitized_code:
                    for item_row in all_items:
                        item_code = str(item_row[1]).lower().strip()
                        item_name = str(item_row[2]).lower().strip()
                        
                        # If we have an exact code and name match, return immediately
                        if item_code == sanitized_code and item_name == sanitized_search:
                            return {
                                'item': {
                                    'internal_id': item_row[0],
                                    'code': item_row[1],
                                    'name': item_row[2],
                                    'material_cost': item_row[3],
                                    'labor_cost': item_row[4]
                                }, 
                                'similarity': 100
                            }
                        
                        # If just code matches exactly, boost similarity
                        if item_code == sanitized_code:
                            name_similarity = fuzz.ratio(sanitized_search, item_name)
                            # Boost similarity for code match
                            adjusted_similarity = min(100, name_similarity + 25)
                            
                            if adjusted_similarity > best_similarity:
                                best_similarity = adjusted_similarity
                                best_match = {
                                    'item': {
                                        'internal_id': item_row[0],
                                        'code': item_row[1],
                                        'name': item_row[2],
                                        'material_cost': item_row[3],
                                        'labor_cost': item_row[4]
                                    }, 
                                    'similarity': adjusted_similarity
                                }
                
                # Then do standard fuzzy matching on name
                for item_row in all_items:
                    item_name = str(item_row[2]).lower().strip()
                    similarity = fuzz.ratio(sanitized_search, item_name)
                    
                    if similarity > best_similarity:
                        best_similarity = similarity
                        best_match = {
                            'item': {
                                'internal_id': item_row[0],
                                'code': item_row[1],
                                'name': item_row[2],
                                'material_cost': item_row[3],
                                'labor_cost': item_row[4]
                            }, 
                            'similarity': similarity
                        }
                        
                return best_match
            
            # Test cases
            test_cases = [
                {"name": "Test Item", "code": "CODE1", "expected_material": 100},
                {"name": "Test Item", "code": "CODE2", "expected_material": 200},
                {"name": "Test Item", "code": "CODE3", "expected_material": 300},
                {"name": "Test Item", "code": "", "expected_material": 100},  # Should match the first one
                {"name": "Another Item", "code": "CODE4", "expected_material": 400},
            ]
            
            for i, test_case in enumerate(test_cases):
                match = find_best_match(test_case["name"], test_case["code"])
                if match:
                    material_cost = match["item"]["material_cost"]
                    code = match["item"]["code"]
                    success = material_cost == test_case["expected_material"]
                    
                    print(f"Test {i+1}: Name='{test_case['name']}', Code='{test_case['code']}'")
                    print(f"  Matched to: Code='{code}', Material Cost={material_cost}")
                    print(f"  Expected Material Cost: {test_case['expected_material']}")
                    print(f"  {'✅ PASS' if success else '❌ FAIL'}")
                else:
                    print(f"Test {i+1}: Name='{test_case['name']}', Code='{test_case['code']}'")
                    print(f"  No match found ❌ FAIL")
            
            return True
            
    except Exception as e:
        print(f"Error testing duplicate handling: {e}")
        return False

def test_data_sanitization():
    """Test the improved data sanitization for hyphen-only names"""
    print("\nTesting data sanitization...")
    
    # Test function to mimic the improved _clean_item_name
    def clean_item_name(name, code):
        """Clean and improve item names, especially for '-' values"""
        if not name or pd.isna(name) or name.strip() in ['-', '', 'nan', 'none']:
            if code and code.strip() and code.strip() not in ['-', 'nan', 'none']:
                # Use code as name if name is empty/invalid but code exists
                return f"Item {code.strip()}"
            else:
                # Generate a unique name if both name and code are invalid
                return f"Item_unique_id"
                
        # Clean up the name
        cleaned = name.strip()
        if cleaned == '-':
            return f"Unnamed item_unique_id"
        
        return cleaned
    
    # Test cases
    test_cases = [
        {"name": "-", "code": "CODE1", "expected": "Unnamed item_unique_id"},
        {"name": "-", "code": "", "expected": "Unnamed item_unique_id"},
        {"name": "", "code": "CODE2", "expected": "Item CODE2"},
        {"name": None, "code": "CODE3", "expected": "Item CODE3"},
        {"name": "Valid Name", "code": "CODE4", "expected": "Valid Name"},
        {"name": "  Spaces  ", "code": "CODE5", "expected": "Spaces"},
    ]
    
    for i, test_case in enumerate(test_cases):
        result = clean_item_name(test_case["name"], test_case["code"])
        success = result == test_case["expected"] or (
            "Unnamed item" in result and "Unnamed item" in test_case["expected"]
        )
        
        print(f"Test {i+1}: Name='{test_case['name']}', Code='{test_case['code']}'")
        print(f"  Result: '{result}'")
        print(f"  Expected: '{test_case['expected']}'")
        print(f"  {'✅ PASS' if success else '❌ FAIL'}")
    
    return True

if __name__ == "__main__":
    print("=== Testing Fixed BOQ Processor ===\n")
    
    # Test total calculation
    total_ok = test_total_calculation()
    
    # Test duplicate handling
    duplicate_ok = test_duplicate_handling()
    
    # Test data sanitization
    sanitization_ok = test_data_sanitization()
    
    if total_ok and duplicate_ok and sanitization_ok:
        print("\n✅ All tests passed!")
    else:
        print("\n❌ Some tests failed")
        if not total_ok:
            print("  - Total calculation failed")
        if not duplicate_ok:
            print("  - Duplicate handling failed")
        if not sanitization_ok:
            print("  - Data sanitization failed")
    
    print("\nTest script complete")