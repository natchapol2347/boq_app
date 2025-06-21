"""
Test script to check the matching logic and cost retrieval from the database
"""
import sqlite3
from pathlib import Path
import os
import pandas as pd
from fuzzywuzzy import fuzz
import openpyxl

def test_database_costs():
    """Test if the database contains costs"""
    db_path = Path.home() / 'AppData' / 'Roaming' / 'BOQProcessor' / 'master_data.db'
    
    if not os.path.exists(db_path):
        print(f"Database not found at {db_path}")
        return False
    
    try:
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # Check if we have items with costs
            cursor.execute("SELECT COUNT(*) FROM master_items WHERE material_cost > 0 OR labor_cost > 0")
            count = cursor.fetchone()[0]
            
            print(f"Database has {count} items with non-zero costs")
            
            # Show some sample items with costs
            cursor.execute("SELECT name, material_cost, labor_cost, total_cost FROM master_items WHERE material_cost > 0 OR labor_cost > 0 LIMIT 10")
            items = cursor.fetchall()
            
            if items:
                print("\nSample items with costs:")
                for item in items:
                    print(f"  - {item['name']}: Material={item['material_cost']}, Labor={item['labor_cost']}, Total={item['total_cost']}")
            else:
                print("No items with costs found in database")
            
            return count > 0
    except Exception as e:
        print(f"Error checking database: {e}")
        return False

def test_matching_logic():
    """Test the fuzzy matching logic used in the original code"""
    # Simulate the original find_best_match method
    def find_best_match(item_name, db_items):
        """Find best matching item from database using fuzzy matching"""
        if not item_name or pd.isna(item_name): 
            return None
            
        if not db_items: 
            return None
        
        sanitized_search = str(item_name).lower().strip()
        best_match = None
        best_similarity = 0
        
        for item_row in db_items:
            item_dict = dict(item_row)
            sanitized_candidate = str(item_dict['name']).lower().strip()
            similarity = fuzz.ratio(sanitized_search, sanitized_candidate)
            
            if similarity > best_similarity:
                best_similarity = similarity
                best_match = {'item': item_dict, 'similarity': similarity}
                
        return best_match
    
    # Get items from database
    db_path = Path.home() / 'AppData' / 'Roaming' / 'BOQProcessor' / 'master_data.db'
    
    if not os.path.exists(db_path):
        print(f"Database not found at {db_path}")
        return False
    
    # Get items from the BOQ file
    boq_file = "/Users/a677022/Desktop/woodman/boq_app/output/final_boq_20250619_122910.xlsx"
    
    if not os.path.exists(boq_file):
        print(f"BOQ file not found: {boq_file}")
        return False
    
    try:
        # Get items from database
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            all_items = conn.execute("SELECT * FROM master_items").fetchall()
        
        print(f"Retrieved {len(all_items)} items from database")
        
        # Get items from BOQ file
        wb = openpyxl.load_workbook(boq_file)
        
        # Test with one sheet
        sheet_name = "Int. 13-05-68"  # Test with Interior sheet
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found in BOQ file")
            return False
        
        sheet = wb[sheet_name]
        
        # Find header row
        header_row = None
        for row_idx in range(1, min(20, sheet.max_row + 1)):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value == "ลำดับ":
                header_row = row_idx
                print(f"Found header at row {row_idx}")
                break
        
        if not header_row:
            print("Could not find header row")
            return False
        
        # Define columns
        code_col = 2  # B
        name_col = 3  # C
        
        # Test matching for items
        matches_found = 0
        items_tested = 0
        matches_with_costs = 0
        
        print("\nTesting matching logic with items from BOQ file:")
        for row_idx in range(header_row + 1, min(header_row + 30, sheet.max_row + 1)):  # Test first 30 rows
            code = sheet.cell(row=row_idx, column=code_col).value
            name = sheet.cell(row=row_idx, column=name_col).value
            
            if not name and not code:
                continue
                
            if not name:
                continue
                
            # Skip header rows
            name_lower = str(name).lower()
            if any(keyword in name_lower for keyword in ['total', 'รวม', 'system', 'ระบบ']):
                continue
            
            items_tested += 1
            print(f"\nTesting item {items_tested}: '{name}'")
            
            match = find_best_match(name, all_items)
            
            if match:
                matches_found += 1
                match_item = match['item']
                print(f"  Match found with {match['similarity']}% similarity")
                print(f"  Matched to: '{match_item['name']}'")
                
                material_cost = match_item.get('material_cost', 0)
                labor_cost = match_item.get('labor_cost', 0)
                
                print(f"  Costs: Material={material_cost}, Labor={labor_cost}")
                
                if material_cost > 0 or labor_cost > 0:
                    matches_with_costs += 1
            else:
                print(f"  No match found")
        
        print(f"\nMatching test results:")
        print(f"  Items tested: {items_tested}")
        print(f"  Matches found: {matches_found} ({matches_found/items_tested*100 if items_tested > 0 else 0:.1f}%)")
        print(f"  Matches with costs: {matches_with_costs} ({matches_with_costs/matches_found*100 if matches_found > 0 else 0:.1f}%)")
        
        return True
    
    except Exception as e:
        print(f"Error testing matching logic: {e}")
        return False

if __name__ == "__main__":
    print("Testing database for costs...")
    test_database_costs()
    
    print("\nTesting matching logic...")
    test_matching_logic()