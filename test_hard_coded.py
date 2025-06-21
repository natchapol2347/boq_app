#!/usr/bin/env python3
# Test script for hard-coded BOQ processor

import sqlite3
from pathlib import Path
import os
import openpyxl
import shutil
import uuid
from datetime import datetime

def test_database_setup():
    """Test if the database is set up with the right tables and sample data"""
    print("Testing database setup...")
    
    # Clear existing database to force fresh setup
    db_path = Path.home() / 'AppData' / 'Roaming' / 'BOQProcessor' / 'master_data.db'
    
    if os.path.exists(db_path):
        print(f"Backing up existing database from {db_path}")
        backup_path = str(db_path) + f".bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        shutil.copy(db_path, backup_path)
        os.remove(db_path)
        print(f"Old database backed up to {backup_path} and removed")
    
    # Import the HardCodedBOQProcessor
    from hard_coded_boq import HardCodedBOQProcessor
    
    # Initialize processor (this should create the database)
    processor = HardCodedBOQProcessor()
    
    # Check if database was created
    if not os.path.exists(db_path):
        print(f"❌ Database was not created at {db_path}")
        return False
    
    print(f"✓ Database created at {db_path}")
    
    # Check database tables
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            
            # Get list of tables
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in cursor.fetchall()]
            
            print(f"Database tables: {tables}")
            
            # Expected tables
            expected_tables = ['interior_items', 'ee_items', 'ac_items', 'fp_items', 'default_items']
            
            # Check if all expected tables exist
            for table in expected_tables:
                if table not in tables:
                    print(f"❌ Missing table: {table}")
                    return False
                
                # Check if table has data
                cursor.execute(f"SELECT COUNT(*) FROM {table}")
                count = cursor.fetchone()[0]
                print(f"Table {table}: {count} items")
                
                # Check if items have costs
                cursor.execute(f"SELECT COUNT(*) FROM {table} WHERE material_cost > 0")
                cost_count = cursor.fetchone()[0]
                print(f"  Items with costs: {cost_count}")
                
                if cost_count == 0:
                    print(f"❌ No items with costs in table {table}")
                    return False
                
                # Show some sample items
                cursor.execute(f"SELECT name, material_cost, labor_cost FROM {table} LIMIT 3")
                items = cursor.fetchall()
                for item in items:
                    print(f"  - {item[0]}: Material={item[1]}, Labor={item[2]}")
        
        print("✓ All expected tables exist with cost data")
        return True
        
    except Exception as e:
        print(f"❌ Error checking database: {e}")
        return False

def test_process_boq():
    """Test processing a BOQ file"""
    print("\nTesting BOQ processing...")
    
    # Check if we have a sample BOQ file
    sample_file = "uploads/Blank BOQ AIS ASP Zeer รังสิต-1.xlsx"
    
    if not os.path.exists(sample_file):
        print(f"❌ Sample BOQ file not found at {sample_file}")
        return False
    
    # Make a copy for testing
    test_file = f"uploads/test_boq_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy(sample_file, test_file)
    print(f"Created test file at {test_file}")
    
    # Import the processor
    from hard_coded_boq import HardCodedBOQProcessor
    
    # Initialize processor
    processor = HardCodedBOQProcessor()
    
    # Check sheet types and formats
    try:
        wb = openpyxl.load_workbook(test_file)
        print(f"BOQ file contains {len(wb.sheetnames)} sheets: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            if "sum" in sheet_name.lower():
                continue
                
            sheet_type = processor.determine_sheet_type(sheet_name)
            print(f"Sheet {sheet_name} determined to be type: {sheet_type}")
            
            # Check header row
            header_row = processor.sheet_types[sheet_type]['header_row']
            print(f"  Header row: {header_row} (Excel row {header_row+1})")
            
            # Check columns
            columns = processor.sheet_types[sheet_type]['columns']
            print(f"  Columns: {columns}")
            
            # Verify that the fixed positions make sense
            sheet = wb[sheet_name]
            header_row_excel = header_row + 1
            
            print(f"  Header row content:")
            for col_name, col_num in columns.items():
                col_letter = openpyxl.utils.get_column_letter(col_num)
                cell_value = sheet.cell(row=header_row_excel, column=col_num).value
                print(f"    {col_name} (column {col_letter}): '{cell_value}'")
        
        wb.close()
        print("✓ Sheet formats verified")
        
        # Simulate processing
        print("\nTesting direct processing...")
        
        # Open the test file and match a few items directly
        wb = openpyxl.load_workbook(test_file)
        
        for sheet_name in wb.sheetnames:
            if "sum" in sheet_name.lower():
                continue
                
            sheet_type = processor.determine_sheet_type(sheet_name)
            sheet = wb[sheet_name]
            columns = processor.sheet_types[sheet_type]['columns']
            header_row = processor.sheet_types[sheet_type]['header_row']
            name_col = columns.get('name')
            
            # Check a few rows
            print(f"\nDirectly checking sheet: {sheet_name} (type: {sheet_type})")
            
            for row_idx in range(header_row + 2, header_row + 7):  # Check a few rows after header
                name_cell = sheet.cell(row=row_idx, column=name_col)
                name = name_cell.value
                
                if not name or any(keyword in str(name).lower() for keyword in ['total', 'รวม', 'system']):
                    continue
                
                print(f"Row {row_idx}, Item: '{name}'")
                
                # Try to match in database
                match = processor.find_best_match(name, sheet_type)
                
                if match:
                    print(f"  ✓ Match found with {match['similarity']}% similarity")
                    print(f"  Matched to: '{match['item']['name']}'")
                    print(f"  Costs: Material={match['item']['material_cost']}, Labor={match['item']['labor_cost']}")
                else:
                    print(f"  ❌ No match found")
        
        wb.close()
        print("\n✓ Direct processing test complete")
        
        # Clean up
        if os.path.exists(test_file):
            os.remove(test_file)
            print(f"Removed test file {test_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error testing process_boq: {e}")
        if os.path.exists(test_file):
            os.remove(test_file)
        return False

if __name__ == "__main__":
    print("=== Testing Hard-Coded BOQ Processor ===\n")
    
    # Test database setup
    database_ok = test_database_setup()
    
    if database_ok:
        # Test BOQ processing
        process_ok = test_process_boq()
        
        if process_ok:
            print("\n✅ All tests passed!")
        else:
            print("\n❌ BOQ processing test failed")
    else:
        print("\n❌ Database setup test failed")
    
    print("\nTest script complete")