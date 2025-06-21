"""
Final fix for BOQ costs that handles merged cells
"""
import openpyxl
import os
from openpyxl.utils import get_column_letter

def apply_costs_safely(input_file, output_file):
    """Apply costs to a BOQ file while handling merged cells"""
    if not os.path.exists(input_file):
        print(f"Input file not found: {input_file}")
        return False
    
    print(f"Opening BOQ file: {input_file}")
    
    # Load workbook
    wb = openpyxl.load_workbook(input_file)
    
    # Prepare fixed costs to apply
    sample_costs = {
        "material": 500.0,
        "labor": 300.0,
        "total": 800.0,
        "markups": {
            100: 1600.0,  # 100% markup: 800 * (1+1)
            130: 1840.0,  # 130% markup: 800 * (1+1.3)
            150: 2000.0,  # 150% markup: 800 * (1+1.5)
            50: 1200.0,   # 50% markup: 800 * (1+0.5)
            30: 1040.0    # 30% markup: 800 * (1+0.3)
        }
    }
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        if "sum" in sheet_name.lower():
            continue
            
        print(f"\n=== Processing sheet: {sheet_name} ===")
        sheet = wb[sheet_name]
        
        # Determine column positions based on sheet type
        if "Int" in sheet_name:
            # Interior sheet
            material_col = 6  # F
            labor_col = 7     # G
            total_col = 8     # H
            print("Using Interior sheet mapping: F=material, G=labor, H=total")
        else:
            # System sheets
            material_col = 8  # H
            labor_col = 10    # J
            total_col = 12    # L
            print("Using System sheet mapping: H=material, J=labor, L=total")
        
        # Find header row
        header_row = None
        for row_idx in range(1, min(20, sheet.max_row + 1)):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value == "ลำดับ":
                header_row = row_idx
                print(f"Found header at row {row_idx}")
                break
        
        if not header_row:
            print(f"Could not find header row in {sheet_name}, skipping")
            continue
        
        # Find markup columns
        markup_cols = []
        for col_idx in range(material_col, sheet.max_column + 1):
            cell_value = str(sheet.cell(row=header_row, column=col_idx).value or '')
            if "markup" in cell_value.lower():
                # Extract markup percentage
                try:
                    pct = int(''.join(filter(str.isdigit, cell_value)))
                    markup_cols.append((col_idx, pct))
                    print(f"Found markup column {col_idx} for {pct}%")
                except:
                    markup_cols.append((col_idx, 100))  # Default to 100% if can't parse
        
        # Get all merged cell ranges in the sheet
        merged_ranges = sheet.merged_cells.ranges
        
        # Check code/name columns
        code_col = 2  # B
        name_col = 3  # C
        
        # Apply costs to actual items
        items_updated = 0
        items_skipped = 0
        
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            code = sheet.cell(row=row_idx, column=code_col).value
            name = sheet.cell(row=row_idx, column=name_col).value
            
            # Skip empty rows
            if not name and not code:
                continue
                
            # If name is empty but there's content in column D, use that
            if not name and "Int" not in sheet_name:
                # For system sheets, check column D for description
                name = sheet.cell(row=row_idx, column=4).value
            
            if not name:
                continue
                
            # Skip header rows
            name_lower = str(name).lower()
            if any(keyword in name_lower for keyword in ['total', 'รวม', 'system', 'ระบบ']):
                continue
            
            try:
                # Apply costs safely (handling merged cells)
                safe_write_to_cell(sheet, row_idx, material_col, sample_costs["material"], merged_ranges)
                safe_write_to_cell(sheet, row_idx, labor_col, sample_costs["labor"], merged_ranges)
                safe_write_to_cell(sheet, row_idx, total_col, sample_costs["total"], merged_ranges)
                
                # Apply markups safely
                for col_idx, pct in markup_cols:
                    markup_value = sample_costs["markups"].get(pct, sample_costs["total"] * 2)
                    safe_write_to_cell(sheet, row_idx, col_idx, markup_value, merged_ranges)
                
                items_updated += 1
                if items_updated % 10 == 0:
                    print(f"Updated {items_updated} items...")
            except Exception as e:
                print(f"Error updating row {row_idx}: {e}")
                items_skipped += 1
        
        print(f"Total items updated in {sheet_name}: {items_updated}")
        print(f"Items skipped due to errors: {items_skipped}")
    
    # Save the workbook
    wb.save(output_file)
    print(f"\nSaved file with costs applied: {output_file}")
    return True

def safe_write_to_cell(sheet, row, col, value, merged_ranges):
    """Safely write a value to a cell, handling merged cells"""
    cell = sheet.cell(row=row, column=col)
    
    # Check if this is a merged cell
    cell_coord = cell.coordinate
    for merged_range in merged_ranges:
        if cell_coord in merged_range:
            # Get the top-left cell of the merged range
            min_row = merged_range.min_row
            min_col = merged_range.min_col
            
            # Use the top-left cell instead
            if row != min_row or col != min_col:
                cell = sheet.cell(row=min_row, column=min_col)
                #print(f"Using main cell {cell.coordinate} instead of merged cell {cell_coord}")
    
    # Now set the value
    cell.value = value
    
    # Set number format for numeric values
    if isinstance(value, (int, float)):
        cell.number_format = '#,##0.00'
    
    return True

if __name__ == "__main__":
    input_file = "/Users/a677022/Desktop/woodman/boq_app/output/final_boq_20250619_122910.xlsx"
    output_file = "/Users/a677022/Desktop/woodman/boq_app/output/final_boq_with_costs.xlsx"
    apply_costs_safely(input_file, output_file)