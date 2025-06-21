import pandas as pd
import sqlite3
import os
from pathlib import Path
import openpyxl

def fix_cost_issues():
    """Diagnose and fix cost issues in the BOQ processor"""
    print("Starting diagnosis of cost issues...")
    
    # 1. Check the database contents to confirm costs are there
    db_path = Path.home() / 'AppData' / 'Roaming' / 'BOQProcessor' / 'master_data.db'
    if not os.path.exists(db_path):
        print(f"Database not found at {db_path}")
        return
    
    # Connect to database and check items
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Check item count
    cursor.execute("SELECT COUNT(*) FROM master_items")
    count = cursor.fetchone()[0]
    print(f"Database contains {count} master items")
    
    # Check items with costs
    cursor.execute("SELECT COUNT(*) FROM master_items WHERE material_cost > 0 OR labor_cost > 0")
    items_with_costs = cursor.fetchone()[0]
    print(f"Items with costs: {items_with_costs} ({items_with_costs/count*100 if count > 0 else 0:.1f}%)")
    
    # Check the blank BOQ structure
    print("\nChecking the blank BOQ structure...")
    boq_file = os.path.join("uploads", "Blank BOQ AIS ASP Zeer รังสิต-1.xlsx")
    
    if not os.path.exists(boq_file):
        print(f"Blank BOQ file not found at {boq_file}")
        return
    
    # Get sheet names
    wb = openpyxl.load_workbook(boq_file, read_only=True)
    sheet_names = wb.sheetnames
    print(f"Blank BOQ contains {len(sheet_names)} sheets: {sheet_names}")
    
    # Check if any recent output files exist to compare
    output_files = [f for f in os.listdir("output") if f.startswith("final_boq_")]
    if output_files:
        latest_output = sorted(output_files)[-1]
        print(f"\nExamining latest output file: {latest_output}")
        output_path = os.path.join("output", latest_output)
        
        # Open the output file and check if costs were written
        output_wb = openpyxl.load_workbook(output_path, read_only=True)
        
        for sheet_name in output_wb.sheetnames:
            if "sum" in sheet_name.lower():
                continue
                
            print(f"\nInspecting sheet: {sheet_name}")
            sheet = output_wb[sheet_name]
            
            # Find header row
            header_row = None
            header_indicators = ['ลำดับ', 'code', 'รายการ', 'จำนวน', 'หน่วย']
            
            for row_idx in range(1, min(20, sheet.max_row + 1)):
                row_values = [str(cell.value).lower() if cell.value else "" for cell in sheet[row_idx]]
                matches = sum(1 for indicator in header_indicators 
                             if any(indicator in cell for cell in row_values if cell))
                if matches >= 3:
                    header_row = row_idx
                    print(f"Header row found at row {row_idx}")
                    print(f"Header content: {[cell.value for cell in sheet[row_idx]]}")
                    break
            
            if not header_row:
                print(f"No header row found in sheet {sheet_name}")
                continue
            
            # Find cost columns
            material_col = None
            labor_col = None
            total_col = None
            
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).lower()
                    if any(term in cell_str for term in ['วัสดุ', 'material']):
                        material_col = col_idx
                    elif any(term in cell_str for term in ['แรงงาน', 'labor', 'labour']):
                        labor_col = col_idx
                    elif any(term in cell_str for term in ['รวม', 'total']):
                        total_col = col_idx
            
            print(f"Material cost column: {material_col}")
            print(f"Labor cost column: {labor_col}")
            print(f"Total cost column: {total_col}")
            
            # Check for non-zero costs
            non_zero_rows = 0
            zero_rows = 0
            rows_checked = 0
            
            for row_idx in range(header_row + 1, min(header_row + 50, sheet.max_row + 1)):
                name_value = None
                material_value = 0
                labor_value = 0
                total_value = 0
                
                # Find the name column
                name_col = None
                for col_idx in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=header_row, column=col_idx).value
                    if cell_value and any(term in str(cell_value).lower() for term in ['รายการ', 'name', 'description']):
                        name_col = col_idx
                        break
                
                if name_col:
                    name_value = sheet.cell(row=row_idx, column=name_col).value
                
                # Only check rows with a name
                if name_value and len(str(name_value).strip()) > 2:
                    rows_checked += 1
                    
                    # Get cost values
                    if material_col:
                        cell = sheet.cell(row=row_idx, column=material_col)
                        try:
                            material_value = float(cell.value or 0)
                        except (ValueError, TypeError):
                            material_value = 0
                    
                    if labor_col:
                        cell = sheet.cell(row=row_idx, column=labor_col)
                        try:
                            labor_value = float(cell.value or 0)
                        except (ValueError, TypeError):
                            labor_value = 0
                    
                    if total_col:
                        cell = sheet.cell(row=row_idx, column=total_col)
                        try:
                            total_value = float(cell.value or 0)
                        except (ValueError, TypeError):
                            total_value = 0
                    
                    # Check if any cost value is non-zero
                    if material_value > 0 or labor_value > 0 or total_value > 0:
                        non_zero_rows += 1
                    else:
                        zero_rows += 1
                        
                        # Look up this item in the database
                        cursor.execute("SELECT name, material_cost, labor_cost, total_cost FROM master_items WHERE name LIKE ?", 
                                     (f"%{name_value}%",))
                        db_matches = cursor.fetchall()
                        
                        print(f"\nRow {row_idx} has zero costs - Item: {name_value}")
                        if db_matches:
                            print(f"  Found {len(db_matches)} potential matches in database:")
                            for match in db_matches:
                                print(f"  - {match[0][:40]}: Material: {match[1]}, Labor: {match[2]}, Total: {match[3]}")
                        else:
                            print(f"  No matches found in database for this item")
            
            print(f"\nSummary for sheet {sheet_name}:")
            print(f"  Rows checked: {rows_checked}")
            print(f"  Rows with non-zero costs: {non_zero_rows}")
            print(f"  Rows with zero costs: {zero_rows}")
    
    # Identify fixes needed
    print("\n===== DIAGNOSIS SUMMARY =====")
    if items_with_costs > 0:
        print("✓ Database has cost data")
    else:
        print("✗ Database is missing cost data - needs to be reloaded from master.xlsx")
    
    if output_files:
        print("\nPossible issues:")
        print("1. Column mapping between master data and blank BOQ is incorrect")
        print("2. Fuzzy matching might not be identifying the correct items")
        print("3. Cost data might not be properly parsed from master.xlsx")
        print("4. Excel cell formatting issues when writing costs")
        
        print("\nRecommended fixes:")
        print("1. Improve column detection in the blank BOQ file")
        print("2. Enhance cost column mapping with more aggressive pattern matching")
        print("3. Ensure cost data is properly parsed and converted to numeric values")
        print("4. Fix Excel cell writing issues with proper number formatting")
        print("5. Debug the matching algorithm to ensure correct items are found")
    
    conn.close()

if __name__ == "__main__":
    fix_cost_issues()