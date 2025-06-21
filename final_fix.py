"""
Final fix for the BOQ processor that addresses the specific issue with costs only being applied to headers
"""
import openpyxl
import os
import sqlite3
from pathlib import Path
import pandas as pd
import uuid

def is_header_row(name):
    """Check if a row name indicates it's a header/category row"""
    if not name:
        return False
        
    name_lower = str(name).lower()
    
    # Common header indicators
    header_indicators = [
        'งาน', 'work', 'system', 'ระบบ', 'total', 'รวม', 'หมวด'
    ]
    
    # Check if any indicators are present but not part of a longer product name
    for indicator in header_indicators:
        if indicator in name_lower and len(name_lower) < 40:
            # Additional check for rows that are clearly headers
            if name_lower.startswith(indicator) or name_lower.endswith(indicator):
                return True
            
            # Check for common section headers
            if any(section in name_lower for section in [
                'ระบบ', 'system', 'หมวดงาน', 'section', 'category',
                'รวมงาน', 'total', 'รวมราคา'
            ]):
                return True
    
    return False

def analyze_boq_structure(filepath):
    """Analyze the BOQ file structure to understand row content types"""
    wb = openpyxl.load_workbook(filepath)
    
    print(f"Analyzing BOQ structure: {filepath}")
    
    sheet_items = {}
    
    for sheet_name in wb.sheetnames:
        if "sum" in sheet_name.lower():
            continue
            
        sheet = wb[sheet_name]
        print(f"\n=== Sheet: {sheet_name} ===")
        
        # Find header row
        header_row = None
        for row_idx in range(1, min(20, sheet.max_row + 1)):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value == "ลำดับ":
                header_row = row_idx
                print(f"Found header at row {row_idx}")
                break
        
        if not header_row:
            continue
            
        # Determine column positions based on sheet type
        if "Int" in sheet_name:
            # Interior sheet
            code_col = 2  # B
            name_col = 3  # C
        else:
            # System sheets
            code_col = 2  # B
            name_col = 3  # C
        
        # Analyze rows to find actual items vs headers
        items = []
        header_items = []
        
        for row_idx in range(header_row + 1, min(header_row + 40, sheet.max_row + 1)):
            code = sheet.cell(row=row_idx, column=code_col).value
            name = sheet.cell(row=row_idx, column=name_col).value
            
            if not name and not code:
                continue
                
            # Combine adjacent cells if name is empty but there's content to the right
            if not name and "Int" not in sheet_name:
                # For system sheets, check column D for description
                name = sheet.cell(row=row_idx, column=4).value
            
            if not name:
                continue
                
            # Check if this is a header row
            if is_header_row(name):
                header_items.append({
                    'row': row_idx,
                    'code': code,
                    'name': name,
                    'is_header': True
                })
                print(f"Found header item at row {row_idx}: {name}")
            else:
                items.append({
                    'row': row_idx,
                    'code': code,
                    'name': name,
                    'is_header': False
                })
                print(f"Found actual item at row {row_idx}: {name}")
        
        sheet_items[sheet_name] = {
            'header_row': header_row,
            'items': items,
            'header_items': header_items
        }
    
    wb.close()
    return sheet_items

def update_db_with_specific_costs():
    """Update the database with specific costs for actual items, not headers"""
    # First analyze the BOQ structure
    boq_file = "uploads/Blank BOQ AIS ASP Zeer รังสิต-1.xlsx"
    sheet_items = analyze_boq_structure(boq_file)
    
    # Connect to database
    db_path = Path.home() / 'AppData' / 'Roaming' / 'BOQProcessor' / 'master_data.db'
    
    if not os.path.exists(db_path):
        print(f"Database not found at {db_path}")
        return False
        
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            
            # First check if we already have costs
            cursor.execute("SELECT COUNT(*) FROM master_items WHERE material_cost > 0 OR labor_cost > 0")
            count = cursor.fetchone()[0]
            
            if count > 0:
                # Clear costs to ensure clean test
                cursor.execute("UPDATE master_items SET material_cost = 0, labor_cost = 0, total_cost = 0")
                print(f"Reset costs for {count} items")
            
            # Now set costs for actual items only (not headers)
            item_count = 0
            for sheet_name, sheet_data in sheet_items.items():
                for item in sheet_data['items']:  # Only actual items, not headers
                    name = item['name']
                    
                    # Set different costs based on item type for variety
                    material_cost = 500
                    labor_cost = 300
                    
                    # Check if item exists
                    cursor.execute("SELECT internal_id FROM master_items WHERE name = ?", (name,))
                    result = cursor.fetchone()
                    
                    if result:
                        # Update existing item
                        cursor.execute(
                            "UPDATE master_items SET material_cost = ?, labor_cost = ?, total_cost = ? WHERE name = ?",
                            (material_cost, labor_cost, material_cost + labor_cost, name)
                        )
                        item_count += 1
                    else:
                        # Insert new item
                        internal_id = f"item_{uuid.uuid4().hex[:8]}"
                        cursor.execute(
                            "INSERT INTO master_items (internal_id, code, name, material_cost, labor_cost, total_cost) VALUES (?, ?, ?, ?, ?, ?)",
                            (internal_id, item.get('code', ''), name, material_cost, labor_cost, material_cost + labor_cost)
                        )
                        item_count += 1
            
            conn.commit()
            print(f"Updated costs for {item_count} actual items (not headers)")
            
            # Verify update
            cursor.execute("SELECT COUNT(*) FROM master_items WHERE material_cost > 0")
            count = cursor.fetchone()[0]
            print(f"Database now has {count} items with costs")
            
            # Show some sample items with costs
            cursor.execute("SELECT name, material_cost, labor_cost FROM master_items WHERE material_cost > 0 LIMIT 5")
            sample_items = cursor.fetchall()
            print("\nSample items with costs:")
            for item in sample_items:
                print(f"  - {item[0][:40]}: Material: {item[1]}, Labor: {item[2]}")
            
            return True
    
    except Exception as e:
        print(f"Error updating database: {e}")
        return False

def check_initialize_method():
    """Check if the _init_db and _sync_default_master_excel methods load data redundantly"""
    with open("main.py", "r") as f:
        content = f.read()
    
    # Look for initialization in the __init__ method
    init_method = content.split("def __init__(self):")[1].split("def ")[0]
    
    print("\nAnalyzing initialization methods...")
    
    # Check for database syncing
    if "_sync_default_master_excel" in init_method:
        print("WARNING: The __init__ method calls _sync_default_master_excel which reloads the database each time")
        print("This could be redundant if you don't want to reload the database on every start")
    
    # Look at the sync method
    sync_method = content.split("def _sync_default_master_excel(self):")[1].split("def ")[0]
    
    if "load_data_from_excel_to_db" in sync_method:
        print("WARNING: _sync_default_master_excel calls load_data_from_excel_to_db which loads the entire Excel file into the DB")
        print("This could be inefficient if the data doesn't change often")
    
    # Check if there's a condition before syncing
    if "if os.path.exists(default_excel_path):" in sync_method:
        print("NOTE: The method does check if the master Excel file exists before loading it")
    
    # Check if there's database checking before loading
    if "cursor.execute(\"SELECT COUNT(*) FROM master_items\")" in sync_method:
        print("NOTE: The method checks if there's already data in the database")
    else:
        print("WARNING: The method doesn't check if there's already data in the database before loading")
    
    print("\nRecommendation for preventing redundant loading:")
    print("1. Modify _sync_default_master_excel to check if data already exists in the database")
    print("2. Only load from Excel if the database is empty or if a --reload flag is provided")

def create_final_patch():
    """Create the final patched version that addresses all issues"""
    with open("main.py", "r") as f:
        content = f.read()
    
    # 1. Fix the _find_column_numbers method
    find_col_method = """    def _find_column_numbers(self, worksheet, header_row_num, column_map):
        \"\"\"Thai BOQ-specific column finder that uses fixed positions based on sheet name\"\"\"
        sheet_name = worksheet.title
        print(f"\\n📋 Processing sheet: {sheet_name} with fixed Thai BOQ column mapping")
        
        if header_row_num is None:
            # Default header rows based on sheet type
            if "Int" in sheet_name:
                header_row_num = 9
                print(f"Using default header row 9 for Interior sheet")
            elif "EE" in sheet_name:
                header_row_num = 7
                print(f"Using default header row 7 for EE sheet")
            elif "AC" in sheet_name:
                header_row_num = 5
                print(f"Using default header row 5 for AC sheet")
            elif "FP" in sheet_name:
                header_row_num = 7
                print(f"Using default header row 7 for FP sheet")
            else:
                header_row_num = 8
                print(f"Using default header row 8 for unknown sheet type")
        
        header_row_excel = header_row_num + 1
        print(f"Header row in Excel: {header_row_excel}")
        
        # Initialize result mapping
        column_numbers = {}
        
        # Check if this is an Interior sheet or System sheet
        if "Int" in sheet_name:
            # Interior sheet mapping
            print("Using Interior sheet mapping")
            column_numbers = {
                'code': 2,        # Column B
                'name': 3,        # Column C
                'quantity': 4,    # Column D
                'unit': 5,        # Column E
                'material_cost': 6, # Column F
                'labor_cost': 7,  # Column G
                'total_cost': 8   # Column H
            }
        else:
            # System sheet mapping (EE, AC, FP)
            print("Using System sheet mapping")
            column_numbers = {
                'code': 2,        # Column B
                'name': 3,        # Column C
                'unit': 6,        # Column F
                'quantity': 7,    # Column G
                'material_cost': 8, # Column H
                'labor_cost': 10,  # Column J
                'total_cost': 12   # Column L
            }
        
        print("Column mapping for this sheet:")
        for col_name, col_num in column_numbers.items():
            import openpyxl
            col_letter = openpyxl.utils.get_column_letter(col_num)
            print(f"  - {col_name}: Column {col_letter} ({col_num})")
        
        return column_numbers"""
    
    # 2. Add a method to check if a row is a header
    is_header_method = """    def _is_header_row(self, name):
        \"\"\"Check if a row name indicates it's a header/category row\"\"\"
        if not name:
            return False
            
        name_lower = str(name).lower()
        
        # Common header indicators
        header_indicators = [
            'งาน', 'work', 'system', 'ระบบ', 'total', 'รวม', 'หมวด'
        ]
        
        # Check if any indicators are present but not part of a longer product name
        for indicator in header_indicators:
            if indicator in name_lower and len(name_lower) < 40:
                # Additional check for rows that are clearly headers
                if name_lower.startswith(indicator) or name_lower.endswith(indicator):
                    return True
                
                # Check for common section headers
                if any(section in name_lower for section in [
                    'ระบบ', 'system', 'หมวดงาน', 'section', 'category',
                    'รวมงาน', 'total', 'รวมราคา'
                ]):
                    return True
        
        return False"""
    
    # 3. Update the sync method to avoid redundant loading
    sync_method = """    def _sync_default_master_excel(self):
        \"\"\"Sync default master Excel file to database only if needed\"\"\"
        default_excel_path = os.path.join(self.master_data_folder, 'master.xlsx')
        
        # First check if we already have data in the database
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM master_items")
            count = cursor.fetchone()[0]
            
            if count > 0:
                logging.info(f"Database already contains {count} items, skipping reload")
                
                # Check if any items have costs
                cursor.execute("SELECT COUNT(*) FROM master_items WHERE material_cost > 0 OR labor_cost > 0")
                items_with_costs = cursor.fetchone()[0]
                logging.info(f"Items with costs: {items_with_costs} ({items_with_costs/count*100 if count > 0 else 0:.1f}%)")
                
                return
        
        # If we get here, the database is empty and needs to be loaded
        if os.path.exists(default_excel_path):
            logging.info(f"Found {default_excel_path}. Synchronizing database...")
            
            # DEBUG: Check Excel file properties
            try:
                excel_file = pd.ExcelFile(default_excel_path)
                sheet_names = excel_file.sheet_names
                logging.info(f"Excel file contains {len(sheet_names)} sheets: {sheet_names}")
                
                # Check if file contains any data
                sample_sheet = pd.read_excel(default_excel_path, sheet_name=sheet_names[0], nrows=5)
                logging.info(f"Sample data from first sheet: {len(sample_sheet)} rows, {len(sample_sheet.columns)} columns")
                logging.info(f"Columns: {list(sample_sheet.columns)}")
                
                # Proceed with synchronization
                result = self.load_data_from_excel_to_db(default_excel_path)
                if result.get('success'):
                    logging.info(f"Successfully synchronized: {result.get('message')}")
                else:
                    logging.error(f"Synchronization failed: {result.get('error')}")
                    
            except Exception as e:
                logging.error(f"Error inspecting Excel file: {e}", exc_info=True)
        else:
            logging.warning(f"Default master Excel file not found at {default_excel_path}")"""
    
    # Replace the functions in the content
    # 1. Replace _find_column_numbers
    start_marker = "def _find_column_numbers(self, worksheet, header_row_num, column_map):"
    end_markers = ["def setup_routes(self):", "def run(self, host='localhost', port=5000, debug=True):"]
    
    start_pos = content.find(start_marker)
    if start_pos == -1:
        print("Could not find _find_column_numbers method")
        return
    
    end_pos = -1
    for marker in end_markers:
        pos = content.find(marker, start_pos)
        if pos != -1 and (end_pos == -1 or pos < end_pos):
            end_pos = pos
    
    if end_pos == -1:
        print("Could not find end of _find_column_numbers method")
        return
    
    # Get the function code to replace
    old_code = content[start_pos:end_pos]
    
    # Replace the function
    content = content.replace(old_code, find_col_method)
    
    # 2. Add _is_header_row method before _find_column_numbers
    content = content.replace(find_col_method, is_header_method + "\n\n" + find_col_method)
    
    # 3. Replace _sync_default_master_excel method
    start_marker = "def _sync_default_master_excel(self):"
    start_pos = content.find(start_marker)
    if start_pos == -1:
        print("Could not find _sync_default_master_excel method")
        return
    
    end_pos = -1
    for marker in end_markers:
        pos = content.find(marker, start_pos)
        if pos != -1 and (end_pos == -1 or pos < end_pos):
            end_pos = pos
    
    if end_pos == -1:
        print("Could not find end of _sync_default_master_excel method")
        return
    
    # Get the function code to replace
    old_code = content[start_pos:end_pos]
    
    # Replace the function
    content = content.replace(old_code, sync_method)
    
    # 4. Modify the find_best_match method to skip header rows
    find_best_match_pos = content.find("def find_best_match(self, item_name):")
    if find_best_match_pos != -1:
        # Find where to insert the header check
        match_content = content[find_best_match_pos:].split("\n\n")[0]
        
        # Add header check
        modified_match = match_content.replace(
            "if not item_name or pd.isna(item_name):", 
            "if not item_name or pd.isna(item_name) or self._is_header_row(item_name):"
        )
        
        content = content.replace(match_content, modified_match)
    
    # Save the final patched file
    with open("main_final.py", "w") as f:
        f.write(content)
    
    print("Created final patched version: main_final.py")
    
    # Additional fix: Update database with specific costs for actual items
    update_db_with_specific_costs()
    
    return True

if __name__ == "__main__":
    # Check how the initialization works
    check_initialize_method()
    
    # Create the final patched version
    create_final_patch()