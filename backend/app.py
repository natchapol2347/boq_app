#!/usr/bin/env python3

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import os
import uuid
from datetime import datetime
import logging
from werkzeug.utils import secure_filename
from pathlib import Path
import sqlite3
import shutil
import openpyxl
from typing import Dict, List, Any, Optional

# Add parent directory to path for imports
import sys
sys.path.append(str(Path(__file__).parent.parent))

# Import the specialized processors
from src.processors.interior_sheet_processor import InteriorSheetProcessor
from src.processors.electrical_sheet_processor import ElectricalSheetProcessor
from src.processors.ac_sheet_processor import ACSheetProcessor
from src.processors.fp_sheet_processor import FPSheetProcessor
from src.config.config_manager import ConfigManager
from models.config_models import (
    ProcessorType,
    ConfigUpdateRequest,
    ConfigInquiryResponse,
    ConfigUpdateResponse
)

logging.basicConfig(level=logging.DEBUG)

class App:
    """Main BOQ processor with CRUD API for master data management"""
    
    def __init__(self):
        self.app = Flask(__name__)
        CORS(self.app)
        
        # Setup directories - repo root only
        self.app_root = Path(__file__).parent.parent.absolute()
        
        # Database in repo root data folder
        self.data_dir = self.app_root / 'data'
        os.makedirs(self.data_dir, exist_ok=True)
        self.db_path = str(self.data_dir / 'master_data.db')
        
        # Session management
        self.processing_sessions = {}
        
        # Folder setup - all in repo root
        self.upload_folder = str(self.app_root / 'storage' / 'uploads')
        self.output_folder = str(self.app_root / 'storage' / 'output')
        
        # Create all necessary directories
        folders = [self.upload_folder, self.output_folder, str(self.data_dir)]
        for folder in folders:
            os.makedirs(folder, exist_ok=True)
        
        # Markup rates
        self.markup_rates = {30: 0.30, 50: 0.50, 100: 1.00, 130: 1.30, 150: 1.50}
        
        # Configuration manager
        self.config_manager = ConfigManager()
        
        # Initialize sheet processors with configuration
        configs = self.config_manager.get_all_configs()
        self.sheet_processors = [
            InteriorSheetProcessor(self.db_path, self.markup_rates, configs.interior),
            ElectricalSheetProcessor(self.db_path, self.markup_rates, configs.electrical),
            ACSheetProcessor(self.db_path, self.markup_rates, configs.ac),
            FPSheetProcessor(self.db_path, self.markup_rates, configs.fp)
        ]
        
        # Initialize database (no Excel sync)
        self._init_database()
        
        # Setup Flask routes
        self.setup_routes()
    
    def _init_database(self):
        """Initialize database with all required tables (no Excel sync)"""
        logging.info(f"Initializing database at {self.db_path}")
        
        with sqlite3.connect(self.db_path) as conn:
            # Create tables for each processor
            for processor in self.sheet_processors:
                processor.create_table(conn)
            
            # Log table creation
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in cursor.fetchall()]
            logging.info(f"Database tables created: {tables}")
            
            # Add sample data if tables are empty (only once)
            self._add_sample_data_if_empty(conn)
    
    def _add_sample_data_if_empty(self, conn):
        """Add sample data only if tables are completely empty"""
        cursor = conn.cursor()
        
        # Check if any table has data
        has_data = False
        for processor in self.sheet_processors:
            cursor.execute(f"SELECT COUNT(*) FROM {processor.table_name}")
            count = cursor.fetchone()[0]
            if count > 0:
                has_data = True
                break
        
        if not has_data:
            logging.info("Adding sample data to empty database...")
            self._add_sample_data(conn)
    
    def _add_sample_data(self, conn):
        """Add sample data for testing purposes"""
        cursor = conn.cursor()
        
        # Sample interior items
        interior_samples = [
            ('INT001', 'ปูกระเบื้อง 60x60 ซม.', 450.0, 200.0, 'ตร.ม.'),
            ('INT002', 'ทาสีผนังภายใน', 80.0, 120.0, 'ตร.ม.'),
            ('INT003', 'ติดตั้งประตูไม้', 2500.0, 800.0, 'บาน'),
            ('INT004', 'ติดตั้งหน้าต่างอลูมิเนียม', 1800.0, 600.0, 'ตร.ม.'),
            ('INT005', 'ทำฝ้าเพดานทีบาร์', 320.0, 180.0, 'ตร.ม.')
        ]
        
        for code, name, mat_cost, lab_cost, unit in interior_samples:
            cursor.execute(
                "INSERT OR IGNORE INTO interior_items (internal_id, code, name, material_unit_cost, labor_unit_cost, total_unit_cost, unit) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (f"item_{uuid.uuid4().hex[:8]}", code, name, mat_cost, lab_cost, mat_cost + lab_cost, unit)
            )
        
        # Sample electrical items
        electrical_samples = [
            ('EE001', 'เดินสายไฟ VCT 2x2.5', 35.0, 25.0, 'เมตร'),
            ('EE002', 'ติดตั้งเต้าเสียบ 3 รู', 150.0, 100.0, 'จุด'),
            ('EE003', 'ติดตั้งสวิทช์เปิด-ปิด', 120.0, 80.0, 'จุด'),
            ('EE004', 'ติดตั้งโคมไฟ LED 18W', 380.0, 120.0, 'ดวง'),
            ('EE005', 'ติดตั้งเบรกเกอร์ 32A', 280.0, 150.0, 'ตัว')
        ]
        
        for code, name, mat_cost, lab_cost, unit in electrical_samples:
            cursor.execute(
                "INSERT OR IGNORE INTO ee_items (internal_id, code, name, material_unit_cost, labor_unit_cost, unit) VALUES (?, ?, ?, ?, ?, ?)",
                (f"item_{uuid.uuid4().hex[:8]}", code, name, mat_cost, lab_cost, unit)
            )
        
        # Sample AC items
        ac_samples = [
            ('AC001', 'แอร์ติดผนัง 12000 BTU', 18000.0, 2500.0, 'เครื่อง'),
            ('AC002', 'แอร์ติดผนัง 18000 BTU', 22000.0, 3000.0, 'เครื่อง'),
            ('AC003', 'เดินท่อแอร์ 1/2 นิ้ว', 180.0, 120.0, 'เมตร'),
            ('AC004', 'เดินสายไฟแอร์ 3x2.5', 45.0, 35.0, 'เมตร'),
            ('AC005', 'ติดตั้งรีโมทแอร์', 150.0, 100.0, 'ตัว')
        ]
        
        for code, name, mat_cost, lab_cost, unit in ac_samples:
            cursor.execute(
                "INSERT OR IGNORE INTO ac_items (internal_id, code, name, material_unit_cost, labor_unit_cost, unit) VALUES (?, ?, ?, ?, ?, ?)",
                (f"item_{uuid.uuid4().hex[:8]}", code, name, mat_cost, lab_cost, unit)
            )
        
        # Sample FP items
        fp_samples = [
            ('FP001', 'หัวดับเพลิง Sprinkler', 180.0, 120.0, 'หัว'),
            ('FP002', 'ท่อดับเพลิง 2 นิ้ว', 120.0, 80.0, 'เมตร'),
            ('FP003', 'วาล์วดับเพลิง', 2500.0, 800.0, 'ตัว'),
            ('FP004', 'ตู้เครื่องสูบน้ำ', 25000.0, 5000.0, 'ชุด'),
            ('FP005', 'เซ็นเซอร์ควัน', 450.0, 200.0, 'ตัว')
        ]
        
        for code, name, mat_cost, lab_cost, unit in fp_samples:
            cursor.execute(
                "INSERT OR IGNORE INTO fp_items (internal_id, code, name, material_unit_cost, labor_unit_cost, unit) VALUES (?, ?, ?, ?, ?, ?)",
                (f"item_{uuid.uuid4().hex[:8]}", code, name, mat_cost, lab_cost, unit)
            )
        
        conn.commit()
        logging.info("Sample data added to all tables")
    
    def _find_processor_for_sheet(self, sheet_name: str):
        """Find the appropriate processor for a given sheet name"""
        for processor in self.sheet_processors:
            if processor.matches_sheet(sheet_name):
                return processor
        return None
    
    def _find_processor_by_type(self, processor_type: str):
        """Find processor by type name"""
        type_mapping = {
            'interior': 'InteriorSheetProcessor',
            'electrical': 'ElectricalSheetProcessor', 
            'ac': 'ACSheetProcessor',
            'fp': 'FPSheetProcessor'
        }
        
        target_class = type_mapping.get(processor_type)
        if not target_class:
            return None
            
        for processor in self.sheet_processors:
            if processor.__class__.__name__ == target_class:
                return processor
        return None
    
    def store_processing_session(self, session_id: str, data: Dict[str, Any]):
        """Store processing session data"""
        self.processing_sessions[session_id] = {
            'data': data,
            'created_at': datetime.now()
        }
    
    def _reload_sheet_processors(self):
        """Reload sheet processors with updated configuration"""
        try:
            configs = self.config_manager.get_all_configs()
            self.sheet_processors = [
                InteriorSheetProcessor(self.db_path, self.markup_rates, configs.interior),
                ElectricalSheetProcessor(self.db_path, self.markup_rates, configs.electrical),
                ACSheetProcessor(self.db_path, self.markup_rates, configs.ac),
                FPSheetProcessor(self.db_path, self.markup_rates, configs.fp)
            ]
            logging.info("Sheet processors reloaded with updated configuration")
        except Exception as e:
            logging.error(f"Error reloading sheet processors: {e}", exc_info=True)

    def setup_routes(self):
        """Setup Flask routes including new CRUD endpoints"""
        
        # ========== EXISTING BOQ PROCESSING ROUTES ==========
        
        @self.app.route('/api/process-boq', methods=['POST'])
        def process_boq_route():
            """Process uploaded BOQ file and store matches + section data"""
            if 'file' not in request.files:
                return jsonify({'success': False, 'error': 'No file uploaded'})
            
            file = request.files['file']
            filepath = os.path.join(self.upload_folder, secure_filename(file.filename))
            file.save(filepath)
            
            try:
                excel_file = pd.ExcelFile(filepath)
                session_data = {'sheets': {}, 'original_filepath': filepath}
                
                sheets_to_process = excel_file.sheet_names
                total_items = 0
                total_matches = 0
                
                for sheet_name in sheets_to_process:
                    processor = self._find_processor_for_sheet(sheet_name)
                    if not processor:
                        logging.info(f"No processor found for sheet: {sheet_name} - skipping")
                        continue
                    
                    logging.info(f"Processing BOQ sheet: {sheet_name} with {processor.__class__.__name__}")
                    
                    df = pd.read_excel(filepath, sheet_name=sheet_name, header=processor.header_row)
                    processed_items = processor.process_boq_sheet(df)
                    
                    try:
                        temp_workbook = openpyxl.load_workbook(filepath, data_only=False)
                        temp_worksheet = temp_workbook[sheet_name]
                        sections = processor.find_section_structure(temp_worksheet, temp_worksheet.max_row)
                        temp_workbook.close()
                        logging.info(f"Pre-calculated {len(sections)} sections for {sheet_name}")
                    except Exception as e:
                        logging.warning(f"Could not pre-calculate sections for {sheet_name}: {e}")
                        sections = {}
                    
                    session_data['sheets'][sheet_name] = {
                        'processor_type': processor.__class__.__name__,
                        'header_row': processor.header_row,
                        'processed_matches': {item['original_row_index']: item['match'] for item in processed_items},
                        'row_details': {item['original_row_index']: {'code': item['row_code'], 'name': item['row_name']} for item in processed_items},
                        'sections': sections,
                        'total_rows': len(df),
                        'matched_count': len(processed_items)
                    }
                    
                    total_items += len(df)
                    total_matches += len(processed_items)
                
                session_id = str(uuid.uuid4())
                self.store_processing_session(session_id, session_data)
                
                return jsonify({
                    'success': True,
                    'session_id': session_id,
                    'summary': {
                        'total_items': total_items,
                        'matched_items': total_matches,
                        'match_rate': (total_matches / total_items * 100) if total_items > 0 else 0,
                        'sheets_processed': len(session_data['sheets'])
                    }
                })
                
            except Exception as e:
                logging.error(f"Error processing BOQ file: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/generate-final-boq', methods=['POST'])
        def generate_final_boq_route():
            """Generate final BOQ with calculated costs"""
            data = request.get_json()
            session_id = data.get('session_id')
            
            if not session_id or session_id not in self.processing_sessions:
                return jsonify({'success': False, 'error': 'Invalid session'})
            
            session_data = self.processing_sessions[session_id]['data']
            original_filepath = session_data['original_filepath']
            markup_options = data.get('markup_options', [30, 50, 100, 130, 150])
            
            try:
                filename = f"final_boq_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                output_filepath = os.path.join(self.output_folder, filename)
                shutil.copy(original_filepath, output_filepath)
                
                workbook = openpyxl.load_workbook(output_filepath)
                data_workbook = openpyxl.load_workbook(original_filepath, data_only=True)
                
                items_processed = 0
                items_failed = 0
                processing_summary = {}
                
                for sheet_name, sheet_info in session_data['sheets'].items():
                    if sheet_name not in workbook.sheetnames:
                        continue
                    
                    processor = self._find_processor_for_sheet(sheet_name)
                    if not processor:
                        logging.warning(f"No processor found for sheet: {sheet_name}")
                        continue
                    
                    logging.info(f"Generating costs for sheet: {sheet_name}")
                    
                    sheet_result = processor.process_final_sheet(
                        worksheet=workbook[sheet_name], 
                        data_worksheet=data_workbook[sheet_name],
                        sheet_info=sheet_info,
                        markup_options=markup_options
                    )
                    
                    items_processed += sheet_result['items_processed']
                    items_failed += sheet_result['items_failed']
                    processing_summary[sheet_name] = sheet_result
                
                workbook.save(output_filepath)
                workbook.close()
                data_workbook.close()
                
                logging.info(f"Processing complete: {items_processed} items processed, {items_failed} failed")
                
                return jsonify({
                    'success': True,
                    'filename': filename,
                    'download_url': f'/api/download/{filename}',
                    'items_processed': items_processed,
                    'items_failed': items_failed,
                    'processing_summary': processing_summary
                })
                
            except Exception as e:
                logging.error(f"Error generating final BOQ: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/apply-markup', methods=['POST'])
        def apply_markup_route():
            """Apply markup directly to all values in all sheets"""
            data = request.get_json()
            session_id = data.get('session_id')
            markup_percent = data.get('markup_percent')
            
            if not session_id or session_id not in self.processing_sessions:
                return jsonify({'success': False, 'error': 'Invalid session'})
            
            if markup_percent is None or not isinstance(markup_percent, (int, float)):
                return jsonify({'success': False, 'error': 'markup_percent must be a valid number'})
            
            session_data = self.processing_sessions[session_id]['data']
            original_filepath = session_data['original_filepath']
            
            try:
                original_name = os.path.splitext(os.path.basename(original_filepath))[0]
                filename = f"{markup_percent}%_{original_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                output_filepath = os.path.join(self.output_folder, filename)
                shutil.copy(original_filepath, output_filepath)
                
                workbook = openpyxl.load_workbook(output_filepath)
                data_workbook = openpyxl.load_workbook(original_filepath, data_only=True)
                
                items_processed = 0
                items_failed = 0
                processing_summary = {}
                
                for sheet_name, sheet_info in session_data['sheets'].items():
                    if sheet_name not in workbook.sheetnames:
                        continue
                    
                    processor = self._find_processor_for_sheet(sheet_name)
                    if not processor:
                        logging.warning(f"No processor found for sheet: {sheet_name}")
                        continue
                    
                    logging.info(f"Applying {markup_percent}% markup to sheet: {sheet_name}")
                    
                    sheet_result = processor.process_final_sheet(
                        worksheet=workbook[sheet_name], 
                        data_worksheet=data_workbook[sheet_name],
                        sheet_info=sheet_info,
                        markup_options=[],
                        apply_markup_percent=markup_percent
                    )
                    
                    items_processed += sheet_result['items_processed']
                    items_failed += sheet_result['items_failed']
                    processing_summary[sheet_name] = sheet_result
                
                workbook.save(output_filepath)
                workbook.close()
                data_workbook.close()
                
                logging.info(f"Markup application complete: {markup_percent}% applied to {items_processed} items, {items_failed} failed")
                
                return jsonify({
                    'success': True,
                    'filename': filename,
                    'download_url': f'/api/download/{filename}',
                    'markup_percent': markup_percent,
                    'items_processed': items_processed,
                    'items_failed': items_failed,
                    'processing_summary': processing_summary
                })
                
            except Exception as e:
                logging.error(f"Error applying markup: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/cleanup-session', methods=['POST'])
        def cleanup_session_route():
            """Cleanup session data and delete associated files"""
            data = request.get_json()
            session_id = data.get('session_id')
            
            if not session_id:
                return jsonify({'success': False, 'error': 'session_id is required'})
            
            files_deleted = []
            errors = []
            
            try:
                if session_id in self.processing_sessions:
                    session_data = self.processing_sessions[session_id]['data']
                    original_filepath = session_data.get('original_filepath')
                    
                    if original_filepath and os.path.exists(original_filepath):
                        try:
                            os.remove(original_filepath)
                            files_deleted.append(original_filepath)
                            logging.info(f"Deleted original file: {original_filepath}")
                        except Exception as e:
                            errors.append(f"Failed to delete {original_filepath}: {e}")
                    
                    del self.processing_sessions[session_id]
                    logging.info(f"Cleaned up session: {session_id}")
                else:
                    return jsonify({'success': False, 'error': 'Invalid session_id'})
                
                if os.path.exists(self.output_folder):
                    for output_file in os.listdir(self.output_folder):
                        output_path = os.path.join(self.output_folder, output_file)
                        if os.path.isfile(output_path):
                            try:
                                os.remove(output_path)
                                files_deleted.append(output_path)
                                logging.info(f"Deleted output file: {output_path}")
                            except Exception as e:
                                errors.append(f"Failed to delete {output_path}: {e}")
                
                return jsonify({
                    'success': True,
                    'session_cleaned': True,
                    'files_deleted': len(files_deleted),
                    'deleted_files': files_deleted,
                    'errors': errors
                })
                
            except Exception as e:
                logging.error(f"Error during cleanup: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        # ========== NEW MASTER DATA CRUD ROUTES ==========
        
        @self.app.route('/api/master-data/list/<processor_type>', methods=['GET'])
        def list_master_data(processor_type):
            """List all master data items for a specific processor type"""
            try:
                processor = self._find_processor_by_type(processor_type)
                if not processor:
                    return jsonify({'success': False, 'error': f'Invalid processor type: {processor_type}'})
                
                with sqlite3.connect(self.db_path) as conn:
                    conn.row_factory = sqlite3.Row
                    cursor = conn.cursor()
                    cursor.execute(f"SELECT * FROM {processor.table_name} ORDER BY code, name")
                    items = [dict(row) for row in cursor.fetchall()]
                
                return jsonify({
                    'success': True,
                    'processor_type': processor_type,
                    'table_name': processor.table_name,
                    'items': items,
                    'count': len(items)
                })
                
            except Exception as e:
                logging.error(f"Error listing master data: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/master-data/get/<processor_type>/<item_id>', methods=['GET'])
        def get_master_data_item(processor_type, item_id):
            """Get a specific master data item"""
            try:
                processor = self._find_processor_by_type(processor_type)
                if not processor:
                    return jsonify({'success': False, 'error': f'Invalid processor type: {processor_type}'})
                
                with sqlite3.connect(self.db_path) as conn:
                    conn.row_factory = sqlite3.Row
                    cursor = conn.cursor()
                    cursor.execute(f"SELECT * FROM {processor.table_name} WHERE internal_id = ?", (item_id,))
                    item = cursor.fetchone()
                
                if item:
                    return jsonify({
                        'success': True,
                        'item': dict(item)
                    })
                else:
                    return jsonify({'success': False, 'error': 'Item not found'})
                    
            except Exception as e:
                logging.error(f"Error getting master data item: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/master-data/create/<processor_type>', methods=['POST'])
        def create_master_data_item(processor_type):
            """Create a new master data item"""
            try:
                processor = self._find_processor_by_type(processor_type)
                if not processor:
                    return jsonify({'success': False, 'error': f'Invalid processor type: {processor_type}'})
                
                data = request.get_json()
                
                # Generate internal ID
                internal_id = f"item_{uuid.uuid4().hex[:8]}"
                
                # Validate required fields
                if not data.get('name'):
                    return jsonify({'success': False, 'error': 'Name is required'})
                
                with sqlite3.connect(self.db_path) as conn:
                    cursor = conn.cursor()
                    
                    if processor_type == 'interior':
                        cursor.execute(
                            f"INSERT INTO {processor.table_name} (internal_id, code, name, material_unit_cost, labor_unit_cost, total_unit_cost, unit) VALUES (?, ?, ?, ?, ?, ?, ?)",
                            (
                                internal_id,
                                data.get('code', ''),
                                data.get('name'),
                                float(data.get('material_unit_cost', 0)),
                                float(data.get('labor_unit_cost', 0)),
                                float(data.get('material_unit_cost', 0)) + float(data.get('labor_unit_cost', 0)),
                                data.get('unit', '')
                            )
                        )
                    else:
                        cursor.execute(
                            f"INSERT INTO {processor.table_name} (internal_id, code, name, material_unit_cost, labor_unit_cost, unit) VALUES (?, ?, ?, ?, ?, ?)",
                            (
                                internal_id,
                                data.get('code', ''),
                                data.get('name'),
                                float(data.get('material_unit_cost', 0)),
                                float(data.get('labor_unit_cost', 0)),
                                data.get('unit', '')
                            )
                        )
                    
                    conn.commit()
                
                return jsonify({
                    'success': True,
                    'message': 'Item created successfully',
                    'internal_id': internal_id
                })
                
            except Exception as e:
                logging.error(f"Error creating master data item: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/master-data/update/<processor_type>/<item_id>', methods=['PUT'])
        def update_master_data_item(processor_type, item_id):
            """Update an existing master data item"""
            try:
                processor = self._find_processor_by_type(processor_type)
                if not processor:
                    return jsonify({'success': False, 'error': f'Invalid processor type: {processor_type}'})
                
                data = request.get_json()
                
                # Validate required fields
                if not data.get('name'):
                    return jsonify({'success': False, 'error': 'Name is required'})
                
                with sqlite3.connect(self.db_path) as conn:
                    cursor = conn.cursor()
                    
                    # Check if item exists
                    cursor.execute(f"SELECT internal_id FROM {processor.table_name} WHERE internal_id = ?", (item_id,))
                    if not cursor.fetchone():
                        return jsonify({'success': False, 'error': 'Item not found'})
                    
                    if processor_type == 'interior':
                        total_unit_cost = float(data.get('material_unit_cost', 0)) + float(data.get('labor_unit_cost', 0))
                        cursor.execute(
                            f"UPDATE {processor.table_name} SET code = ?, name = ?, material_unit_cost = ?, labor_unit_cost = ?, total_unit_cost = ?, unit = ? WHERE internal_id = ?",
                            (
                                data.get('code', ''),
                                data.get('name'),
                                float(data.get('material_unit_cost', 0)),
                                float(data.get('labor_unit_cost', 0)),
                                total_unit_cost,
                                data.get('unit', ''),
                                item_id
                            )
                        )
                    else:
                        cursor.execute(
                            f"UPDATE {processor.table_name} SET code = ?, name = ?, material_unit_cost = ?, labor_unit_cost = ?, unit = ? WHERE internal_id = ?",
                            (
                                data.get('code', ''),
                                data.get('name'),
                                float(data.get('material_unit_cost', 0)),
                                float(data.get('labor_unit_cost', 0)),
                                data.get('unit', ''),
                                item_id
                            )
                        )
                    
                    conn.commit()
                
                return jsonify({
                    'success': True,
                    'message': 'Item updated successfully'
                })
                
            except Exception as e:
                logging.error(f"Error updating master data item: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/master-data/delete/<processor_type>/<item_id>', methods=['DELETE'])
        def delete_master_data_item(processor_type, item_id):
            """Delete a master data item"""
            try:
                processor = self._find_processor_by_type(processor_type)
                if not processor:
                    return jsonify({'success': False, 'error': f'Invalid processor type: {processor_type}'})
                
                with sqlite3.connect(self.db_path) as conn:
                    cursor = conn.cursor()
                    
                    # Check if item exists
                    cursor.execute(f"SELECT internal_id FROM {processor.table_name} WHERE internal_id = ?", (item_id,))
                    if not cursor.fetchone():
                        return jsonify({'success': False, 'error': 'Item not found'})
                    
                    # Delete the item
                    cursor.execute(f"DELETE FROM {processor.table_name} WHERE internal_id = ?", (item_id,))
                    conn.commit()
                
                return jsonify({
                    'success': True,
                    'message': 'Item deleted successfully'
                })
                
            except Exception as e:
                logging.error(f"Error deleting master data item: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/master-data/bulk-import/<processor_type>', methods=['POST'])
        def bulk_import_master_data(processor_type):
            """Bulk import master data from uploaded Excel file"""
            try:
                processor = self._find_processor_by_type(processor_type)
                if not processor:
                    return jsonify({'success': False, 'error': f'Invalid processor type: {processor_type}'})
                
                if 'file' not in request.files:
                    return jsonify({'success': False, 'error': 'No file uploaded'})
                
                file = request.files['file']
                filepath = os.path.join(self.upload_folder, secure_filename(file.filename))
                file.save(filepath)
                
                # Read Excel file
                df = pd.read_excel(filepath, header=0)
                
                imported_count = 0
                errors = []
                
                with sqlite3.connect(self.db_path) as conn:
                    cursor = conn.cursor()
                    
                    for idx, row in df.iterrows():
                        try:
                            internal_id = f"import_{uuid.uuid4().hex[:8]}"
                            
                            if processor_type == 'interior':
                                mat_cost = float(row.get('material_unit_cost', 0))
                                lab_cost = float(row.get('labor_unit_cost', 0))
                                cursor.execute(
                                    f"INSERT INTO {processor.table_name} (internal_id, code, name, material_unit_cost, labor_unit_cost, total_unit_cost, unit) VALUES (?, ?, ?, ?, ?, ?, ?)",
                                    (
                                        internal_id,
                                        str(row.get('code', '')),
                                        str(row.get('name', '')),
                                        mat_cost,
                                        lab_cost,
                                        mat_cost + lab_cost,
                                        str(row.get('unit', ''))
                                    )
                                )
                            else:
                                cursor.execute(
                                    f"INSERT INTO {processor.table_name} (internal_id, code, name, material_unit_cost, labor_unit_cost, unit) VALUES (?, ?, ?, ?, ?, ?)",
                                    (
                                        internal_id,
                                        str(row.get('code', '')),
                                        str(row.get('name', '')),
                                        float(row.get('material_unit_cost', 0)),
                                        float(row.get('labor_unit_cost', 0)),
                                        str(row.get('unit', ''))
                                    )
                                )
                            
                            imported_count += 1
                            
                        except Exception as e:
                            errors.append(f"Row {idx + 2}: {str(e)}")
                    
                    conn.commit()
                
                # Clean up uploaded file
                os.remove(filepath)
                
                return jsonify({
                    'success': True,
                    'message': f'Bulk import completed',
                    'imported_count': imported_count,
                    'errors': errors
                })
                
            except Exception as e:
                logging.error(f"Error in bulk import: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/master-data/export/<processor_type>', methods=['GET'])
        def export_master_data(processor_type):
            """Export master data to Excel file"""
            try:
                processor = self._find_processor_by_type(processor_type)
                if not processor:
                    return jsonify({'success': False, 'error': f'Invalid processor type: {processor_type}'})
                
                with sqlite3.connect(self.db_path) as conn:
                    df = pd.read_sql_query(f"SELECT * FROM {processor.table_name}", conn)
                
                if df.empty:
                    return jsonify({'success': False, 'error': 'No data to export'})
                
                # Generate export filename
                filename = f"{processor_type}_master_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                filepath = os.path.join(self.output_folder, filename)
                
                # Export to Excel
                df.to_excel(filepath, index=False)
                
                return jsonify({
                    'success': True,
                    'filename': filename,
                    'download_url': f'/api/download/{filename}',
                    'exported_count': len(df)
                })
                
            except Exception as e:
                logging.error(f"Error exporting master data: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        # ========== EXISTING CONFIGURATION ROUTES ==========
        
        @self.app.route('/api/config/inquiry', methods=['GET'])
        def config_inquiry_route():
            """Get current processor configurations"""
            try:
                configs = self.config_manager.get_all_configs()
                summary = self.config_manager.get_config_summary()
                
                return ConfigInquiryResponse(
                    success=True,
                    configs=configs
                ).model_dump()
                
            except Exception as e:
                logging.error(f"Error getting config: {e}", exc_info=True)
                return ConfigInquiryResponse(
                    success=False,
                    error=str(e)
                ).model_dump()
        
        @self.app.route('/api/config/update', methods=['POST'])
        def config_update_route():
            """Update processor configuration"""
            try:
                data = request.get_json()
                
                update_request = ConfigUpdateRequest(**data)
                success = self.config_manager.update_config(update_request)
                
                if success:
                    self._reload_sheet_processors()
                    
                    return ConfigUpdateResponse(
                        success=True,
                        message="Configuration updated successfully",
                        updated_processor=update_request.processor_name.value
                    ).model_dump()
                else:
                    return ConfigUpdateResponse(
                        success=False,
                        message="Failed to update configuration",
                        error="Update operation failed"
                    ).model_dump()
                    
            except Exception as e:
                logging.error(f"Error updating config: {e}", exc_info=True)
                return ConfigUpdateResponse(
                    success=False,
                    message="Configuration update failed",
                    error=str(e)
                ).model_dump()
        
        @self.app.route('/api/download/<filename>')
        def download_file(filename):
            """Download generated BOQ file"""
            filepath = os.path.join(self.output_folder, filename)
            if os.path.exists(filepath):
                return send_file(filepath, as_attachment=True)
            return jsonify({'error': 'File not found'}), 404
    
    def run(self, host: str = 'localhost', port: int = 5000, debug: bool = True):
        """Run the Flask application"""
        logging.info(f"BOQ Processing Server with CRUD API starting on http://{host}:{port}")
        self.app.run(host=host, port=port, debug=debug)

if __name__ == '__main__':
    processor = App()
    processor.run()