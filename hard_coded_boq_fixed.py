#!/usr/bin/env python3
# Hard-coded BOQ Processor with fixes for:
# 1. Total cost calculation
# 2. Duplicate item name handling (using code + name)
# 3. Better data sanitization for non-interior sheets

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import numpy as np
from fuzzywuzzy import fuzz
import os
import uuid
from datetime import datetime
import re
from werkzeug.utils import secure_filename
from pathlib import Path
import sqlite3
import logging
import shutil
import openpyxl

logging.basicConfig(level=logging.INFO)

class HardCodedBOQProcessor:
    def __init__(self):
        self.app = Flask(__name__)
        CORS(self.app)
        
        # --- Database and Folder Setup ---
        self.data_dir = Path.home() / 'AppData' / 'Roaming' / 'BOQProcessor'
        os.makedirs(self.data_dir, exist_ok=True)
        self.db_path = self.data_dir / 'master_data.db'
        
        # Session management for processing sessions
        self.processing_sessions = {}
        
        self.master_data_folder = 'master_data'
        self.upload_folder = 'uploads'
        self.output_folder = 'output'
        os.makedirs(self.master_data_folder, exist_ok=True)
        os.makedirs(self.upload_folder, exist_ok=True)
        os.makedirs(self.output_folder, exist_ok=True)

        # Hardcoded markup rates
        self.markup_rates = {100: 1.00, 130: 1.30, 150: 1.50, 50: 0.50, 30: 0.30}
        
        # Sheet type definitions
        self.sheet_types = {
            'INT': {
                'pattern': 'int',
                'header_row': 9,  # 0-based
                'columns': {
                    'code': 2,          # Column B
                    'name': 3,          # Column C
                    'quantity': 4,      # Column D
                    'unit': 5,          # Column E
                    'material_cost': 6, # Column F
                    'labor_cost': 7,    # Column G
                    'total_cost': 8     # Column H
                },
                'table_name': 'interior_items'
            },
            'EE': {
                'pattern': 'ee',
                'header_row': 7,  # 0-based
                'columns': {
                    'code': 2,          # Column B
                    'name': 3,          # Column C
                    'unit': 6,          # Column F
                    'quantity': 7,      # Column G
                    'material_cost': 8, # Column H
                    'labor_cost': 10,   # Column J
                    'total_cost': 12    # Column L
                },
                'table_name': 'ee_items'
            },
            'AC': {
                'pattern': 'ac',
                'header_row': 5,  # 0-based
                'columns': {
                    'code': 2,          # Column B
                    'name': 3,          # Column C
                    'unit': 6,          # Column F
                    'quantity': 7,      # Column G
                    'material_cost': 8, # Column H
                    'labor_cost': 10,   # Column J
                    'total_cost': 12    # Column L
                },
                'table_name': 'ac_items'
            },
            'FP': {
                'pattern': 'fp',
                'header_row': 7,  # 0-based
                'columns': {
                    'code': 2,          # Column B
                    'name': 3,          # Column C
                    'unit': 6,          # Column F
                    'quantity': 7,      # Column G
                    'material_cost': 8, # Column H
                    'labor_cost': 10,   # Column J
                    'total_cost': 12    # Column L
                },
                'table_name': 'fp_items'
            },
            'DEFAULT': {
                'pattern': '',
                'header_row': 8,  # 0-based
                'columns': {
                    'code': 2,          # Column B
                    'name': 3,          # Column C
                    'quantity': 4,      # Column D
                    'unit': 5,          # Column E
                    'material_cost': 6, # Column F
                    'labor_cost': 7,    # Column G
                    'total_cost': 8     # Column H
                },
                'table_name': 'default_items'
            }
        }
        
        self._init_db()
        self._sync_default_master_excel()
        self.setup_routes()

    def _init_db(self):
        """Initialize SQLite database with separate tables for each sheet type"""
        logging.info(f"Initializing database at {self.db_path}")
        
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Create a table for each sheet type
                for sheet_type, config in self.sheet_types.items():
                    table_name = config['table_name']
                    cursor.execute(f'''
                        CREATE TABLE IF NOT EXISTS {table_name} (
                            internal_id TEXT PRIMARY KEY, 
                            code TEXT, 
                            name TEXT NOT NULL,
                            material_cost REAL DEFAULT 0, 
                            labor_cost REAL DEFAULT 0, 
                            total_cost REAL DEFAULT 0,
                            unit TEXT
                        )
                    ''')
                conn.commit()
                
                # Debug: Check tables
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                tables = cursor.fetchall()
                logging.info(f"Database tables: {tables}")
                
                for sheet_type, config in self.sheet_types.items():
                    table_name = config['table_name']
                    # Check if table has any data
                    count = cursor.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
                    logging.info(f"Table {table_name} contains {count} items")
                    
                    # Check items with costs
                    if count > 0:
                        items_with_costs = cursor.execute(
                            f"SELECT COUNT(*) FROM {table_name} WHERE material_cost > 0 OR labor_cost > 0"
                        ).fetchone()[0]
                        logging.info(f"Items with costs in {table_name}: {items_with_costs} ({items_with_costs/count*100:.1f}%)")
                
                logging.info(f"Database initialized successfully at {self.db_path}")
                
        except Exception as e:
            logging.error(f"Error initializing database: {e}", exc_info=True)
            raise

    def _sync_default_master_excel(self):
        """Sync default master Excel file to database with sheet-specific tables"""
        default_excel_path = os.path.join(self.master_data_folder, 'master.xlsx')
        
        if os.path.exists(default_excel_path):
            logging.info(f"Found {default_excel_path}. Synchronizing database...")
            
            try:
                excel_file = pd.ExcelFile(default_excel_path)
                sheet_names = excel_file.sheet_names
                logging.info(f"Excel file contains {len(sheet_names)} sheets: {sheet_names}")
                
                # Process each sheet
                for sheet_name in sheet_names:
                    if "sum" in sheet_name.lower():
                        continue
                        
                    # Determine sheet type
                    sheet_type = 'DEFAULT'
                    for key, config in self.sheet_types.items():
                        if config['pattern'] and config['pattern'].lower() in sheet_name.lower():
                            sheet_type = key
                            break
                    
                    logging.info(f"Processing sheet {sheet_name} as type {sheet_type}")
                    
                    # Get the sheet configuration
                    config = self.sheet_types[sheet_type]
                    header_row = config['header_row']
                    table_name = config['table_name']
                    
                    # Read sheet with fixed header row
                    df = pd.read_excel(default_excel_path, sheet_name=sheet_name, header=header_row)
                    
                    # Process the dataframe using fixed column positions
                    processed_df = self._process_master_sheet(df, sheet_type)
                    
                    if not processed_df.empty:
                        # Insert into the appropriate table
                        with sqlite3.connect(self.db_path) as conn:
                            # Clear existing data for this sheet type
                            conn.execute(f"DELETE FROM {table_name}")
                            
                            # Insert new data
                            for _, row in processed_df.iterrows():
                                try:
                                    # Check if a similar item already exists
                                    cursor = conn.cursor()
                                    cursor.execute(
                                        f"SELECT internal_id FROM {table_name} WHERE name = ? AND code = ?",
                                        (row['name'], row['code'])
                                    )
                                    existing = cursor.fetchone()
                                    
                                    if existing:
                                        # Log database duplicate to alert the user
                                        logging.warning(f"⚠️ DATABASE DUPLICATE: Code='{row['code']}', Name='{row['name']}' in table {table_name}")
                                        print(f"⚠️ DATABASE DUPLICATE: Code='{row['code']}', Name='{row['name']}' in table {table_name}")
                                        
                                        # Update existing item with non-zero costs
                                        if row['material_cost'] > 0 or row['labor_cost'] > 0:
                                            conn.execute(
                                                f"UPDATE {table_name} SET material_cost = ?, labor_cost = ?, total_cost = ? WHERE internal_id = ?",
                                                (
                                                    row['material_cost'],
                                                    row['labor_cost'],
                                                    row['total_cost'],
                                                    existing[0]
                                                )
                                            )
                                            logging.info(f"  → Updated costs for database duplicate: Material={row['material_cost']}, Labor={row['labor_cost']}")
                                    else:
                                        # Insert new item
                                        conn.execute(
                                            f"INSERT INTO {table_name} (internal_id, code, name, material_cost, labor_cost, total_cost, unit) VALUES (?, ?, ?, ?, ?, ?, ?)",
                                            (
                                                row['internal_id'],
                                                row['code'],
                                                row['name'],
                                                row['material_cost'],
                                                row['labor_cost'],
                                                row['total_cost'],
                                                row.get('unit', '')
                                            )
                                        )
                                except sqlite3.IntegrityError as e:
                                    # Detailed error logging for database integrity errors
                                    error_msg = f"DATABASE INTEGRITY ERROR: {e}"
                                    logging.error(f"🚨 {error_msg}")
                                    print(f"🚨 {error_msg}")
                                    logging.error(f"  → Item: Code='{row['code']}', Name='{row['name']}' in table {table_name}")
                                    print(f"  → Skipping duplicate item: {row['name']} (code: {row['code']})")
                            conn.commit()
                            
                        logging.info(f"Synchronized {len(processed_df)} items into {table_name}")
                        
                        # Log duplicate summary for this table
                        cursor = conn.cursor()
                        cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                        total_items = cursor.fetchone()[0]
                        
                        cursor.execute(f"SELECT COUNT(name), name, code FROM {table_name} GROUP BY name, code HAVING COUNT(name) > 1")
                        duplicates = cursor.fetchall()
                        
                        if duplicates:
                            logging.warning(f"⚠️ Found {len(duplicates)} duplicate items in {table_name} table")
                            for dup in duplicates:
                                logging.warning(f"  → Duplicate: Count={dup[0]}, Name='{dup[1]}', Code='{dup[2]}'")
                        else:
                            logging.info(f"✅ No duplicate items found in {table_name} table")
                
                # Add sample costs if no costs are found
                self._ensure_costs_exist()
                        
            except Exception as e:
                logging.error(f"Error synchronizing master data: {e}", exc_info=True)
        else:
            logging.warning(f"Default master Excel file not found at {default_excel_path}")
            
            # Add sample data if no master data is available
            self._add_sample_data()

    def _process_master_sheet(self, df, sheet_type):
        """Process a master data sheet using fixed column positions with improved handling of item names"""
        if df.empty:
            return pd.DataFrame()
        
        # Get column configuration for this sheet type
        config = self.sheet_types[sheet_type]
        column_mapping = config['columns']
        
        # Create a new dataframe with selected columns
        result_data = []
        # Keep track of processed items to handle duplicates
        processed_items = {}
        
        for idx, row in df.iterrows():
            try:
                # Get values from fixed positions
                code_idx = column_mapping['code'] - 1  # Convert to 0-based
                name_idx = column_mapping['name'] - 1
                material_idx = column_mapping['material_cost'] - 1
                labor_idx = column_mapping['labor_cost'] - 1
                unit_idx = column_mapping['unit'] - 1 if 'unit' in column_mapping else None
                
                # Get values (safely)
                if idx >= len(df):
                    continue
                    
                row_values = row.values
                if len(row_values) <= max(code_idx, name_idx, material_idx, labor_idx):
                    continue
                
                code = str(row_values[code_idx]) if code_idx < len(row_values) else ''
                name = str(row_values[name_idx]) if name_idx < len(row_values) else ''
                
                # For non-Interior sheets: check if name is "-" and get description from next column
                # This ensures we store the proper item name in the database from column D
                if sheet_type in ["EE", "AC", "FP"] and (name == "-" or name.strip() == "-") and name_idx + 1 < len(row_values):
                    description = str(row_values[name_idx + 1]).strip()
                    if description and description != "nan" and description != "-":
                        print(f"  Found description in master sheet: '{description}' for {sheet_type} item with code '{code}'")
                        name = description  # Use the description as the name
                
                # IMPROVED: Clean item name to handle '-' or empty names and non-Interior sheets
                name = self._clean_item_name(name, code, sheet_type)
                if not name:
                    continue
                
                # Clean and convert cost values
                try:
                    material_cost = float(row_values[material_idx]) if material_idx < len(row_values) and pd.notna(row_values[material_idx]) else 0
                except (ValueError, TypeError):
                    material_cost = 0
                    
                try:
                    labor_cost = float(row_values[labor_idx]) if labor_idx < len(row_values) and pd.notna(row_values[labor_idx]) else 0
                except (ValueError, TypeError):
                    labor_cost = 0
                
                # Get unit if available
                unit = str(row_values[unit_idx]) if unit_idx is not None and unit_idx < len(row_values) else ''
                
                # Skip empty or total rows
                if not name or name.lower() in ['nan', 'none', '']:
                    continue
                    
                if any(keyword in name.lower() for keyword in ['total', 'รวม', 'sum', 'subtotal']):
                    continue
                
                # Handle duplicates in the same sheet
                item_key = f"{code}|{name}"
                if item_key in processed_items:
                    # Log duplicate items to alert the user
                    logging.warning(f"⚠️ DUPLICATE ITEM DETECTED: Code='{code}', Name='{name}' in sheet type {sheet_type}")
                    print(f"⚠️ DUPLICATE ITEM: Code='{code}', Name='{name}' in sheet type {sheet_type}")
                    
                    # Only update if this item has costs and the previous one doesn't
                    existing_item = processed_items[item_key]
                    if (material_cost > 0 or labor_cost > 0) and (existing_item['material_cost'] == 0 and existing_item['labor_cost'] == 0):
                        existing_item['material_cost'] = material_cost
                        existing_item['labor_cost'] = labor_cost
                        existing_item['total_cost'] = material_cost + labor_cost
                        logging.info(f"  → Updated costs for duplicate item: Material={material_cost}, Labor={labor_cost}")
                    continue
                
                # Create item
                total_cost = material_cost + labor_cost
                item = {
                    'internal_id': f"item_{uuid.uuid4().hex[:8]}",
                    'code': code,
                    'name': name,
                    'material_cost': material_cost,
                    'labor_cost': labor_cost,
                    'total_cost': total_cost,
                    'unit': unit
                }
                
                processed_items[item_key] = item
                result_data.append(item)
                
                # Debug costs
                if material_cost > 0 or labor_cost > 0:
                    print(f"Found item with costs: {name[:30]}... -> Material: {material_cost}, Labor: {labor_cost}")
                    
            except Exception as e:
                print(f"Error processing row {idx}: {e}")
                continue
        
        if not result_data:
            return pd.DataFrame()
            
        result_df = pd.DataFrame(result_data)
        print(f"Processed {len(result_df)} items from sheet type {sheet_type}")
        print(f"Items with costs: {len(result_df[(result_df['material_cost'] > 0) | (result_df['labor_cost'] > 0)])}")
        
        return result_df

    def _clean_item_name(self, name, code, sheet_type=None):
        """Clean and improve item names, especially for '-' values and non-Interior sheets"""
        if not name or pd.isna(name) or name.strip() in ['-', '', 'nan', 'none']:
            if code and code.strip() and code.strip() not in ['-', 'nan', 'none']:
                # Use code as name if name is empty/invalid but code exists
                item_prefix = "" if sheet_type == "INT" else ""  # No prefix for any sheet type
                return f"{item_prefix}{code.strip()}"
            else:
                # Generate a unique name if both name and code are invalid
                return f"Item_{uuid.uuid4().hex[:6]}"
                
        # Clean up the name
        cleaned = name.strip()
        if cleaned == '-':
            return f"Unnamed item {uuid.uuid4().hex[:6]}"
        
        # For non-interior sheets, remove "Item " prefix if present
        if sheet_type and sheet_type != "INT" and cleaned.lower().startswith("item "):
            cleaned = cleaned[5:].strip()
            print(f"  Cleaned item name: Removed 'Item ' prefix -> '{cleaned}'")
        
        return cleaned

    def _ensure_costs_exist(self):
        """Make sure all tables have at least some items with costs"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            for sheet_type, config in self.sheet_types.items():
                table_name = config['table_name']
                
                # Check if table has costs
                cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE material_cost > 0 OR labor_cost > 0")
                count = cursor.fetchone()[0]
                
                if count == 0:
                    logging.info(f"No costs found in {table_name}, adding sample costs")
                    
                    # Add sample costs to this table
                    cursor.execute(f"UPDATE {table_name} SET material_cost = 500, labor_cost = 300, total_cost = 800")
                    conn.commit()
                    
                    cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE material_cost > 0")
                    updated = cursor.fetchone()[0]
                    logging.info(f"Added sample costs to {updated} items in {table_name}")

    def _add_sample_data(self):
        """Add sample data to the database if no master data is available"""
        logging.info("Adding sample data to database")
        
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Add sample items to each table
            for sheet_type, config in self.sheet_types.items():
                table_name = config['table_name']
                
                # Check if table already has data
                cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                count = cursor.fetchone()[0]
                
                if count > 0:
                    logging.info(f"Table {table_name} already has {count} items, skipping")
                    continue
                
                # Add 10 sample items
                for i in range(1, 11):
                    item_id = f"sample_{sheet_type.lower()}_{i}"
                    name = f"Sample {sheet_type} item {i}"
                    code = f"CODE{sheet_type}{i}"
                    material_cost = 500
                    labor_cost = 300
                    total_cost = material_cost + labor_cost
                    
                    cursor.execute(
                        f"INSERT INTO {table_name} (internal_id, code, name, material_cost, labor_cost, total_cost, unit) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (item_id, code, name, material_cost, labor_cost, total_cost, 'EA')
                    )
                
                conn.commit()
                logging.info(f"Added 10 sample items to {table_name}")

    def store_processing_session(self, session_id, data):
        """Store processing session data"""
        self.processing_sessions[session_id] = { 
            'data': data, 
            'created_at': datetime.now() 
        }

    def find_best_match(self, name, code, sheet_type):
        """Find best matching item from database using fuzzy matching - specific to sheet type
        Now uses both name and code for better matching of duplicate items"""
        if not name or pd.isna(name): 
            return None
            
        # Get the appropriate table for this sheet type
        table_name = self.sheet_types[sheet_type]['table_name']
            
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            all_items = conn.execute(f"SELECT * FROM {table_name}").fetchall()
            
        if not all_items: 
            return None
        
        # For non-INT sheets, remove item prefix from search name if present
        original_name = str(name).strip()
        sanitized_search = original_name.lower()
        
        # Special debug for important info
        print(f"  🔍 Searching for item: '{original_name}' with code '{code}' in {sheet_type} sheet")
        
        # Handle item prefix for non-interior sheets (e.g., "Item B2" -> "B2")
        if sheet_type != 'INT' and sanitized_search.startswith('item '):
            sanitized_search = sanitized_search.replace('item ', '', 1)
            print(f"  Removed 'Item' prefix: '{original_name}' -> '{sanitized_search}'")
        
        sanitized_code = str(code).lower().strip() if code and not pd.isna(code) else ""
        best_match = None
        best_similarity = 0
        
        # Special handling for hyphen-only names in non-INT sheet types
        is_hyphen_only = sanitized_search == '-'
        
        # EXACT MATCHING FIRST - multiple approaches
        
        # 1. Exact code+name match
        for item_row in all_items:
            item_dict = dict(item_row)
            item_code = str(item_dict['code']).lower().strip()
            item_name = str(item_dict['name']).lower().strip()
            
            # If we have an exact code and name match, return immediately
            if item_code == sanitized_code and item_name == sanitized_search:
                print(f"  ✅ EXACT MATCH found (code+name): {item_code} - {item_name}")
                return {'item': item_dict, 'similarity': 100}
        
        # 2. Exact code match (if code exists)
        if sanitized_code and sanitized_code != "nan" and sanitized_code != "-":
            for item_row in all_items:
                item_dict = dict(item_row)
                item_code = str(item_dict['code']).lower().strip()
                
                if item_code == sanitized_code:
                    print(f"  ✅ EXACT CODE MATCH found: {item_code}")
                    return {'item': item_dict, 'similarity': 95}
        
        # 3. Exact name match
        for item_row in all_items:
            item_dict = dict(item_row)
            item_name = str(item_dict['name']).lower().strip()
            
            if item_name == sanitized_search:
                print(f"  ✅ EXACT NAME MATCH found: {item_name}")
                return {'item': item_dict, 'similarity': 90}
        
        # FUZZY MATCHING (only if exact matching fails)
        
        # Special handling for hyphen-only names in non-INT sheet types
        if is_hyphen_only and sheet_type != 'INT':
            for item_row in all_items:
                item_dict = dict(item_row)
                item_code = str(item_dict['code']).lower().strip()
                
                # For hyphen-only names, prioritize code matching heavily
                if sanitized_code and (item_code.startswith(sanitized_code) or sanitized_code.startswith(item_code)):
                    # Partial code match for hyphen-only items
                    code_similarity = 75
                    if code_similarity > best_similarity:
                        best_similarity = code_similarity
                        best_match = {'item': item_dict, 'similarity': code_similarity}
        
        # Code match with similarity boost
        if sanitized_code:
            for item_row in all_items:
                item_dict = dict(item_row)
                item_code = str(item_dict['code']).lower().strip()
                item_name = str(item_dict['name']).lower().strip()
                
                # Partial code match with boost
                if (item_code.startswith(sanitized_code) or sanitized_code.startswith(item_code)):
                    name_similarity = fuzz.ratio(sanitized_search, item_name)
                    # Boost similarity for code match
                    adjusted_similarity = min(100, name_similarity + 20)
                    
                    if adjusted_similarity > best_similarity:
                        best_similarity = adjusted_similarity
                        best_match = {'item': item_dict, 'similarity': adjusted_similarity}
        
        # Finally, fuzzy name matching
        for item_row in all_items:
            item_dict = dict(item_row)
            item_name = str(item_dict['name']).lower().strip()
            similarity = fuzz.ratio(sanitized_search, item_name)
            
            if similarity > best_similarity:
                best_similarity = similarity
                best_match = {'item': item_dict, 'similarity': similarity}
                
        return best_match

    def determine_sheet_type(self, sheet_name):
        """Determine the sheet type based on its name"""
        sheet_name_lower = sheet_name.lower()
        
        for sheet_type, config in self.sheet_types.items():
            if config['pattern'] and config['pattern'].lower() in sheet_name_lower:
                return sheet_type
                
        # Default to the DEFAULT type if no match
        return 'DEFAULT'

    def _safe_write_to_cell(self, worksheet, row_num, col_num, value):
        """Write value to Excel cell with error handling"""
        if row_num is None or col_num is None or row_num < 1 or col_num < 1:
            print(f"❌ Invalid cell coordinates: row={row_num}, col={col_num}")
            return False
            
        try:
            # Debug cell info
            print(f"  🔍 Writing to cell ({row_num}, {col_num}): value={value}, type={type(value)}")
            
            cell = worksheet.cell(row=int(row_num), column=int(col_num))
            
            # Handle merged cells
            if hasattr(cell, 'coordinate'):
                for merged_range in worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Write to the top-left cell of merged range
                        print(f"  🔄 Cell is part of merged range: {merged_range}")
                        main_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        main_cell.value = value
                        return True
            
            # Write to regular cell
            cell.value = value
            
            # Ensure proper number formatting for numeric values
            if isinstance(value, (int, float)):
                cell.number_format = '0.00'
            
            return True
            
        except Exception as e:
            print(f"❌ Error writing to cell ({row_num}, {col_num}): {e}")
            return False

    def setup_routes(self):
        @self.app.route('/api/process-boq', methods=['POST'])
        def process_boq_route():
            if 'file' not in request.files: 
                return jsonify({'success': False, 'error': 'No file uploaded'})
                
            file = request.files['file']
            filepath = os.path.join(self.upload_folder, secure_filename(file.filename))
            file.save(filepath)
            
            try:
                excel_file = pd.ExcelFile(filepath)
                session_data = {'sheets': {}, 'original_filepath': filepath}
                
                sheets_to_process = [s for s in excel_file.sheet_names if "sum" not in s.lower()]
                
                for sheet_name in sheets_to_process:
                    # Determine sheet type
                    sheet_type = self.determine_sheet_type(sheet_name)
                    sheet_config = self.sheet_types[sheet_type]
                    
                    print(f"\nProcessing sheet: {sheet_name} as type {sheet_type}")
                    
                    # Use fixed header row from configuration
                    header_row = sheet_config['header_row']
                    fixed_columns = sheet_config['columns']
                    
                    # Read with fixed header row
                    df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row)
                    
                    # Simple processing
                    processed_items = []
                    total_rows = len(df)
                    matched_count = 0
                    
                    # Process each row in the sheet
                    for idx, row in df.iterrows():
                        try:
                            # Get name and code from fixed columns
                            name_col = fixed_columns['name'] - 1  # Convert to 0-based
                            code_col = fixed_columns['code'] - 1
                            
                            if name_col >= len(row):
                                continue
                                
                            # Get the basic name and code
                            name = str(row.iloc[name_col]).strip()
                            code = str(row.iloc[code_col]).strip() if code_col < len(row) else ""
                            
                            # Special handling for non-Interior sheets with dash (-) and description
                            # Format in these sheets is typically:
                            # Column C (name): dash (-) 
                            # Column D (next column): actual item description
                            if sheet_type in ["EE", "AC", "FP"] and (name == "-" or name.strip() == "-"):
                                # For EE, AC, FP sheets, *always* check the next column for description
                                # This is specifically for the format we see in the sample where dash is in col C and description in col D
                                next_col_idx = name_col + 1
                                if next_col_idx < len(row):
                                    description = str(row.iloc[next_col_idx]).strip()
                                    if description and description != "nan" and description != "-":
                                        print(f"  Found description in next column: '{description}' in {sheet_type} sheet")
                                        name = description  # Use the description as the name
                                        
                                # If no description found or next column doesn't exist, we'll use the code as fallback
                                if name == "-" and code and code.strip():
                                    print(f"  Using code '{code}' for dash item with no description in {sheet_type} sheet")
                                    name = code
                            
                            # Clean up name for comparison
                            clean_name = name.strip()
                            
                            # Skip empty, header, or total rows
                            if (not clean_name or 
                                clean_name.lower() in ['nan', 'none', ''] or 
                                any(keyword in clean_name.lower() for keyword in ['total', 'รวม', 'system', 'ระบบ'])):
                                continue
                            
                            # For non-interior sheets, handle item prefix
                            if sheet_type != "INT" and clean_name.lower().startswith("item "):
                                original_name = clean_name
                                clean_name = clean_name[5:].strip()  # Remove "Item " prefix
                                print(f"  Processed non-interior item: '{original_name}' -> '{clean_name}'")
                                
                            # Special handling for "-" in non-interior sheets
                            if clean_name == "-" and sheet_type != "INT":
                                # Use code to help with matching when name is just "-"
                                if code and code.strip():
                                    print(f"  Found hyphen-only name with code '{code}' in {sheet_type} sheet")
                                    # We'll proceed with matching using code
                            
                            # Find match in database using improved matching with code+name
                            # Use the cleaned name (without "Item " prefix) for non-interior sheets
                            search_name = clean_name if sheet_type != "INT" and clean_name.lower() != name.lower() else name
                            match = self.find_best_match(search_name, code, sheet_type)
                            
                            if match and match['similarity'] >= 50:
                                processed_items.append({
                                    'original_row_index': idx,
                                    'row_code': code,
                                    'row_name': name,
                                    'match': match
                                })
                                matched_count += 1
                                print(f"  Match: '{name[:40]}...' -> {match['similarity']:.0f}% similarity")
                        except Exception as e:
                            print(f"Error processing row {idx}: {e}")
                            continue
                    
                    print(f"Sheet {sheet_name}: {matched_count}/{total_rows} items matched")
                    
                    # Store sheet info
                    session_data['sheets'][sheet_name] = {
                        'sheet_type': sheet_type,
                        'header_row_num': header_row,
                        'processed_matches': {item['original_row_index']: item['match'] for item in processed_items},
                        'row_details': {item['original_row_index']: {'code': item['row_code'], 'name': item['row_name']} for item in processed_items},
                        'total_rows': total_rows,
                        'matched_count': matched_count
                    }
                
                # Store session
                session_id = str(uuid.uuid4())
                self.store_processing_session(session_id, session_data)
                
                # Calculate summary
                total_items = sum(sheet['total_rows'] for sheet in session_data['sheets'].values())
                total_matches = sum(sheet['matched_count'] for sheet in session_data['sheets'].values())
                
                return jsonify({
                    'success': True, 
                    'session_id': session_id,
                    'summary': {
                        'total_items': total_items,
                        'matched_items': total_matches,
                        'match_rate': (total_matches / total_items * 100) if total_items > 0 else 0
                    }
                })
                
            except Exception as e:
                logging.error(f"Error processing BOQ file: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})

        @self.app.route('/api/generate-final-boq', methods=['POST'])
        def generate_final_boq_route():
            data = request.get_json()
            session_id = data.get('session_id')
            if not session_id or session_id not in self.processing_sessions: 
                return jsonify({'success': False, 'error': 'Invalid session'})
            
            session_data = self.processing_sessions[session_id]['data']
            original_filepath = session_data['original_filepath']
            markup_options = data.get('markup_options', [100, 130, 150, 50, 30])
            
            try:
                # Generate unique output filename
                filename = f"final_boq_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                output_filepath = os.path.join(self.output_folder, filename)
                shutil.copy(original_filepath, output_filepath)

                # Open workbooks
                workbook = openpyxl.load_workbook(output_filepath)
                data_workbook = openpyxl.load_workbook(original_filepath, data_only=True)

                # Initialize counters
                items_processed = 0
                items_failed = 0
                items_with_zero_cost = 0
                items_with_zero_qty = 0

                # Process each sheet
                for sheet_name, sheet_info in session_data['sheets'].items():
                    if sheet_name not in workbook.sheetnames: 
                        continue
                        
                    print(f"\n📋 Processing sheet: {sheet_name}")
                    worksheet = workbook[sheet_name]
                    data_worksheet = data_workbook[sheet_name]
                    
                    # Get sheet info
                    sheet_type = sheet_info['sheet_type']
                    header_row_num = sheet_info['header_row_num']
                    
                    # Get column mappings for this sheet type
                    fixed_columns = self.sheet_types[sheet_type]['columns']
                    print(f"Using fixed columns for {sheet_type}: {fixed_columns}")
                    
                    # Add markup headers
                    header_row_excel = header_row_num + 1
                    start_markup_col = worksheet.max_column + 1
                    for i, p in enumerate(markup_options):
                        self._safe_write_to_cell(worksheet, header_row_excel, start_markup_col + i, f'Markup {p}%')

                    # Process matches
                    match_count = len(sheet_info['processed_matches'])
                    print(f"Processing {match_count} matches for sheet {sheet_name}")
                    
                    # Track section totals and also direct รวมรายการที่ entries
                    current_section = None
                    section_rows = {}  # Map section name to total details
                    
                    # Scan the worksheet for all total row formats
                    sum_rows = {}  # Map total row numbers to their section IDs
                    for row_idx in range(1, worksheet.max_row + 1):
                        # Check BOTH code and name columns for total rows
                        code_cell_value = worksheet.cell(row=row_idx, column=fixed_columns['code']).value
                        name_cell_value = worksheet.cell(row=row_idx, column=fixed_columns['name']).value
                        
                        # Convert to text for comparison
                        code_text = str(code_cell_value).strip() if code_cell_value else ""
                        name_text = str(name_cell_value).strip() if name_cell_value else ""
                        
                        # Use either column's text (prioritize code column for Interior sheets)
                        cell_text = code_text if ("INT" in sheet_type and code_text) else name_text
                        if not cell_text:
                            continue
                        
                        # Check for all possible total row formats
                        section_id = None
                        
                        # Format 1: "รวมรายการที่ N"
                        if 'รวมรายการที่' in cell_text:
                            try:
                                section_num = re.search(r'รวมรายการที่\s*(\d+)', cell_text)
                                if section_num:
                                    section_id = f"SECTION {section_num.group(1)}"
                                    print(f"Found total row 'รวมรายการที่ {section_num.group(1)}' at row {row_idx}")
                            except:
                                print(f"Could not extract section number from '{cell_text}'")
                        
                        # Format 2: Simple "รวมรายการ" (without section number)
                        elif cell_text == 'รวมรายการ' or cell_text.lower() == 'รวมรายการ':
                            section_id = "MAIN_SECTION"
                            print(f"Found main total row 'รวมรายการ' at row {row_idx}")
                            
                        # Also check for other common patterns in AC sheet
                        elif cell_text.lower() == 'electrical woks':
                            section_id = "ELECTRICAL WOKS"
                            print(f"Found electrical work header at row {row_idx}")
                        
                        # Format 3: "Total" or "TOTAL" (case insensitive) - for Interior sheets
                        # AGGRESSIVE DETECTION FOR INTERIOR SHEETS
                        elif cell_text.lower() == 'total' or (("INT" in sheet_type or "Int" in sheet_name) and 
                                                           (cell_text.lower() == 'total' or cell_text.upper() == 'TOTAL')):
                            # For Interior sheets, we want to be more aggressive in detecting total rows
                            section_id = "MAIN_SECTION"
                            # Find the section name by looking at previous rows
                            for prev_row in range(row_idx-1, max(1, row_idx-10), -1):
                                prev_cell = worksheet.cell(row=prev_row, column=fixed_columns['name']).value
                                if prev_cell and str(prev_cell).strip() and not str(prev_cell).strip().lower().startswith('total'):
                                    section_title = str(prev_cell).strip()
                                    section_id = section_title
                                    print(f"  ↑ Section title for this total: '{section_title}'")
                                    break
                            
                            print(f"Found Interior simple 'Total' row at row {row_idx}, assigned to section '{section_id}'")
                            
                        # Format 4: "Total <section name>" - improved to handle various formats
                        elif cell_text.lower().startswith('total'):
                            section_name = cell_text.lower().replace('total', '').strip()
                            section_id = section_name if section_name else "MAIN_SECTION"
                            print(f"Found 'Total {section_name}' row at row {row_idx}")
                            
                        # If we identified this as a total row, store it
                        if section_id:
                            sum_rows[row_idx] = section_id
                            # Initialize section totals
                            section_rows[section_id] = {
                                'material': 0,
                                'labor': 0, 
                                'total': 0,
                                'rows': [],
                                'total_row': row_idx
                            }
                    
                    # First pass - collect and calculate totals
                    all_items_data = []
                    
                    for original_row_index, match_info in sheet_info['processed_matches'].items():
                        if match_info['similarity'] < 50:
                            continue
                            
                        # Calculate target Excel row (1-based)
                        # The original_row_index is already a 0-based index from the dataframe after header
                        target_row_excel = header_row_excel + 1 + int(original_row_index)
                        
                        master_item = match_info['item']
                        
                        # Debug match details
                        print(f"\n🔄 Processing match: '{master_item.get('name', '')[:40]}...'")
                        print(f"  - Similarity: {match_info['similarity']}%")
                        
                        # Get quantity from data worksheet
                        quantity = 0
                        qty_col = fixed_columns.get('quantity')
                        if qty_col:
                            qty_cell = data_worksheet.cell(row=target_row_excel, column=qty_col)
                            try:
                                quantity = float(qty_cell.value or 0)
                                print(f"  - Quantity: {quantity}")
                            except (ValueError, TypeError) as e:
                                print(f"  ⚠️ Quantity conversion error: {e}")
                                quantity = 0
                        
                        if quantity == 0:
                            items_with_zero_qty += 1
                            print(f"  ⚠️ Zero quantity detected!")
                        
                        # Get costs from master data
                        mat_cost = float(master_item.get('material_cost') or 0)
                        lab_cost = float(master_item.get('labor_cost') or 0)
                        total_cost = mat_cost + lab_cost
                        
                        print(f"  - Material cost: {mat_cost}")
                        print(f"  - Labor cost: {lab_cost}")
                        print(f"  - Total unit cost: {total_cost}")
                        
                        if total_cost == 0:
                            items_with_zero_cost += 1
                            print(f"  ⚠️ Zero cost detected!")
                        
                        # Calculate final values
                        material_total = mat_cost * quantity  # Material cost is multiplied by quantity
                        labor_total = lab_cost  # Labor cost is NOT multiplied by quantity
                        total_with_qty = material_total + labor_total  # Total is sum of material*qty + labor
                        
                        # Find section title by looking at rows above
                        section_title = self._find_section_title(worksheet, target_row_excel, fixed_columns['name'])
                        
                        # Find which section this item belongs to
                        section_id = None
                        
                        # Special handling for AC sheets - put all items in a single "MAIN_SECTION"
                        if "AC" in sheet_name:
                            section_id = "MAIN_SECTION"
                            
                        # For other sheets, try to find the appropriate section
                        else:
                            # First, try the รวมรายการที่ approach - find the next total row
                            closest_sum_row = float('inf')
                            for sum_row, section_key in sum_rows.items():
                                if target_row_excel < sum_row and sum_row < closest_sum_row:
                                    closest_sum_row = sum_row
                                    section_id = section_key
                            
                            # If we couldn't find a section by the sum row, use the title-based method
                            if not section_id and section_title:
                                section_id = section_title
                                
                            # Default to MAIN_SECTION for any items without a proper section
                            if not section_id:
                                section_id = "MAIN_SECTION"
                            
                        print(f"  Item assigned to section: '{section_id}'")
                        
                        # Store item data for writing
                        item_data = {
                            'row': target_row_excel,
                            'material_cost': mat_cost,
                            'labor_cost': lab_cost, 
                            'total_cost': total_cost,
                            'quantity': quantity,
                            'material_total': material_total,
                            'labor_total': labor_total,
                            'total_with_qty': total_with_qty,
                            'section': section_id
                        }
                        all_items_data.append(item_data)
                        
                        # Update section totals
                        if section_id:
                            if section_id not in section_rows:
                                # Find the total row if it's not a predefined sum row
                                total_row = None
                                if section_id not in [s for s in sum_rows.values()]:
                                    # Try to find a corresponding total row
                                    total_row = self._find_total_row(worksheet, section_id, fixed_columns['name'])
                                
                                section_rows[section_id] = {
                                    'material': 0,
                                    'labor': 0, 
                                    'total': 0,
                                    'rows': [],
                                    'total_row': total_row
                                }
                            
                            # For section totals, properly accumulate the values
                            section_rows[section_id]['material'] += material_total
                            section_rows[section_id]['labor'] += labor_total
                            section_rows[section_id]['total'] += total_with_qty
                            section_rows[section_id]['rows'].append(target_row_excel)
                    
                    # Second pass - write values and totals
                    for item_data in all_items_data:
                        target_row_excel = item_data['row']
                        mat_cost = item_data['material_cost']
                        lab_cost = item_data['labor_cost']
                        total_with_qty = item_data['total_with_qty']
                        
                        # Write costs to Excel
                        success_count = 0
                        
                        # Material cost
                        mat_col = fixed_columns.get('material_cost')
                        if mat_col:
                            print(f"  - Writing material cost {mat_cost} to cell ({target_row_excel}, {mat_col})")
                            if self._safe_write_to_cell(worksheet, target_row_excel, mat_col, mat_cost):
                                success_count += 1
                        
                        # Labor cost
                        lab_col = fixed_columns.get('labor_cost')
                        if lab_col:
                            print(f"  - Writing labor cost {lab_cost} to cell ({target_row_excel}, {lab_col})")
                            if self._safe_write_to_cell(worksheet, target_row_excel, lab_col, lab_cost):
                                success_count += 1
                        
                        # Total cost
                        total_col = fixed_columns.get('total_cost')
                        if total_col:
                            print(f"  - Writing total cost {total_with_qty} to cell ({target_row_excel}, {total_col})")
                            if self._safe_write_to_cell(worksheet, target_row_excel, total_col, total_with_qty):
                                success_count += 1
                        
                        # Find and write to (บาท) column - typically the next column after total
                        # Look for column header containing "บาท" in header row
                        baht_col = None
                        for col_idx in range(1, worksheet.max_column + 1):
                            header_cell = worksheet.cell(row=header_row_excel, column=col_idx)
                            if header_cell.value and 'บาท' in str(header_cell.value):
                                baht_col = col_idx
                                break
                        
                        # If found, write the same value as total_with_qty
                        if baht_col:
                            print(f"  - Writing baht value {total_with_qty} to cell ({target_row_excel}, {baht_col})")
                            self._safe_write_to_cell(worksheet, target_row_excel, baht_col, total_with_qty)
                        
                        # Write markup values
                        for i, p in enumerate(markup_options):
                            rate = self.markup_rates.get(p, 1.0)
                            # Use the same formula as total_with_qty but with markup
                            markup_val = round((item_data['material_total'] + item_data['labor_total']) * (1 + rate), 2)
                            col_num = start_markup_col + i
                            print(f"  - Writing markup {p}% value {markup_val} to cell ({target_row_excel}, {col_num})")
                            self._safe_write_to_cell(worksheet, target_row_excel, col_num, markup_val)
                        
                        if success_count > 0:
                            items_processed += 1
                        else:
                            items_failed += 1
                            print(f"  ❌ Failed to write any values for row {target_row_excel}")
                    
                    # Third pass - write section totals
                    for section_id, totals in section_rows.items():
                        # Use the pre-identified total row if available
                        total_row = totals.get('total_row')
                        
                        # If we don't have a total row yet, try to find it
                        if not total_row:
                            total_row = self._find_total_row(worksheet, section_id, fixed_columns['name'])
                            
                        if total_row:
                            print(f"\n📊 Writing section total for '{section_id}' at row {total_row}")
                            print(f"  - Total material: {totals['material']}")
                            print(f"  - Total labor: {totals['labor']}")
                            print(f"  - Total sum: {totals['total']}")
                            print(f"  - Item count: {len(totals['rows'])}")
                            
                            # Write the total values
                            mat_col = fixed_columns.get('material_cost')
                            lab_col = fixed_columns.get('labor_cost')
                            total_col = fixed_columns.get('total_cost')
                            
                            if mat_col:
                                self._safe_write_to_cell(worksheet, total_row, mat_col, totals['material'])
                                
                            if lab_col:
                                self._safe_write_to_cell(worksheet, total_row, lab_col, totals['labor'])
                                
                            if total_col:
                                self._safe_write_to_cell(worksheet, total_row, total_col, totals['total'])
                            
                            # Find and write to (บาท) column for section total
                            baht_col = None
                            for col_idx in range(1, worksheet.max_column + 1):
                                header_cell = worksheet.cell(row=header_row_excel, column=col_idx)
                                if header_cell.value and 'บาท' in str(header_cell.value):
                                    baht_col = col_idx
                                    break
                            
                            # If found, write the same total value
                            if baht_col:
                                self._safe_write_to_cell(worksheet, total_row, baht_col, totals['total'])
                                
                            # Write markup totals
                            for i, p in enumerate(markup_options):
                                rate = self.markup_rates.get(p, 1.0)
                                # Use the same formula for section totals
                                markup_total = round((totals['material'] + totals['labor']) * (1 + rate), 2)
                                col_num = start_markup_col + i
                                self._safe_write_to_cell(worksheet, total_row, col_num, markup_total)
                    
                    # ADDITIONAL PASS FOR INTERIOR SHEETS - Handle any missed "Total" rows
                    if "INT" in sheet_type or "Int" in sheet_name:
                        print("\n🔍 Extra pass for Interior sheet total rows...")
                        # Scan for all rows with just "Total" text - check BOTH code and name columns
                        for row_idx in range(1, worksheet.max_row + 1):
                            # Check the code column (column B in Interior sheets) - THIS IS LIKELY WHERE "Total" IS
                            code_cell_value = worksheet.cell(row=row_idx, column=fixed_columns['code']).value
                            name_cell_value = worksheet.cell(row=row_idx, column=fixed_columns['name']).value
                            
                            # Convert to text for comparison
                            code_text = str(code_cell_value).strip() if code_cell_value else ""
                            name_text = str(name_cell_value).strip() if name_cell_value else ""
                            
                            # Check both columns for "Total" text
                            is_total_row = False
                            if code_text.lower() == 'total' or code_text == 'Total' or code_text == 'TOTAL':
                                print(f"  🎯 Found Interior 'Total' in CODE column at row {row_idx}")
                                is_total_row = True
                            elif name_text.lower() == 'total' or name_text == 'Total' or name_text == 'TOTAL':
                                print(f"  🎯 Found Interior 'Total' in NAME column at row {row_idx}")
                                is_total_row = True
                                
                            if is_total_row:
                                print(f"  🛠️ Processing Interior 'Total' row at {row_idx}")
                                
                                # Calculate all totals above this row
                                material_sum = 0
                                labor_sum = 0
                                total_sum = 0
                                
                                # Find start row (first non-empty cell above)
                                start_row = 1
                                for i in range(row_idx-1, 1, -1):
                                    cell = worksheet.cell(row=i, column=fixed_columns['name']).value
                                    if cell and str(cell).strip() and not str(cell).strip().lower().startswith('total'):
                                        start_row = i
                                        break
                                
                                # Calculate total for all items between start_row and row_idx
                                for i in range(start_row, row_idx):
                                    # Only include rows with costs
                                    mat_cell = worksheet.cell(row=i, column=fixed_columns['material_cost']).value
                                    lab_cell = worksheet.cell(row=i, column=fixed_columns['labor_cost']).value
                                    
                                    try:
                                        mat_val = float(mat_cell) if mat_cell and pd.notna(mat_cell) else 0
                                        lab_val = float(lab_cell) if lab_cell and pd.notna(lab_cell) else 0
                                        
                                        if mat_val > 0 or lab_val > 0:
                                            material_sum += mat_val
                                            labor_sum += lab_val
                                            total_sum += (mat_val + lab_val)
                                    except (ValueError, TypeError):
                                        continue
                                
                                # Only write if we have values
                                if material_sum > 0 or labor_sum > 0 or total_sum > 0:
                                    print(f"  📊 Writing calculated totals to row {row_idx}:")
                                    print(f"    - Material: {material_sum}")
                                    print(f"    - Labor: {labor_sum}")
                                    print(f"    - Total: {total_sum}")
                                    
                                    # Write values to cells
                                    self._safe_write_to_cell(worksheet, row_idx, fixed_columns['material_cost'], material_sum)
                                    self._safe_write_to_cell(worksheet, row_idx, fixed_columns['labor_cost'], labor_sum)
                                    self._safe_write_to_cell(worksheet, row_idx, fixed_columns['total_cost'], total_sum)
                                    
                                    # Find and write to (บาท) column
                                    baht_col = None
                                    for col_idx in range(1, worksheet.max_column + 1):
                                        header_cell = worksheet.cell(row=header_row_excel, column=col_idx)
                                        if header_cell.value and 'บาท' in str(header_cell.value):
                                            baht_col = col_idx
                                            break
                                    
                                    if baht_col:
                                        self._safe_write_to_cell(worksheet, row_idx, baht_col, total_sum)
                                        
                                    # Write markup values
                                    for i, p in enumerate(markup_options):
                                        rate = self.markup_rates.get(p, 1.0)
                                        markup_val = round(total_sum * (1 + rate), 2)
                                        col_num = start_markup_col + i
                                        self._safe_write_to_cell(worksheet, row_idx, col_num, markup_val)

                # Save the workbook
                print(f"\n💾 Saving workbook to {output_filepath}")
                workbook.save(output_filepath)
                workbook.close()
                data_workbook.close()
                
                # Cleanup
                if os.path.exists(original_filepath): 
                    os.remove(original_filepath)
                if session_id in self.processing_sessions: 
                    del self.processing_sessions[session_id]

                print(f"\n✅ Processing complete:")
                print(f"  - Items processed successfully: {items_processed}")
                print(f"  - Items failed: {items_failed}")
                print(f"  - Items with zero cost: {items_with_zero_cost}")
                print(f"  - Items with zero quantity: {items_with_zero_qty}")
                
                return jsonify({
                    'success': True, 
                    'filename': filename,
                    'items_processed': items_processed,
                    'items_failed': items_failed,
                    'debug_info': {
                        'items_with_zero_cost': items_with_zero_cost,
                        'items_with_zero_qty': items_with_zero_qty
                    }
                })
                
            except Exception as e:
                logging.error(f"Error generating final BOQ: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})

        @self.app.route('/api/download/<filename>')
        def download_file(filename):
            filepath = os.path.join(self.output_folder, filename)
            if os.path.exists(filepath):
                return send_file(filepath, as_attachment=True)
            return jsonify({'error': 'File not found'}), 404

    def _find_section_title(self, worksheet, row_index, name_col):
        """Find the section title for a given row by looking upward"""
        # Keep track of the most recent total row and section header we've seen
        last_total_row = None
        
        # Try to find a section header between the current row and the previous total row
        for i in range(row_index-1, 0, -1):
            cell_value = worksheet.cell(row=i, column=name_col).value
            if not cell_value:
                continue
                
            cell_text = str(cell_value).strip()
            
            # Stop if we encounter a total row - this marks the boundary of the current section
            if ('Total' in cell_text or 'รวม' in cell_text or 'รวมรายการที่' in cell_text):
                last_total_row = i
                break
            
            # Check if this is a section header (not an item with a hyphen)
            if cell_text and '-' not in cell_text and len(cell_text) > 2:
                # If all caps or clear header format, it's likely a section title
                is_header = (cell_text.isupper() or 
                           cell_text.lower() not in ['panelboard', 'conduit', 'conductor', 'lighting fixture'])
                
                # Check if this row has no numeric cells (likely a header)
                has_numbers = False
                for j in range(1, worksheet.max_column + 1):
                    try:
                        val = worksheet.cell(row=i, column=j).value
                        if isinstance(val, (int, float)) and val > 0:
                            has_numbers = True
                            break
                    except:
                        pass
                
                if not has_numbers and is_header:
                    print(f"  Found section title: '{cell_text}' at row {i}")
                    return cell_text
        
        # If we didn't find a section header, but found a total row, look above it for a major header
        if last_total_row:
            # Try to find a major section header before the last total row
            for i in range(last_total_row-1, max(1, last_total_row-20), -1):
                cell_value = worksheet.cell(row=i, column=name_col).value
                if not cell_value:
                    continue
                    
                cell_text = str(cell_value).strip()
                
                # Look for all-caps headers or known major section types
                if (cell_text.isupper() or 
                    cell_text in ['PANELBOARD', 'CONDUIT AND RACEWAY', 'CONDUCTOR', 'LIGHTING FIXTURE', 
                                'RECEPTACLE AND SWITCH']):
                    print(f"  Found major section title: '{cell_text}' at row {i}")
                    return cell_text
        
        # If all else fails, find the first non-item header above
        for i in range(row_index-1, max(1, row_index-30), -1):
            cell_value = worksheet.cell(row=i, column=name_col).value
            if cell_value and '-' not in str(cell_value) and len(str(cell_value).strip()) > 3:
                cell_text = str(cell_value).strip()
                print(f"  Fallback section title: '{cell_text}' at row {i}")
                return cell_text
        
        print(f"  ⚠️ No section title found for row {row_index}")
        return "UNDEFINED SECTION"

    def _find_total_row(self, worksheet, section_title, name_col):
        """Find the total row for a given section"""
        # For accurate debugging
        print(f"\n🔍 Looking for total row for section: '{section_title}'")
        
        # First check for any simple "รวมรายการ" (without the section number)
        # This is often used as the final total in AC and other sheets
        for i in range(1, worksheet.max_row + 1):
            # Check both code and name columns for Interior sheets
            code_col = name_col - 1  # Code column is typically 1 column before name
            code_cell_value = worksheet.cell(row=i, column=code_col).value if code_col > 0 else None
            name_cell_value = worksheet.cell(row=i, column=name_col).value
            
            # Convert to text for comparison
            code_text = str(code_cell_value).strip() if code_cell_value else ""
            name_text = str(name_cell_value).strip() if name_cell_value else ""
            
            # Check both columns, prioritizing the code column for "Total" text
            cell_text = code_text if "total" in code_text.lower() else name_text
            if not cell_text:
                continue
            
            # Exact match for "รวมรายการ" (final total)
            if cell_text == 'รวมรายการ':
                print(f"  Found main total row: '{cell_text}' at row {i}")
                return i
                
            # Check for the specific format "รวมรายการที่ N" where N is a number
            if 'รวมรายการที่' in cell_text:
                print(f"  Found total row candidate: '{cell_text}' at row {i}")
                # This is definitely a total row
                return i
        
        # Fall back to standard total row patterns
        for i in range(1, worksheet.max_row + 1):
            # Check both code and name columns
            code_col = name_col - 1  # Code column is typically 1 column before name
            code_cell_value = worksheet.cell(row=i, column=code_col).value if code_col > 0 else None
            name_cell_value = worksheet.cell(row=i, column=name_col).value
            
            # Convert to text for comparison
            code_text = str(code_cell_value).strip() if code_cell_value else ""
            name_text = str(name_cell_value).strip() if name_cell_value else ""
            
            # Check both columns, prioritizing the code column for "Total" text
            if "total" in code_text.lower():
                cell_text = code_text
                print(f"  Found 'Total' in CODE column at row {i}")
            else:
                cell_text = name_text
                
            if cell_text:
                
                # Check for standard total row format (improved with case insensitivity)
                if ('total' in cell_text.lower() or 'รวม' in cell_text) and section_title.lower() in cell_text.lower():
                    print(f"  Found standard total row: '{cell_text}' at row {i}")
                    return i
                
                # Check for "ร่วมรายการ" format used in non-interior sheets
                if ('ร่วมรายการ' in cell_text) and section_title in cell_text:
                    print(f"  Found 'ร่วมรายการ' row: '{cell_text}' at row {i}")
                    return i
                
                # Alternative check - "Total" followed by section name
                if ('total' in cell_text.lower() or 'รวม' in cell_text or 'ร่วมรายการ' in cell_text):
                    # Check if this is just a plain "Total" (common in Interior sheets)
                    if cell_text.strip().lower() == 'total' or code_text.strip().lower() == 'total':
                        print(f"  Found plain 'Total' row at row {i}")
                        return i
                        
                    # Check if this section matches our title
                    section_from_cell = self._extract_section_from_total(cell_text)
                    if section_from_cell and section_from_cell.lower() in section_title.lower():
                        print(f"  Found total row by section name match: '{cell_text}' at row {i}")
                        return i
        
        print(f"  ❌ No total row found for section: '{section_title}'")
        return None
        
    def _extract_section_from_total(self, total_text):
        """Extract section name from a total row text"""
        # Try to extract section name from "Total <section_name>" (case insensitive)
        if 'total' in total_text.lower():
            return total_text.lower().replace('total', '').strip()
        if 'รวม' in total_text:
            return total_text.replace('รวม', '').strip()
        if 'ร่วมรายการ' in total_text:
            return total_text.replace('ร่วมรายการ', '').strip()
        # Handle "รวมรายการที่ N" format - this doesn't contain a section name
        if 'รวมรายการที่' in total_text:
            # Extract the section number
            try:
                section_num = re.search(r'รวมรายการที่\s*(\d+)', total_text)
                if section_num:
                    # Return "SECTION N" as a placeholder
                    return f"SECTION {section_num.group(1)}"
            except:
                pass
            # If regex fails, just return the whole text for debugging
            return total_text
        return None

    def run(self, host='localhost', port=5000, debug=True):
        logging.info(f"Hard-coded BOQ Processing Server starting on http://{host}:{port}")
        self.app.run(host=host, port=port, debug=debug)

if __name__ == '__main__':
    processor = HardCodedBOQProcessor()
    processor.run()