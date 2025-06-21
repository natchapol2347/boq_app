"""
Debug script to analyze column structure and determine exact column positions
"""
import openpyxl
import os
import pandas as pd

def debug_excel_structure(filepath):
    """Analyze Excel structure precisely to identify column mappings"""
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        return
    
    print(f"Analyzing Excel file: {filepath}")
    
    # Use openpyxl to inspect the file with coordinates
    wb = openpyxl.load_workbook(filepath)
    
    for sheet_name in wb.sheetnames:
        if "sum" in sheet_name.lower():
            print(f"Skipping summary sheet: {sheet_name}")
            continue
            
        sheet = wb[sheet_name]
        print(f"\n=== Sheet: {sheet_name} ===")
        print(f"Dimensions: {sheet.dimensions}")
        print(f"Max rows: {sheet.max_row}, Max columns: {sheet.max_column}")
        
        # Look for header row (rows 1-15)
        header_row = None
        header_markers = ['ลำดับ', 'code', 'รายการ', 'จำนวน', 'หน่วย', 'ราคา', 'วัสดุ', 'แรง']
        
        for row_idx in range(1, min(20, sheet.max_row + 1)):
            row_values = [str(sheet.cell(row=row_idx, column=col_idx).value or '').lower() 
                         for col_idx in range(1, sheet.max_column + 1)]
            
            matches = sum(1 for marker in header_markers 
                         if any(marker in val for val in row_values if val))
            
            if matches >= 3:
                header_row = row_idx
                print(f"Header found at row {row_idx}")
                
                # Print exact cell coordinates with content
                print("Cell coordinates:")
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        print(f"  {col_letter}{row_idx}: '{cell.value}'")
                
                break
        
        if not header_row:
            print("No header row found")
            continue
        
        # Map columns exactly
        column_mapping = {}
        cost_columns = []
        
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = str(sheet.cell(row=header_row, column=col_idx).value or '').lower()
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            
            if 'code' in cell_value:
                column_mapping['code'] = col_letter
            elif 'รายการ' in cell_value:
                column_mapping['name'] = col_letter
            elif 'จำนวน' in cell_value:
                column_mapping['quantity'] = col_letter
            elif 'หน่วย' in cell_value:
                column_mapping['unit'] = col_letter
            elif any(term in cell_value for term in ['วัสดุ', 'material']):
                column_mapping['material_cost'] = col_letter
                cost_columns.append(col_letter)
            elif any(term in cell_value for term in ['แรงงาน', 'แรง', 'labor', 'labour']):
                column_mapping['labor_cost'] = col_letter
                cost_columns.append(col_letter)
            elif any(term in cell_value for term in ['รวม', 'total']):
                column_mapping['total_cost'] = col_letter
                cost_columns.append(col_letter)
        
        print(f"Column mapping: {column_mapping}")
        
        # Check for cost columns that might be empty/missing
        if not cost_columns:
            print("No cost columns found! Looking for cost columns by position...")
            
            # Try by position (typically columns F, G, H for Thai BOQ)
            for col_letter in ['F', 'G', 'H']:
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                cell_value = sheet.cell(row=header_row, column=col_idx).value
                print(f"  Position {col_letter}: '{cell_value}'")
            
            # Look at the header row structure
            header_structure = []
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row, column=col_idx).value
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                header_structure.append(f"{col_letter}: '{cell_value}'")
            
            print("Header row structure:")
            print("  " + ", ".join(header_structure))
            
            # Look at surrounding rows to see if header is multi-row
            for offset in [-1, 1]:
                surrounding_row = header_row + offset
                if 1 <= surrounding_row <= sheet.max_row:
                    surrounding_values = []
                    for col_idx in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=surrounding_row, column=col_idx).value
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        if cell_value:
                            surrounding_values.append(f"{col_letter}: '{cell_value}'")
                    
                    if surrounding_values:
                        print(f"Row {surrounding_row} values:")
                        print("  " + ", ".join(surrounding_values))
        
        # Check sample data rows
        print("\nSample data rows:")
        for row_idx in range(header_row + 1, min(header_row + 5, sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                if cell_value:
                    row_data.append(f"{col_letter}: '{cell_value}'")
            
            print(f"Row {row_idx}: {', '.join(row_data)}")
    
    # Also check with pandas to get a different view
    print("\nPandas analysis:")
    for sheet_name in pd.ExcelFile(filepath).sheet_names:
        if "sum" in sheet_name.lower():
            continue
            
        print(f"\nSheet: {sheet_name}")
        # Read without headers first
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        
        # Find header row
        header_row = None
        for i in range(min(20, len(df))):
            row_str = df.iloc[i].astype(str).str.lower()
            matches = sum(1 for marker in header_markers 
                         if any(marker in val for val in row_str if 'nan' not in val))
            if matches >= 3:
                header_row = i
                print(f"Header at row {i}: {df.iloc[i].tolist()}")
                break
        
        # Read with proper header
        if header_row is not None:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row)
            print(f"Columns after reading with header at row {header_row}:")
            print(df.columns.tolist())
            
            # Check for numeric columns that might be costs
            numeric_cols = []
            for col in df.columns:
                try:
                    numeric_data = pd.to_numeric(df[col], errors='coerce')
                    non_na_count = numeric_data.notna().sum()
                    if non_na_count > 0:
                        numeric_cols.append(f"{col} ({non_na_count} numeric values)")
                except:
                    pass
            
            if numeric_cols:
                print("Potential numeric columns that might contain costs:")
                for col in numeric_cols:
                    print(f"  {col}")

if __name__ == "__main__":
    filepath = "uploads/Blank BOQ AIS ASP Zeer รังสิต-1.xlsx"
    debug_excel_structure(filepath)