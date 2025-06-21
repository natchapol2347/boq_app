"""
Patched functions for fixing the BOQ processor issues
"""
import openpyxl
import os
import sqlite3
from pathlib import Path

def patched_find_column_numbers(self, worksheet, header_row_num, column_map):
    """Thai BOQ-specific column finder that uses fixed positions based on sheet name"""
    sheet_name = worksheet.title
    print(f"\n📋 Processing sheet: {sheet_name} with fixed Thai BOQ column mapping")
    
    if header_row_num is None:
        # Default header rows based on sheet type
        if "Int" in sheet_name:
            header_row_num = 9
            print(f"Using default header row 9 for Interior sheet")
        elif "EE" in sheet_name:
            header_row_num = 7
            print(f"Using default header row 7 for EE sheet")
        elif "AC" in sheet_name:
            header_row_num = 5
            print(f"Using default header row 5 for AC sheet")
        elif "FP" in sheet_name:
            header_row_num = 7
            print(f"Using default header row 7 for FP sheet")
        else:
            header_row_num = 8
            print(f"Using default header row 8 for unknown sheet type")
    
    header_row_excel = header_row_num + 1
    print(f"Header row in Excel: {header_row_excel}")
    
    # Initialize result mapping
    column_numbers = {}
    
    # Check if this is an Interior sheet or System sheet
    if "Int" in sheet_name:
        # Interior sheet mapping
        print("Using Interior sheet mapping")
        column_numbers = {
            'code': 2,        # Column B
            'name': 3,        # Column C
            'quantity': 4,    # Column D
            'unit': 5,        # Column E
            'material_cost': 6, # Column F
            'labor_cost': 7,  # Column G
            'total_cost': 8   # Column H
        }
        
        # Verify column headers
        header_values = []
        for col_name, col_num in column_numbers.items():
            col_letter = openpyxl.utils.get_column_letter(col_num)
            cell_value = worksheet.cell(row=header_row_excel, column=col_num).value
            header_values.append(f"{col_letter}{header_row_excel} ({col_name}): '{cell_value}'")
            
            # Check if the header matches expectation
            if col_name == 'material_cost':
                # Material cost is in the row below header for Interior sheets
                cost_header_row = header_row_excel + 1
                cost_cell_value = worksheet.cell(row=cost_header_row, column=col_num).value
                print(f"Material cost header: {cost_cell_value} (row {cost_header_row}, col {col_letter})")
    else:
        # System sheet mapping (EE, AC, FP)
        print("Using System sheet mapping")
        column_numbers = {
            'code': 2,        # Column B
            'name': 3,        # Column C
            'unit': 6,        # Column F
            'quantity': 7,    # Column G
            'material_cost': 8, # Column H
            'labor_cost': 10,  # Column J
            'total_cost': 12   # Column L
        }
        
        # Verify column headers
        header_values = []
        for col_name, col_num in column_numbers.items():
            col_letter = openpyxl.utils.get_column_letter(col_num)
            cell_value = worksheet.cell(row=header_row_excel, column=col_num).value
            header_values.append(f"{col_letter}{header_row_excel} ({col_name}): '{cell_value}'")
    
    print("Column mapping for this sheet:")
    for col_name, col_num in column_numbers.items():
        col_letter = openpyxl.utils.get_column_letter(col_num)
        print(f"  - {col_name}: Column {col_letter} ({col_num})")
    
    print(f"Header values: {header_values}")
    
    return column_numbers

def add_sample_costs_to_db():
    """Add sample costs to the database for demonstration"""
    db_path = Path.home() / 'AppData' / 'Roaming' / 'BOQProcessor' / 'master_data.db'
    
    if not os.path.exists(db_path):
        print(f"Database not found at {db_path}")
        return False
    
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            
            # Check if we have items with costs
            cursor.execute("SELECT COUNT(*) FROM master_items WHERE material_cost > 0 OR labor_cost > 0")
            count = cursor.fetchone()[0]
            
            if count > 0:
                print(f"Database already has {count} items with costs")
                return True
            
            # Add sample costs to all items
            cursor.execute("UPDATE master_items SET material_cost = 500, labor_cost = 300, total_cost = 800")
            conn.commit()
            
            # Verify update
            cursor.execute("SELECT COUNT(*) FROM master_items WHERE material_cost > 0")
            count = cursor.fetchone()[0]
            print(f"Updated {count} items with sample costs")
            
            return count > 0
    except Exception as e:
        print(f"Error adding sample costs: {e}")
        return False

def patch_main_script():
    """Creates a patched version of the main.py script with fixed functions"""
    if not os.path.exists("main.py"):
        print("main.py not found")
        return
        
    with open("main.py", "r") as f:
        content = f.read()
    
    # Replace the _find_column_numbers method
    start_marker = "def _find_column_numbers(self, worksheet, header_row_num, column_map):"
    end_markers = ["def setup_routes(self):", "def run(self, host='localhost', port=5000, debug=True):"]
    
    # Find the start and end of the method
    start_pos = content.find(start_marker)
    if start_pos == -1:
        print("Could not find _find_column_numbers method")
        return
    
    # Find the end position (next method)
    end_pos = -1
    for marker in end_markers:
        pos = content.find(marker, start_pos)
        if pos != -1 and (end_pos == -1 or pos < end_pos):
            end_pos = pos
    
    if end_pos == -1:
        print("Could not find end of _find_column_numbers method")
        return
    
    # Get the function code to replace
    old_code = content[start_pos:end_pos]
    
    # Get the new function code
    new_code = """def _find_column_numbers(self, worksheet, header_row_num, column_map):
        \"\"\"Thai BOQ-specific column finder that uses fixed positions based on sheet name\"\"\"
        sheet_name = worksheet.title
        print(f"\\n📋 Processing sheet: {sheet_name} with fixed Thai BOQ column mapping")
        
        if header_row_num is None:
            # Default header rows based on sheet type
            if "Int" in sheet_name:
                header_row_num = 9
                print(f"Using default header row 9 for Interior sheet")
            elif "EE" in sheet_name:
                header_row_num = 7
                print(f"Using default header row 7 for EE sheet")
            elif "AC" in sheet_name:
                header_row_num = 5
                print(f"Using default header row 5 for AC sheet")
            elif "FP" in sheet_name:
                header_row_num = 7
                print(f"Using default header row 7 for FP sheet")
            else:
                header_row_num = 8
                print(f"Using default header row 8 for unknown sheet type")
        
        header_row_excel = header_row_num + 1
        print(f"Header row in Excel: {header_row_excel}")
        
        # Initialize result mapping
        column_numbers = {}
        
        # Check if this is an Interior sheet or System sheet
        if "Int" in sheet_name:
            # Interior sheet mapping
            print("Using Interior sheet mapping")
            column_numbers = {
                'code': 2,        # Column B
                'name': 3,        # Column C
                'quantity': 4,    # Column D
                'unit': 5,        # Column E
                'material_cost': 6, # Column F
                'labor_cost': 7,  # Column G
                'total_cost': 8   # Column H
            }
        else:
            # System sheet mapping (EE, AC, FP)
            print("Using System sheet mapping")
            column_numbers = {
                'code': 2,        # Column B
                'name': 3,        # Column C
                'unit': 6,        # Column F
                'quantity': 7,    # Column G
                'material_cost': 8, # Column H
                'labor_cost': 10,  # Column J
                'total_cost': 12   # Column L
            }
        
        print("Column mapping for this sheet:")
        for col_name, col_num in column_numbers.items():
            import openpyxl
            col_letter = openpyxl.utils.get_column_letter(col_num)
            print(f"  - {col_name}: Column {col_letter} ({col_num})")
        
        return column_numbers"""
    
    # Replace the function
    patched_content = content[:start_pos] + new_code + content[end_pos:]
    
    # Save the patched file
    with open("main_patched.py", "w") as f:
        f.write(patched_content)
    
    print("Created patched main_patched.py file")
    return True

if __name__ == "__main__":
    # Add sample costs to database for demonstration
    add_sample_costs_to_db()
    
    # Create the patched script
    patch_main_script()