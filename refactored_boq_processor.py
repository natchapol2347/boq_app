#!/usr/bin/env python3
"""
Refactored BOQ Processor - Main application class that orchestrates all sheet processors.
CLEANED VERSION: Removed summary logic, moved total writing to processors.
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import os
import uuid
from datetime import datetime
import logging
from werkzeug.utils import secure_filename
from pathlib import Path
import sqlite3
import shutil
import openpyxl
import sys
from typing import Dict, List, Any, Optional

# Import the specialized processors
from interior_sheet_processor import InteriorSheetProcessor
from electrical_sheet_processor import ElectricalSheetProcessor
from ac_sheet_processor import ACSheetProcessor
from fp_sheet_processor import FPSheetProcessor
from config_manager import ConfigManager
from models.config_models import (
    ProcessorType,
    ConfigUpdateRequest,
    ConfigInquiryResponse,
    ConfigUpdateResponse
)

logging.basicConfig(level=logging.DEBUG)

class RefactoredBOQProcessor:
    """Main BOQ processor that orchestrates all sheet-specific processors"""
    
    def __init__(self):
        self.app = Flask(__name__)
        CORS(self.app)
        
        # Setup repo root directories (no more AppData!)
        if getattr(sys, 'frozen', False):
            # Running as packaged executable
            self.app_root = Path(sys.executable).parent
        else:
            # Running as Python script
            self.app_root = Path(__file__).parent.absolute()
        
        # Use environment variables if set by launcher, otherwise use repo paths
        self.data_dir = Path(os.getenv('BOQ_DATA_DIR', self.app_root / 'data'))
        self.config_dir = Path(os.getenv('BOQ_CONFIG_DIR', self.app_root / 'config'))
        self.upload_folder = Path(os.getenv('BOQ_UPLOADS_DIR', self.app_root / 'uploads'))
        self.output_folder = Path(os.getenv('BOQ_OUTPUT_DIR', self.app_root / 'output'))
        
        # Create directories if they don't exist
        for directory in [self.data_dir, self.config_dir, self.upload_folder, self.output_folder]:
            directory.mkdir(exist_ok=True)
        
        self.db_path = str(self.data_dir / 'master_data.db')
        
        # Session management
        self.processing_sessions = {}
        
        # Folder setup
        self.master_data_folder = 'master_data'
        self.upload_folder = 'uploads'
        self.output_folder = 'output'
        for folder in [self.master_data_folder, self.upload_folder, self.output_folder]:
            os.makedirs(folder, exist_ok=True)
        
        # Markup rates
        self.markup_rates = {30: 0.30, 50: 0.50, 100: 1.00, 130: 1.30, 150: 1.50}
        
        # Configuration manager
        self.config_manager = ConfigManager()
        
        # Initialize sheet processors with configuration
        configs = self.config_manager.get_all_configs()
        self.sheet_processors = [
            InteriorSheetProcessor(self.db_path, self.markup_rates, configs.interior),
            ElectricalSheetProcessor(self.db_path, self.markup_rates, configs.electrical),
            ACSheetProcessor(self.db_path, self.markup_rates, configs.ac),
            FPSheetProcessor(self.db_path, self.markup_rates, configs.fp)
        ]
        
        # Initialize database and sync master data
        self._init_database()
        self._sync_master_data()
        
        # Setup Flask routes
        self.setup_routes()
    
    def _init_database(self):
        """Initialize database with all required tables"""
        logging.info(f"Initializing database at {self.db_path}")
        
        with sqlite3.connect(self.db_path) as conn:
            # Create tables for each processor
            for processor in self.sheet_processors:
                processor.create_table(conn)
            
            # Log table creation
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in cursor.fetchall()]
            logging.info(f"Database tables created: {tables}")
    
    def _sync_master_data(self):
        """Sync master data from Excel files to database"""
        master_excel_path = os.path.join(self.master_data_folder, 'master.xlsx')
        
        if not os.path.exists(master_excel_path):
            logging.error(f"Master Excel file not found at {master_excel_path}")
            return
        
        logging.info(f"Synchronizing master data from {master_excel_path}")
        
        try:
            excel_file = pd.ExcelFile(master_excel_path)
            sheet_names = excel_file.sheet_names
            logging.info(f"Found {len(sheet_names)} sheets: {sheet_names}")
            
            for sheet_name in sheet_names:
                # Find the appropriate processor for this sheet
                processor = self._find_processor_for_sheet(sheet_name)
                if not processor:
                    logging.warning(f"No processor found for sheet: {sheet_name}")
                    continue
                
                logging.info(f"Processing sheet {sheet_name} with {processor.__class__.__name__}")
                
                # Read and process the sheet
                df = pd.read_excel(master_excel_path, sheet_name=sheet_name, header=processor.header_row)
                processed_df = processor.process_master_sheet(df)
                
                if not processed_df.empty:
                    processor.sync_to_database(processed_df)
                else:
                    logging.warning(f"No data processed for sheet: {sheet_name}")
            
     
        except Exception as e:
            logging.error(f"Error synchronizing master data: {e}", exc_info=True)
    
    def _find_processor_for_sheet(self, sheet_name: str):
        """Find the appropriate processor for a given sheet name"""
        for processor in self.sheet_processors:
            if processor.matches_sheet(sheet_name):
                return processor
        return None
    
    
    
    def store_processing_session(self, session_id: str, data: Dict[str, Any]):
        """Store processing session data"""
        self.processing_sessions[session_id] = {
            'data': data,
            'created_at': datetime.now()
        }
    
    def _reload_sheet_processors(self):
        """Reload sheet processors with updated configuration"""
        try:
            configs = self.config_manager.get_all_configs()
            self.sheet_processors = [
                InteriorSheetProcessor(self.db_path, self.markup_rates, configs.interior),
                ElectricalSheetProcessor(self.db_path, self.markup_rates, configs.electrical),
                ACSheetProcessor(self.db_path, self.markup_rates, configs.ac),
                FPSheetProcessor(self.db_path, self.markup_rates, configs.fp)
            ]
            logging.info("Sheet processors reloaded with updated configuration")
        except Exception as e:
            logging.error(f"Error reloading sheet processors: {e}", exc_info=True)
    
    def setup_routes(self):
        """Setup Flask routes"""
        
        @self.app.route('/api/process-boq', methods=['POST'])
        def process_boq_route():
            """Process uploaded BOQ file and store matches + section data"""
            if 'file' not in request.files:
                return jsonify({'success': False, 'error': 'No file uploaded'})
            
            file = request.files['file']
            filepath = os.path.join(self.upload_folder, secure_filename(file.filename))
            file.save(filepath)
            
            try:
                excel_file = pd.ExcelFile(filepath)
                session_data = {'sheets': {}, 'original_filepath': filepath}
                
                # Get all sheets to process
                sheets_to_process = excel_file.sheet_names
                
                total_items = 0
                total_matches = 0
                
                for sheet_name in sheets_to_process:
                    # Find appropriate processor
                    processor = self._find_processor_for_sheet(sheet_name)
                    if not processor:
                        logging.info(f"No processor found for sheet: {sheet_name} - skipping")
                        continue
                    
                    logging.info(f"Processing BOQ sheet: {sheet_name} with {processor.__class__.__name__}")
                    
                    # Read sheet data
                    df = pd.read_excel(filepath, sheet_name=sheet_name, header=processor.header_row)
                    
                    # Process the sheet (find matches)
                    processed_items = processor.process_boq_sheet(df)
                    
                    # Pre-calculate section boundaries and store them
                    try:
                        # Read the Excel sheet for section analysis
                        temp_workbook = openpyxl.load_workbook(filepath, data_only=False)
                        temp_worksheet = temp_workbook[sheet_name]
                        
                        # Find section structure only (no cost calculation)
                        sections = processor.find_section_structure(temp_worksheet, temp_worksheet.max_row)
                        temp_workbook.close()
                        
                        logging.info(f"Pre-calculated {len(sections)} sections for {sheet_name}")
                        
                    except Exception as e:
                        logging.warning(f"Could not pre-calculate sections for {sheet_name}: {e}")
                        sections = {}
                    
                    # Store enhanced sheet data
                    session_data['sheets'][sheet_name] = {
                        'processor_type': processor.__class__.__name__,
                        'header_row': processor.header_row,
                        'processed_matches': {item['original_row_index']: item['match'] for item in processed_items},
                        'row_details': {item['original_row_index']: {'code': item['row_code'], 'name': item['row_name']} for item in processed_items},
                        'sections': sections,  
                        'total_rows': len(df),
                        'matched_count': len(processed_items)
                    }
                    
                    total_items += len(df)
                    total_matches += len(processed_items)
                
                # Store session
                session_id = str(uuid.uuid4())
                self.store_processing_session(session_id, session_data)
                
                return jsonify({
                    'success': True,
                    'session_id': session_id,
                    'summary': {
                        'total_items': total_items,
                        'matched_items': total_matches,
                        'match_rate': (total_matches / total_items * 100) if total_items > 0 else 0,
                        'sheets_processed': len(session_data['sheets'])
                    }
                })
                
            except Exception as e:
                logging.error(f"Error processing BOQ file: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/generate-final-boq', methods=['POST'])
        def generate_final_boq_route():
            """Generate final BOQ with calculated costs"""
            data = request.get_json()
            session_id = data.get('session_id')
            
            if not session_id or session_id not in self.processing_sessions:
                return jsonify({'success': False, 'error': 'Invalid session'})
            
            session_data = self.processing_sessions[session_id]['data']
            original_filepath = session_data['original_filepath']
            markup_options = data.get('markup_options', [30, 50, 100, 130, 150])
            
            try:
                # Generate output filename
                filename = f"final_boq_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                output_filepath = os.path.join(self.output_folder, filename)
                shutil.copy(original_filepath, output_filepath)
                
                # Open workbooks
                workbook = openpyxl.load_workbook(output_filepath)
                data_workbook = openpyxl.load_workbook(original_filepath, data_only=True)
                
                # Process each sheet
                items_processed = 0
                items_failed = 0
                processing_summary = {}
                
                for sheet_name, sheet_info in session_data['sheets'].items():
                    if sheet_name not in workbook.sheetnames:
                        continue
                    
                    # Find the processor for this sheet
                    processor = self._find_processor_for_sheet(sheet_name)
                    if not processor:
                        logging.warning(f"No processor found for sheet: {sheet_name}")
                        continue
                    
                    logging.info(f"Generating costs for sheet: {sheet_name}")
                    
                    # Process the sheet - Let processor handle everything
                    sheet_result = processor.process_final_sheet(
                        worksheet=workbook[sheet_name], 
                        data_worksheet=data_workbook[sheet_name],
                        sheet_info=sheet_info,
                        markup_options=markup_options
                    )
                    
                    items_processed += sheet_result['items_processed']
                    items_failed += sheet_result['items_failed']
                    processing_summary[sheet_name] = sheet_result
                
                # Save workbook
                workbook.save(output_filepath)
                workbook.close()
                data_workbook.close()
                
                # Keep session alive for potential markup application
                # No cleanup here - user might want to apply markup next
                
                logging.info(f"Processing complete: {items_processed} items processed, {items_failed} failed")
                
                return jsonify({
                    'success': True,
                    'filename': filename,
                    'download_url': f'/api/download/{filename}',
                    'items_processed': items_processed,
                    'items_failed': items_failed,
                    'processing_summary': processing_summary
                })
                
            except Exception as e:
                logging.error(f"Error generating final BOQ: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/apply-markup', methods=['POST'])
        def apply_markup_route():
            """Apply markup directly to all values in all sheets (not just display columns)"""
            data = request.get_json()
            session_id = data.get('session_id')
            markup_percent = data.get('markup_percent')
            
            if not session_id or session_id not in self.processing_sessions:
                return jsonify({'success': False, 'error': 'Invalid session'})
            
            if markup_percent is None or not isinstance(markup_percent, (int, float)):
                return jsonify({'success': False, 'error': 'markup_percent must be a valid number'})
            
            session_data = self.processing_sessions[session_id]['data']
            original_filepath = session_data['original_filepath']
            
            try:
                # Extract original filename without extension for better naming
                original_name = os.path.splitext(os.path.basename(original_filepath))[0]
                
                # Generate output filename with markup percentage and original name
                filename = f"{markup_percent}%_{original_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                output_filepath = os.path.join(self.output_folder, filename)
                shutil.copy(original_filepath, output_filepath)
                
                # Open workbooks
                workbook = openpyxl.load_workbook(output_filepath)
                data_workbook = openpyxl.load_workbook(original_filepath, data_only=True)
                
                # Process each sheet with markup application flag
                items_processed = 0
                items_failed = 0
                processing_summary = {}
                
                for sheet_name, sheet_info in session_data['sheets'].items():
                    if sheet_name not in workbook.sheetnames:
                        continue
                    
                    # Find the processor for this sheet
                    processor = self._find_processor_for_sheet(sheet_name)
                    if not processor:
                        logging.warning(f"No processor found for sheet: {sheet_name}")
                        continue
                    
                    logging.info(f"Applying {markup_percent}% markup to sheet: {sheet_name}")
                    
                    # Process the sheet with markup application flag
                    sheet_result = processor.process_final_sheet(
                        worksheet=workbook[sheet_name], 
                        data_worksheet=data_workbook[sheet_name],
                        sheet_info=sheet_info,
                        markup_options=[],  # Empty list since we're applying markup, not displaying it
                        apply_markup_percent=markup_percent  # New parameter
                    )
                    
                    items_processed += sheet_result['items_processed']
                    items_failed += sheet_result['items_failed']
                    processing_summary[sheet_name] = sheet_result
                
                # Save workbook
                workbook.save(output_filepath)
                workbook.close()
                data_workbook.close()
                
                # Keep session alive - user may want to apply different markup rates
                # Session will be cleaned up when app closes or via cleanup-session endpoint
                
                logging.info(f"Markup application complete: {markup_percent}% applied to {items_processed} items, {items_failed} failed")
                
                return jsonify({
                    'success': True,
                    'filename': filename,
                    'download_url': f'/api/download/{filename}',
                    'markup_percent': markup_percent,
                    'items_processed': items_processed,
                    'items_failed': items_failed,
                    'processing_summary': processing_summary
                })
                
            except Exception as e:
                logging.error(f"Error applying markup: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/cleanup-session', methods=['POST'])
        def cleanup_session_route():
            """Cleanup session data and delete associated files"""
            data = request.get_json()
            session_id = data.get('session_id')
            
            if not session_id:
                return jsonify({'success': False, 'error': 'session_id is required'})
            
            files_deleted = []
            errors = []
            
            try:
                # Get session data before deletion
                if session_id in self.processing_sessions:
                    session_data = self.processing_sessions[session_id]['data']
                    original_filepath = session_data.get('original_filepath')
                    
                    # Delete original uploaded file
                    if original_filepath and os.path.exists(original_filepath):
                        try:
                            os.remove(original_filepath)
                            files_deleted.append(original_filepath)
                            logging.info(f"Deleted original file: {original_filepath}")
                        except Exception as e:
                            errors.append(f"Failed to delete {original_filepath}: {e}")
                    
                    # Delete session from memory
                    del self.processing_sessions[session_id]
                    logging.info(f"Cleaned up session: {session_id}")
                else:
                    return jsonify({'success': False, 'error': 'Invalid session_id'})
                
                # Find and delete related output files (optional - files generated for this session)
                # We'll look for files that might be related to this session's original filename
                if original_filepath:
                    original_basename = os.path.splitext(os.path.basename(original_filepath))[0]
                    for output_file in os.listdir(self.output_folder):
                        if original_basename in output_file:
                            output_path = os.path.join(self.output_folder, output_file)
                            try:
                                os.remove(output_path)
                                files_deleted.append(output_path)
                                logging.info(f"Deleted output file: {output_path}")
                            except Exception as e:
                                errors.append(f"Failed to delete {output_path}: {e}")
                
                return jsonify({
                    'success': True,
                    'session_cleaned': True,
                    'files_deleted': len(files_deleted),
                    'deleted_files': files_deleted,
                    'errors': errors
                })
                
            except Exception as e:
                logging.error(f"Error during cleanup: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        @self.app.route('/api/config/inquiry', methods=['GET'])
        def config_inquiry_route():
            """Get current processor configurations"""
            try:
                configs = self.config_manager.get_all_configs()
                summary = self.config_manager.get_config_summary()
                
                return ConfigInquiryResponse(
                    success=True,
                    configs=configs
                ).model_dump()
                
            except Exception as e:
                logging.error(f"Error getting config: {e}", exc_info=True)
                return ConfigInquiryResponse(
                    success=False,
                    error=str(e)
                ).model_dump()
        
        @self.app.route('/api/config/update', methods=['POST'])
        def config_update_route():
            """Update processor configuration"""
            try:
                data = request.get_json()
                
                # Validate request using Pydantic
                update_request = ConfigUpdateRequest(**data)
                
                # Update configuration
                success = self.config_manager.update_config(update_request)
                
                if success:
                    # Reload sheet processors with new configuration
                    self._reload_sheet_processors()
                    
                    return ConfigUpdateResponse(
                        success=True,
                        message="Configuration updated successfully",
                        updated_processor=update_request.processor_name.value
                    ).model_dump()
                else:
                    return ConfigUpdateResponse(
                        success=False,
                        message="Failed to update configuration",
                        error="Update operation failed"
                    ).model_dump()
                    
            except Exception as e:
                logging.error(f"Error updating config: {e}", exc_info=True)
                return ConfigUpdateResponse(
                    success=False,
                    message="Configuration update failed",
                    error=str(e)
                ).model_dump()
        
        @self.app.route('/api/download/<filename>')
        def download_file(filename):
            """Download generated BOQ file"""
            filepath = os.path.join(self.output_folder, filename)
            if os.path.exists(filepath):
                return send_file(filepath, as_attachment=True)
            return jsonify({'error': 'File not found'}), 404
    
    def run(self, host: str = 'localhost', port: int = 5000, debug: bool = True):
        """Run the Flask application"""
        logging.info(f"Refactored BOQ Processing Server starting on http://{host}:{port}")
        self.app.run(host=host, port=port, debug=debug)

if __name__ == '__main__':
    processor = RefactoredBOQProcessor()
    processor.run()