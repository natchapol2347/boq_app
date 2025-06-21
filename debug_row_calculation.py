#!/usr/bin/env python3
# Debug script to trace row calculations

import openpyxl
import os
import pandas as pd
import sqlite3
from pathlib import Path

def debug_row_calculation():
    """Debug row calculations between original Excel and final output"""
    print("Debugging row calculation issues...")
    
    # Sample file to check
    sample_file = "uploads/Blank BOQ AIS ASP Zeer รังสิต-1.xlsx"
    if not os.path.exists(sample_file):
        print(f"Sample file not found at {sample_file}")
        return
    
    # Sheet type configurations
    sheet_types = {
        'INT': {
            'pattern': 'int',
            'header_row': 9,  # 0-based
            'columns': {
                'code': 2,          # Column B
                'name': 3,          # Column C
                'quantity': 4,      # Column D
                'unit': 5,          # Column E
                'material_cost': 6, # Column F
                'labor_cost': 7,    # Column G
                'total_cost': 8     # Column H
            },
            'table_name': 'interior_items'
        },
        'EE': {
            'pattern': 'ee',
            'header_row': 7,  # 0-based
            'columns': {
                'code': 2,          # Column B
                'name': 3,          # Column C
                'unit': 6,          # Column F
                'quantity': 7,      # Column G
                'material_cost': 8, # Column H
                'labor_cost': 10,   # Column J
                'total_cost': 12    # Column L
            },
            'table_name': 'ee_items'
        },
        'AC': {
            'pattern': 'ac',
            'header_row': 5,  # 0-based
            'columns': {
                'code': 2,          # Column B
                'name': 3,          # Column C
                'unit': 6,          # Column F
                'quantity': 7,      # Column G
                'material_cost': 8, # Column H
                'labor_cost': 10,   # Column J
                'total_cost': 12    # Column L
            },
            'table_name': 'ac_items'
        },
        'FP': {
            'pattern': 'fp',
            'header_row': 7,  # 0-based
            'columns': {
                'code': 2,          # Column B
                'name': 3,          # Column C
                'unit': 6,          # Column F
                'quantity': 7,      # Column G
                'material_cost': 8, # Column H
                'labor_cost': 10,   # Column J
                'total_cost': 12    # Column L
            },
            'table_name': 'fp_items'
        },
        'DEFAULT': {
            'pattern': '',
            'header_row': 8,  # 0-based
            'columns': {
                'code': 2,          # Column B
                'name': 3,          # Column C
                'quantity': 4,      # Column D
                'unit': 5,          # Column E
                'material_cost': 6, # Column F
                'labor_cost': 7,    # Column G
                'total_cost': 8     # Column H
            },
            'table_name': 'default_items'
        }
    }
    
    # Open the workbook
    wb = openpyxl.load_workbook(sample_file)
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        if "sum" in sheet_name.lower():
            continue
            
        print(f"\n==== Processing sheet: {sheet_name} ====")
        
        # Determine sheet type
        sheet_type = 'DEFAULT'
        for key, config in sheet_types.items():
            if config['pattern'] and config['pattern'].lower() in sheet_name.lower():
                sheet_type = key
                break
        
        print(f"Sheet type determined: {sheet_type}")
        
        # Get sheet configuration
        config = sheet_types[sheet_type]
        header_row = config['header_row']
        columns = config['columns']
        
        print(f"Header row (0-based): {header_row}")
        print(f"Header row (Excel): {header_row + 1}")
        
        # Check columns in Excel
        sheet = wb[sheet_name]
        
        # Examine what's in the header row
        header_row_excel = header_row + 1
        header_values = []
        for col_name, col_num in columns.items():
            cell_value = sheet.cell(row=header_row_excel, column=col_num).value
            header_values.append(f"{col_name}: '{cell_value}'")
        
        print(f"Header values: {header_values}")
        
        # Now check actual data rows
        print("\nChecking data rows:")
        
        # Read with pandas to get DataFrame indices
        df = pd.read_excel(sample_file, sheet_name=sheet_name, header=header_row)
        print(f"DataFrame shape: {df.shape}")
        
        # Check a few rows to see the matching between pandas index and Excel rows
        for idx in range(min(5, len(df))):
            # Get pandas row
            row = df.iloc[idx]
            
            # Get name from fixed column in pandas
            name_col = columns['name'] - 1  # Convert to 0-based for pandas
            if name_col < len(row):
                name = str(row.iloc[name_col]).strip()
                
                # Calculate Excel row
                excel_row = header_row_excel + 1 + idx
                
                # Get value directly from Excel
                excel_name = sheet.cell(row=excel_row, column=columns['name']).value
                
                # Check materials/labor columns
                mat_col = columns['material_cost']
                lab_col = columns['labor_cost']
                material_value = sheet.cell(row=excel_row, column=mat_col).value
                labor_value = sheet.cell(row=excel_row, column=lab_col).value
                
                print(f"Row {idx} (pandas) -> Row {excel_row} (Excel):")
                print(f"  Name from pandas: '{name}'")
                print(f"  Name from Excel: '{excel_name}'")
                print(f"  Material cell: {material_value}")
                print(f"  Labor cell: {labor_value}")
                
                # Check calculation used in the app
                target_row_calc1 = header_row_excel + 1 + idx
                target_row_calc2 = header_row_excel + 1 + idx - header_row
                
                print(f"  Row calculation 1 (used in app): {target_row_calc1}")
                print(f"  Row calculation 2 (alternative): {target_row_calc2}")
        
        # Now check if any items are matched in the database
        db_path = Path.home() / 'AppData' / 'Roaming' / 'BOQProcessor' / 'master_data.db'
        if os.path.exists(db_path):
            print("\nChecking database matches:")
            try:
                with sqlite3.connect(db_path) as conn:
                    conn.row_factory = sqlite3.Row
                    cursor = conn.cursor()
                    
                    # Check if table exists
                    table_name = config['table_name']
                    cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}'")
                    if cursor.fetchone():
                        # Get sample items from this table
                        cursor.execute(f"SELECT name, material_cost, labor_cost FROM {table_name} LIMIT 5")
                        items = cursor.fetchall()
                        
                        print(f"Sample items from {table_name}:")
                        for item in items:
                            print(f"  - '{item['name']}': Material={item['material_cost']}, Labor={item['labor_cost']}")
                        
                        # Try to match some items from the sheet
                        for idx in range(min(5, len(df))):
                            # Get name from DataFrame
                            row = df.iloc[idx]
                            name_col = columns['name'] - 1
                            if name_col < len(row):
                                name = str(row.iloc[name_col]).strip()
                                
                                if name and name.lower() not in ['nan', 'none', ''] and not any(keyword in name.lower() for keyword in ['total', 'รวม']):
                                    print(f"\nTrying to match: '{name}'")
                                    
                                    # Find fuzzy matches
                                    import fuzzywuzzy.fuzz as fuzz
                                    
                                    cursor.execute(f"SELECT name, material_cost, labor_cost FROM {table_name}")
                                    all_items = cursor.fetchall()
                                    
                                    best_match = None
                                    best_similarity = 0
                                    
                                    for item in all_items:
                                        item_name = item['name']
                                        similarity = fuzz.ratio(name.lower(), item_name.lower())
                                        
                                        if similarity > best_similarity:
                                            best_similarity = similarity
                                            best_match = item
                                    
                                    if best_match and best_similarity >= 50:
                                        print(f"  Match found: '{best_match['name']}' ({best_similarity}%)")
                                        print(f"  Costs: Material={best_match['material_cost']}, Labor={best_match['labor_cost']}")
                                        
                                        # Calculate target Excel row
                                        excel_row = header_row_excel + 1 + idx
                                        
                                        print(f"  Target Excel row: {excel_row}")
                                        print(f"  Material column: {columns['material_cost']}")
                                        print(f"  Labor column: {columns['labor_cost']}")
                                    else:
                                        print(f"  No good match found")
                    else:
                        print(f"Table {table_name} not found in database")
            except Exception as e:
                print(f"Error checking database: {e}")
        else:
            print(f"Database not found at {db_path}")
    
    wb.close()

if __name__ == "__main__":
    debug_row_calculation()