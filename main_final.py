# BOQ Cost Automation Backend - Complete Implementation with simplified item-by-item matching

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import numpy as np
from fuzzywuzzy import fuzz
import os
import uuid
from datetime import datetime
import re
from werkzeug.utils import secure_filename
from pathlib import Path
import sqlite3
import logging
import shutil
import openpyxl

logging.basicConfig(level=logging.INFO)

class BOQProcessor:
    def __init__(self):
        self.app = Flask(__name__)
        CORS(self.app)
        
        # --- Database and Folder Setup ---
        self.data_dir = Path.home() / 'AppData' / 'Roaming' / 'BOQProcessor'
        os.makedirs(self.data_dir, exist_ok=True)
        self.db_path = self.data_dir / 'master_data.db'
        
        # Session management for processing sessions
        self.processing_sessions = {}
        
        self.master_data_folder = 'master_data'
        self.upload_folder = 'uploads'
        self.output_folder = 'output'
        os.makedirs(self.master_data_folder, exist_ok=True)
        os.makedirs(self.upload_folder, exist_ok=True)
        os.makedirs(self.output_folder, exist_ok=True)

        self.markup_rates = {100: 1.00, 130: 1.30, 150: 1.50, 50: 0.50, 30: 0.30}
        
        self._init_db()
        self._sync_default_master_excel()
        self.setup_routes()

    def _init_db(self):
        """Initialize SQLite database with enhanced debugging"""
        logging.info(f"Initializing database at {self.db_path}")
        
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS master_items (
                        internal_id TEXT PRIMARY KEY, 
                        code TEXT, 
                        name TEXT NOT NULL UNIQUE,
                        material_cost REAL DEFAULT 0, 
                        labor_cost REAL DEFAULT 0, 
                        total_cost REAL DEFAULT 0
                    )
                ''')
                conn.commit()
                
                # DEBUG: Check if table exists and has correct structure
                table_info = cursor.execute("PRAGMA table_info(master_items)").fetchall()
                logging.info(f"Database table structure: {table_info}")
                
                # DEBUG: Check if table has any data
                count = cursor.execute("SELECT COUNT(*) FROM master_items").fetchone()[0]
                logging.info(f"Database contains {count} master items")
                
                # DEBUG: Check items with costs
                if count > 0:
                    items_with_costs = cursor.execute(
                        "SELECT COUNT(*) FROM master_items WHERE material_cost > 0 OR labor_cost > 0"
                    ).fetchone()[0]
                    logging.info(f"Items with costs: {items_with_costs} ({items_with_costs/count*100:.1f}%)")
                
                logging.info(f"Database initialized successfully at {self.db_path}")
                
        except Exception as e:
            logging.error(f"Error initializing database: {e}", exc_info=True)
            raise

    def _sync_default_master_excel(self):
        """Sync default master Excel file to database only if needed"""
        default_excel_path = os.path.join(self.master_data_folder, 'master.xlsx')
        
        # First check if we already have data in the database
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM master_items")
            count = cursor.fetchone()[0]
            
            if count > 0:
                logging.info(f"Database already contains {count} items, skipping reload")
                
                # Check if any items have costs
                cursor.execute("SELECT COUNT(*) FROM master_items WHERE material_cost > 0 OR labor_cost > 0")
                items_with_costs = cursor.fetchone()[0]
                logging.info(f"Items with costs: {items_with_costs} ({items_with_costs/count*100 if count > 0 else 0:.1f}%)")
                
                return
        
        # If we get here, the database is empty and needs to be loaded
        if os.path.exists(default_excel_path):
            logging.info(f"Found {default_excel_path}. Synchronizing database...")
            
            # DEBUG: Check Excel file properties
            try:
                excel_file = pd.ExcelFile(default_excel_path)
                sheet_names = excel_file.sheet_names
                logging.info(f"Excel file contains {len(sheet_names)} sheets: {sheet_names}")
                
                # Check if file contains any data
                sample_sheet = pd.read_excel(default_excel_path, sheet_name=sheet_names[0], nrows=5)
                logging.info(f"Sample data from first sheet: {len(sample_sheet)} rows, {len(sample_sheet.columns)} columns")
                logging.info(f"Columns: {list(sample_sheet.columns)}")
                
                # Proceed with synchronization
                result = self.load_data_from_excel_to_db(default_excel_path)
                if result.get('success'):
                    logging.info(f"Successfully synchronized: {result.get('message')}")
                else:
                    logging.error(f"Synchronization failed: {result.get('error')}")
                    
            except Exception as e:
                logging.error(f"Error inspecting Excel file: {e}", exc_info=True)
        else:
            logging.warning(f"Default master Excel file not found at {default_excel_path}")
    def setup_routes(self):
        @self.app.route('/api/process-boq', methods=['POST'])
        def process_boq_route():
            if 'file' not in request.files: 
                return jsonify({'success': False, 'error': 'No file uploaded'})
                
            file = request.files['file']
            filepath = os.path.join(self.upload_folder, secure_filename(file.filename))
            file.save(filepath)
            
            try:
                excel_file = pd.ExcelFile(filepath)
                session_data = {'sheets': {}, 'original_filepath': filepath}
                
                sheets_to_process = [s for s in excel_file.sheet_names if "sum" not in s.lower()]
                
                for sheet_name in sheets_to_process:
                    raw_df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
                    header_row = self.find_header_row(raw_df)
                    df = pd.read_excel(filepath, sheet_name=sheet_name, 
                                     header=header_row if header_row is not None else 0)
                    
                    # Simple processing - no grouping, match each row individually
                    match_df = self._prepare_dataframe_simple(df.copy())
                    
                    processed_matches = {}
                    total_rows = len(match_df)
                    matched_count = 0
                    
                    for _, row in match_df.iterrows():
                        item_name = str(row['name']).strip()
                        if len(item_name) > 2:  # Only match meaningful names
                            match = self.find_best_match(item_name)
                            if match:
                                processed_matches[row['original_row_index']] = match
                                matched_count += 1
                                print(f"  Match: '{item_name[:40]}...' -> {match['similarity']:.0f}% similarity")
                    
                    print(f"Sheet {sheet_name}: {matched_count}/{total_rows} items matched")

                    session_data['sheets'][sheet_name] = {
                        'header_row_num': header_row,
                        'processed_matches': processed_matches,
                        'column_map': self.detect_column_mapping(df),
                        'total_rows': total_rows,
                        'matched_count': matched_count
                    }

                session_id = str(uuid.uuid4())
                self.store_processing_session(session_id, session_data)
                
                # Calculate summary
                total_items = sum(sheet['total_rows'] for sheet in session_data['sheets'].values())
                total_matches = sum(sheet['matched_count'] for sheet in session_data['sheets'].values())
                
                return jsonify({
                    'success': True, 
                    'session_id': session_id,
                    'summary': {
                        'total_items': total_items,
                        'matched_items': total_matches,
                        'match_rate': (total_matches / total_items * 100) if total_items > 0 else 0
                    }
                })
                
            except Exception as e:
                logging.error(f"Error processing BOQ file: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})

        @self.app.route('/api/generate-final-boq', methods=['POST'])
        def generate_final_boq_route():
            data = request.get_json()
            session_id = data.get('session_id')
            if not session_id or session_id not in self.processing_sessions: 
                return jsonify({'success': False, 'error': 'Invalid session'})
            
            session_data = self.processing_sessions[session_id]['data']
            original_filepath = session_data['original_filepath']
            markup_options = data.get('markup_options', [100, 130, 150, 50, 30])
            
            # DEBUG: Check if we have a valid database and it has cost data
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                items_with_costs = conn.execute(
                    "SELECT COUNT(*) as count FROM master_items WHERE material_cost > 0 OR labor_cost > 0"
                ).fetchone()
                total_items = conn.execute("SELECT COUNT(*) as count FROM master_items").fetchone()
                
                print(f"\n🔍 DATABASE CHECK: {items_with_costs['count']} items with costs out of {total_items['count']} total items")
                
                # Sample some items with costs
                sample_items = conn.execute(
                    "SELECT name, material_cost, labor_cost FROM master_items WHERE material_cost > 0 OR labor_cost > 0 LIMIT 5"
                ).fetchall()
                
                print(f"📊 SAMPLE ITEMS WITH COSTS:")
                for item in sample_items:
                    print(f"  - {item['name'][:40]}: Material: {item['material_cost']}, Labor: {item['labor_cost']}")
            
            try:
                filename = f"final_boq_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                output_filepath = os.path.join(self.output_folder, filename)
                shutil.copy(original_filepath, output_filepath)

                workbook = openpyxl.load_workbook(output_filepath)
                data_workbook = openpyxl.load_workbook(original_filepath, data_only=True)

                items_processed = 0
                items_failed = 0
                items_with_zero_cost = 0
                items_with_zero_qty = 0

                for sheet_name, sheet_info in session_data['sheets'].items():
                    if sheet_name not in workbook.sheetnames: 
                        continue
                        
                    print(f"\n📋 Processing sheet: {sheet_name}")
                    worksheet = workbook[sheet_name]
                    data_worksheet = data_workbook[sheet_name]
                    
                    header_row_num = sheet_info['header_row_num']
                    column_map = sheet_info['column_map']
                    
                    # DEBUG: Print column map for this sheet
                    print(f"🔄 Column map for {sheet_name}: {column_map}")
                    
                    # Find actual column numbers in Excel
                    column_numbers = self._find_column_numbers(worksheet, header_row_num, column_map)
                    
                    # DEBUG: Print found column numbers
                    print(f"📍 Column numbers found: {column_numbers}")
                    
                    if not column_numbers:
                        print(f"⚠️ No columns found for {sheet_name}, skipping")
                        continue
                    
                    # DEBUG: Check if cost columns were found
                    if 'material_cost' not in column_numbers:
                        print(f"⚠️ WARNING: Material cost column not found!")
                    if 'labor_cost' not in column_numbers:
                        print(f"⚠️ WARNING: Labor cost column not found!")
                    if 'total_cost' not in column_numbers:
                        print(f"⚠️ WARNING: Total cost column not found!")
                    
                    # Add markup headers
                    header_row_excel = (header_row_num or 0) + 1
                    start_markup_col = worksheet.max_column + 1
                    for i, p in enumerate(markup_options):
                        self._safe_write_to_cell(worksheet, header_row_excel, start_markup_col + i, f'Markup {p}%')

                    # DEBUG: Print match count
                    print(f"🔍 Total matches to process: {len(sheet_info['processed_matches'])}")
                    
                    # Process matches
                    for original_row_index, match_info in sheet_info['processed_matches'].items():
                        if match_info['similarity'] < 50:
                            continue
                            
                        # Calculate target Excel row (1-based)
                        target_row_excel = header_row_excel + 1 + int(original_row_index)
                        
                        master_item = match_info['item']
                        
                        # DEBUG: Print match details
                        print(f"\n🔄 Processing match: '{master_item.get('name', '')[:40]}...'")
                        print(f"  - Similarity: {match_info['similarity']}%")
                        print(f"  - Raw master item data: {master_item}")
                        
                        # Get quantity from data worksheet
                        quantity = 0
                        qty_raw_value = None
                        if 'quantity' in column_numbers:
                            qty_cell = data_worksheet.cell(row=target_row_excel, column=column_numbers['quantity'])
                            qty_raw_value = qty_cell.value
                            try:
                                quantity = float(qty_cell.value or 0)
                                print(f"  - Quantity: {quantity} (raw value: {qty_raw_value}, type: {type(qty_raw_value)})")
                            except (ValueError, TypeError) as e:
                                print(f"  ⚠️ Quantity conversion error: {e}, raw value: {qty_raw_value}")
                                quantity = 0
                        else:
                            print(f"  ⚠️ No quantity column found!")
                        
                        if quantity == 0:
                            items_with_zero_qty += 1
                            print(f"  ⚠️ Zero quantity detected!")
                        
                        # Get costs from master data
                        mat_cost_raw = master_item.get('material_cost')
                        lab_cost_raw = master_item.get('labor_cost')
                        
                        print(f"  - Raw material cost: {mat_cost_raw}, type: {type(mat_cost_raw)}")
                        print(f"  - Raw labor cost: {lab_cost_raw}, type: {type(lab_cost_raw)}")
                        
                        mat_cost = float(mat_cost_raw or 0)
                        lab_cost = float(lab_cost_raw or 0)
                        total_cost = mat_cost + lab_cost
                        
                        print(f"  - Converted material cost: {mat_cost}")
                        print(f"  - Converted labor cost: {lab_cost}")
                        print(f"  - Total unit cost: {total_cost}")
                        
                        if total_cost == 0:
                            items_with_zero_cost += 1
                            print(f"  ⚠️ Zero cost detected!")
                        
                        # DEBUG: Calculate final values to write
                        material_total = mat_cost
                        labor_total = lab_cost
                        total_with_qty = total_cost * quantity
                        
                        print(f"  - Final values to write:")
                        print(f"    * Material cost: {material_total}")
                        print(f"    * Labor cost: {labor_total}")
                        print(f"    * Total cost with quantity: {total_with_qty}")
                        
                        # Write costs to Excel
                        success_count = 0
                        if 'material_cost' in column_numbers:
                            col_num = column_numbers['material_cost']
                            print(f"  - Writing material cost {material_total} to cell ({target_row_excel}, {col_num})")
                            if self._safe_write_to_cell(worksheet, target_row_excel, col_num, material_total):
                                success_count += 1
                                print(f"    ✅ Material cost written successfully")
                            else:
                                print(f"    ❌ Failed to write material cost")
                                
                        if 'labor_cost' in column_numbers:
                            col_num = column_numbers['labor_cost']
                            print(f"  - Writing labor cost {labor_total} to cell ({target_row_excel}, {col_num})")
                            if self._safe_write_to_cell(worksheet, target_row_excel, col_num, labor_total):
                                success_count += 1
                                print(f"    ✅ Labor cost written successfully")
                            else:
                                print(f"    ❌ Failed to write labor cost")
                                
                        if 'total_cost' in column_numbers:
                            col_num = column_numbers['total_cost']
                            print(f"  - Writing total cost {total_with_qty} to cell ({target_row_excel}, {col_num})")
                            if self._safe_write_to_cell(worksheet, target_row_excel, col_num, total_with_qty):
                                success_count += 1
                                print(f"    ✅ Total cost written successfully")
                            else:
                                print(f"    ❌ Failed to write total cost")
                        
                        # Write markup values
                        for i, p in enumerate(markup_options):
                            rate = self.markup_rates.get(p, 1.0)
                            markup_val = round(total_cost * quantity * (1 + rate), 2)
                            col_num = start_markup_col + i
                            print(f"  - Writing markup {p}% value {markup_val} to cell ({target_row_excel}, {col_num})")
                            result = self._safe_write_to_cell(worksheet, target_row_excel, col_num, markup_val)
                            if result:
                                print(f"    ✅ Markup {p}% written successfully")
                            else:
                                print(f"    ❌ Failed to write markup {p}%")
                        
                        if success_count > 0:
                            items_processed += 1
                        else:
                            items_failed += 1
                            print(f"    ❌ Failed to write any values for row {target_row_excel}")

                # Save the workbook
                print(f"\n💾 Saving workbook to {output_filepath}")
                workbook.save(output_filepath)
                workbook.close()
                data_workbook.close()
                
                # Cleanup
                if os.path.exists(original_filepath): 
                    os.remove(original_filepath)
                if session_id in self.processing_sessions: 
                    del self.processing_sessions[session_id]

                print(f"\n✅ Processing complete:")
                print(f"  - Items processed successfully: {items_processed}")
                print(f"  - Items failed: {items_failed}")
                print(f"  - Items with zero cost: {items_with_zero_cost}")
                print(f"  - Items with zero quantity: {items_with_zero_qty}")
                
                return jsonify({
                    'success': True, 
                    'filename': filename,
                    'items_processed': items_processed,
                    'items_failed': items_failed,
                    'debug_info': {
                        'items_with_zero_cost': items_with_zero_cost,
                        'items_with_zero_qty': items_with_zero_qty
                    }
                })
                
            except Exception as e:
                logging.error(f"Error generating final BOQ: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})

        @self.app.route('/api/download/<filename>')
        def download_file(filename):
            filepath = os.path.join(self.output_folder, filename)
            if os.path.exists(filepath):
                return send_file(filepath, as_attachment=True)
            return jsonify({'error': 'File not found'}), 404

    def run(self, host='localhost', port=5000, debug=True):
        logging.info(f"BOQ Processing Server starting on http://{host}:{port}")
        self.app.run(host=host, port=port, debug=debug)

if __name__ == '__main__':
    processor = BOQProcessor()
    processor.run()