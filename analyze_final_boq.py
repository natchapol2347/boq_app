"""
Analyze the latest final BOQ file to verify that our fixes worked
"""
import openpyxl
import os
from openpyxl.utils import get_column_letter

def is_header_row(name):
    """Check if a row name indicates it's a header/category row"""
    if not name:
        return False
        
    name_lower = str(name).lower()
    
    # Common header indicators
    header_indicators = [
        'งาน', 'work', 'system', 'ระบบ', 'total', 'รวม', 'หมวด'
    ]
    
    # Check if any indicators are present but not part of a longer product name
    for indicator in header_indicators:
        if indicator in name_lower and len(name_lower) < 40:
            # Additional check for rows that are clearly headers
            if name_lower.startswith(indicator) or name_lower.endswith(indicator):
                return True
            
            # Check for common section headers
            if any(section in name_lower for section in [
                'ระบบ', 'system', 'หมวดงาน', 'section', 'category',
                'รวมงาน', 'total', 'รวมราคา'
            ]):
                return True
    
    return False

def analyze_boq_file(filepath):
    """Analyze the final BOQ file to verify our fixes worked"""
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        return
    
    print(f"Analyzing BOQ file: {filepath}\n")
    
    wb = openpyxl.load_workbook(filepath)
    
    for sheet_name in wb.sheetnames:
        if "sum" in sheet_name.lower():
            continue
            
        print(f"=== Sheet: {sheet_name} ===")
        sheet = wb[sheet_name]
        
        # Determine expected cost columns based on sheet type
        if "Int" in sheet_name:
            # Interior sheet
            material_col = 6  # F
            labor_col = 7     # G
            total_col = 8     # H
            print("Expected cost columns: F=material, G=labor, H=total")
        else:
            # System sheets
            material_col = 8  # H
            labor_col = 10    # J
            total_col = 12    # L
            print("Expected cost columns: H=material, J=labor, L=total")
        
        # Find header row
        header_row = None
        for row_idx in range(1, min(20, sheet.max_row + 1)):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value == "ลำดับ":
                header_row = row_idx
                print(f"Found header at row {row_idx}")
                break
        
        if not header_row:
            print("Could not find header row")
            continue
        
        # Check code/name columns
        code_col = 2  # B
        name_col = 3  # C
        
        # Analyze cost data
        items_with_costs = 0
        headers_with_costs = 0
        items_without_costs = 0
        total_items = 0
        
        for row_idx in range(header_row + 1, min(header_row + 40, sheet.max_row + 1)):
            code = sheet.cell(row=row_idx, column=code_col).value
            name = sheet.cell(row=row_idx, column=name_col).value
            
            # Skip empty rows
            if not name and not code:
                continue
                
            # If name is empty but there's content in column D, use that
            if not name and "Int" not in sheet_name:
                # For system sheets, check column D for description
                name = sheet.cell(row=row_idx, column=4).value
            
            if not name:
                continue
                
            # Check if this is a header row
            is_header = is_header_row(name)
            
            # Get cost values
            material_value = sheet.cell(row=row_idx, column=material_col).value
            labor_value = sheet.cell(row=row_idx, column=labor_col).value
            total_value = sheet.cell(row=row_idx, column=total_col).value
            
            # Try to convert to numeric
            try:
                material_cost = float(material_value or 0)
            except (ValueError, TypeError):
                material_cost = 0
                
            try:
                labor_cost = float(labor_value or 0)
            except (ValueError, TypeError):
                labor_cost = 0
                
            try:
                total_cost = float(total_value or 0)
            except (ValueError, TypeError):
                total_cost = 0
            
            # Check if there are costs
            has_costs = material_cost > 0 or labor_cost > 0 or total_cost > 0
            
            if is_header:
                if has_costs:
                    headers_with_costs += 1
                    print(f"HEADER with costs: Row {row_idx}: {name}")
                    print(f"  Material: {material_cost}, Labor: {labor_cost}, Total: {total_cost}")
            else:
                total_items += 1
                if has_costs:
                    items_with_costs += 1
                    print(f"Item with costs: Row {row_idx}: {name}")
                    print(f"  Material: {material_cost}, Labor: {labor_cost}, Total: {total_cost}")
                else:
                    items_without_costs += 1
                    print(f"Item WITHOUT costs: Row {row_idx}: {name}")
        
        # Check markup columns
        markup_found = False
        for col_idx in range(material_col, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col_idx).value
            if cell_value and "markup" in str(cell_value).lower():
                markup_found = True
                print(f"Found markup column at {get_column_letter(col_idx)}: {cell_value}")
                
                # Check if markup values are calculated
                markup_values = 0
                for row_idx in range(header_row + 1, min(header_row + 40, sheet.max_row + 1)):
                    markup_value = sheet.cell(row=row_idx, column=col_idx).value
                    try:
                        markup_numeric = float(markup_value or 0)
                        if markup_numeric > 0:
                            markup_values += 1
                    except (ValueError, TypeError):
                        pass
                
                print(f"Markup column contains {markup_values} non-zero values")
        
        # Summary
        print(f"\nSummary for {sheet_name}:")
        print(f"  Total items: {total_items}")
        print(f"  Items with costs: {items_with_costs} ({items_with_costs/total_items*100 if total_items > 0 else 0:.1f}%)")
        print(f"  Items without costs: {items_without_costs}")
        print(f"  Headers with costs: {headers_with_costs}")
        print(f"  Markup columns found: {'Yes' if markup_found else 'No'}")
        print()
    
    wb.close()

if __name__ == "__main__":
    latest_boq = "/Users/a677022/Desktop/woodman/boq_app/output/final_boq_20250619_122910.xlsx"
    analyze_boq_file(latest_boq)